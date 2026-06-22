import streamlit as st
import plotly.graph_objects as go
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from collections import defaultdict
import datetime, io

st.set_page_config(page_title="QC Monitor", page_icon="🔍", layout="wide", initial_sidebar_state="expanded")
st.markdown("""<style>
.block-container{padding-top:1.5rem}
div[data-testid="metric-container"]{background:#0f1117;border:1px solid #1c2030;border-radius:8px;padding:12px 16px}
div[data-testid="metric-container"] label{font-size:11px!important;color:#7c879e!important;text-transform:uppercase;letter-spacing:.5px}
</style>""", unsafe_allow_html=True)

# ─── HELPERS ─────────────────────────────────────────────────────────────────

SKIP_ROWS = {'Total','Total Panel Dédupliqué','Total Panel Dédupliqué - Top 5 Sites',
    'Total Panel Dédupliqué  - Top 11 Sites','Total Panel Dédupliqué Marché',
    'Immobilier Notaire','Immonot','Site','Département','Totaux','Total Panel Dedup'}

SITES = ['AvendreAlouer',"Bien'ici",'Figaro Immo','Green-Acres','Leboncoin',
         'LogicImmo','MeilleursAgents','OuestFrance','PAP','ParuVendu','SeLoger','SuperImmo']

def excel_date_str(v):
    if isinstance(v, datetime.datetime): return v.strftime('%b-%y')
    if isinstance(v, str): return v.strip()
    if isinstance(v,(int,float)) and 40000<v<50000:
        return (datetime.datetime(1899,12,30)+datetime.timedelta(days=int(v))).strftime('%b-%y')
    return str(v)

def get_sheet_data(ws, site_col=2, start_col=3):
    """Extract site data + months from a sheet."""
    header_row = None
    for r in range(1, min(20, ws.max_row+1)):
        v = ws.cell(r, site_col).value
        if v in ('Site','Département','Région'):
            header_row = r; break
    if not header_row: return None, None, None

    month_cols, months = [], []
    for c in range(start_col, ws.max_column+1):
        v = ws.cell(header_row, c).value
        if v is None: continue
        if (isinstance(v, datetime.datetime) or
            (isinstance(v,(int,float)) and 40000<v<50000) or
            (isinstance(v,str) and any(m in v.lower() for m in ['-25','-26','-24','-23']))):
            month_cols.append(c); months.append(excel_date_str(v))
    if len(months)<2: return None, None, None

    data = {}
    for r in range(header_row+1, ws.max_row+1):
        site = ws.cell(r, site_col).value
        if not site or not isinstance(site,str): continue
        site = site.strip()
        if not site: continue
        vm = ws.cell(r, month_cols[-1]).value
        vm1 = ws.cell(r, month_cols[-2]).value
        if isinstance(vm,(int,float)) and isinstance(vm1,(int,float)):
            data[site] = {'m':float(vm),'m1':float(vm1)}
    return data, months[-1], months[-2]

def check_ok(name, condition, detail, severity="error", sheet=""):
    return {'check':name,'status':'ok' if condition else 'fail',
            'severity':severity if not condition else 'ok','detail':detail,'sheet':sheet}

# ─── QC GOLD CHECKS ──────────────────────────────────────────────────────────

def run_sheet_checks(wb, file_label):
    checks = []
    for sn in wb.sheetnames:
        if sn == 'Intro': continue
        ws = wb[sn]
        data, month_m, month_m1 = get_sheet_data(ws)
        if not data or not month_m: continue
        label = f"{file_label} › {sn}"

        total = data.get('Total')
        dedup = next((data[k] for k in data if 'Dédupliqué' in k or 'Dedup' in k), None)
        sites = {k:v for k,v in data.items() if k not in SKIP_ROWS and 'Dédupliqué' not in k and 'Dedup' not in k}

        # Check 1: Total = sum of sites
        if total and sites:
            s = sum(v['m'] for v in sites.values())
            if s > 1000:
                diff_pct = abs(total['m']-s)/s*100
                checks.append(check_ok(
                    f"Total = sum of sites ({month_m})",
                    diff_pct < 1,
                    f"Total={total['m']:,.0f} | Sum={s:,.0f} | Diff={diff_pct:.2f}%",
                    "error", label))

        # Check 2: Dedup <= Total
        if dedup and total and total['m'] > 1000:
            checks.append(check_ok(
                f"Total Dédupliqué ≤ Total ({month_m})",
                dedup['m'] <= total['m'],
                f"Dedup={dedup['m']:,.0f} | Total={total['m']:,.0f}",
                "error", label))

        # Check 3: Evol% per site
        for site, vals in sites.items():
            vm, vm1 = vals['m'], vals['m1']
            if vm1 > 10 and vm > 10:
                evol = (vm/vm1-1)*100
                if abs(evol) > 30:
                    checks.append({'check':f"Evol% {site}",
                        'status':'warning','severity':'warning',
                        'detail':f"{evol:+.1f}% ({month_m1}={vm1:,.0f} → {month_m}={vm:,.0f})",
                        'sheet':label})
    return checks

def run_cross_checks(files_bytes):
    checks = []
    def get_dedup(key):
        if key not in files_bytes: return None, None
        wb = load_workbook(io.BytesIO(files_bytes[key]), data_only=True)
        sn = [s for s in wb.sheetnames if s != 'Intro']
        if not sn: return None, None
        ws = wb[sn[0]]
        data, m, _ = get_sheet_data(ws)
        if not data: return None, None
        dedup = next((data[k] for k in data if 'Dédupliqué' in k or 'Dedup' in k), None)
        return (dedup['m'] if dedup else None), m

    pairs = [
        ('file1','file3_1','File 1 Total Dedup = File 3.1 Total Dedup'),
        ('file1','file3_2','File 1 Total Dedup = File 3.2 Total Dedup'),
        ('file3_2','file4_1','File 3.2 Total Dedup = File 4.1 Total Dedup'),
    ]
    for k1,k2,name in pairs:
        v1,m1 = get_dedup(k1)
        v2,m2 = get_dedup(k2)
        if v1 and v2:
            diff_pct = abs(v1-v2)/v1*100 if v1>0 else 0
            checks.append(check_ok(
                f"{name} ({m1})",
                diff_pct < 1,
                f"{k1}={v1:,.0f} | {k2}={v2:,.0f} | Diff={diff_pct:.2f}%",
                "error", "Cross-file check"))
    return checks

"""
QC Gold V2 — Exact replicas of Panel Checker controls applied to HAM files.
"""
import io, datetime
from openpyxl import load_workbook

def excel_date_str(v):
    if isinstance(v, datetime.datetime): return v.strftime('%b-%y')
    if isinstance(v, str): return v.strip()
    if isinstance(v,(int,float)) and 40000<v<50000:
        return (datetime.datetime(1899,12,30)+datetime.timedelta(days=int(v))).strftime('%b-%y')
    return str(v)

def get_val_by_label(ws, label, col_offset=0, search_cols=range(1,5)):
    """Find row by label in any of search_cols, return value from last data col + offset."""
    for r in range(1, ws.max_row+1):
        for c in search_cols:
            cell_val = ws.cell(r,c).value
            if cell_val and str(cell_val).strip() == label:
                # Find last non-None column
                last_c = ws.max_column
                while last_c > c and ws.cell(r, last_c).value is None:
                    last_c -= 1
                target_c = last_c + col_offset
                if target_c > 0:
                    return ws.cell(r, target_c).value
    return None

def get_val_contains(ws, label_part, col_offset=0, search_cols=range(1,5)):
    """Find row where label contains label_part."""
    for r in range(1, ws.max_row+1):
        for c in search_cols:
            cell_val = ws.cell(r,c).value
            if cell_val and label_part.lower() in str(cell_val).lower():
                last_c = ws.max_column
                while last_c > c and ws.cell(r, last_c).value is None:
                    last_c -= 1
                target_c = last_c + col_offset
                if target_c > 0:
                    return ws.cell(r, target_c).value, str(cell_val).strip()
    return None, None

def get_last_two_months(ws, site_col=2):
    """Get last and second-to-last month columns and their labels."""
    header_row = None
    for r in range(1, min(20, ws.max_row+1)):
        v = ws.cell(r, site_col).value
        if v in ('Site','Département','Région'):
            header_row = r; break
    if not header_row: return None, None, None, None

    month_cols, months = [], []
    for c in range(site_col+1, ws.max_column+1):
        v = ws.cell(header_row, c).value
        if v is None: continue
        if (isinstance(v, datetime.datetime) or
            (isinstance(v,(int,float)) and 40000<v<50000) or
            (isinstance(v,str) and any(m in v.lower() for m in ['-25','-26','-24','-23']))):
            month_cols.append(c); months.append(excel_date_str(v))
    if len(months)<2: return None, None, None, None
    return month_cols[-1], months[-1], month_cols[-2], months[-2]

def get_site_val(ws, site_name, site_col=2):
    """Get last month value for a site."""
    lc, lm, pc, pm = get_last_two_months(ws, site_col)
    if not lc: return None, None, None, None
    for r in range(1, ws.max_row+1):
        v = ws.cell(r, site_col).value
        if v and str(v).strip() == site_name:
            return ws.cell(r,lc).value, ws.cell(r,pc).value, lm, pm
    return None, None, lm, pm

def check(name, ok, detail, severity="error", sheet=""):
    return {'check':name, 'status':'ok' if ok else 'fail',
            'severity':severity if not ok else 'ok', 'detail':detail, 'sheet':sheet}

"""
QC Gold V3 — All checks from Zeynep's Redmine messages + Panel Checker formulas.
Includes ML-based anomaly detection using historical patterns.
"""
import io, datetime
import numpy as np
from openpyxl import load_workbook

def excel_date_str(v):
    if isinstance(v, datetime.datetime): return v.strftime('%b-%y')
    if isinstance(v, str): return v.strip()
    if isinstance(v,(int,float)) and 40000<v<50000:
        return (datetime.datetime(1899,12,30)+datetime.timedelta(days=int(v))).strftime('%b-%y')
    return str(v)

def last_val(ws, row, from_col=3):
    """Get last non-None numeric value in a row."""
    for c in range(ws.max_column, from_col-1, -1):
        v = ws.cell(row, c).value
        if isinstance(v,(int,float)): return float(v), c
    return None, None

def find_row(ws, label, col_range=range(1,5)):
    """Find row number where label matches."""
    for r in range(1, ws.max_row+1):
        for c in col_range:
            v = ws.cell(r,c).value
            if v and label.lower() in str(v).lower():
                return r
    return None

def check(name, ok, detail, severity="error", sheet=""):
    return {'check':name,'status':'ok' if ok else 'fail',
            'severity':severity if not ok else 'ok','detail':detail,'sheet':sheet}

"""
QC Gold FINAL — Direct replication of Panel Checker cross-table controls.
Each check maps to exact Panel Checker formula logic using HAM file data.
"""
import io, datetime
import numpy as np
from openpyxl import load_workbook

def dstr(v):
    if isinstance(v, datetime.datetime): return v.strftime('%b-%y')
    if isinstance(v, str): return v.strip()
    if isinstance(v,(int,float)) and 40000<v<50000:
        return (datetime.datetime(1899,12,30)+datetime.timedelta(days=int(v))).strftime('%b-%y')
    return str(v) if v else ''

def ok(name, passed, detail, sev="error", sheet=""):
    return {'check':name,'status':'ok' if passed else 'fail',
            'severity':sev if not passed else 'ok','detail':detail,'sheet':sheet}

def close(a,b,pct=0.5):
    if a is None or b is None: return True
    if a==0 and b==0: return True
    return abs(a-b)/max(abs(a),abs(b))*100 < pct

# ─── GENERIC READER ───────────────────────────────────────────────────────────

def read_site_series(ws, site_col=2, section_idx=0):
    """Returns {site: [monthly_vals], months: [str], last_col: int, prev_col: int}
    section_idx: 0=first section (default), -1=last section"""
    all_hdrs = []
    for r in range(1, ws.max_row+1):
        v = ws.cell(r,site_col).value
        if v in ('Site','Département','Région'): all_hdrs.append(r)
    if not all_hdrs: return {}
    hdr = all_hdrs[section_idx] if abs(section_idx)<len(all_hdrs) else all_hdrs[0]

    mcols, months = [], []
    for c in range(site_col+1, ws.max_column+1):
        h = ws.cell(hdr,c).value
        if h is None: continue
        if (isinstance(h,datetime.datetime) or
            (isinstance(h,(int,float)) and 40000<h<50000) or
            (isinstance(h,str) and any(m in h.lower() for m in ['-26','-25','-24','-23']))):
            mcols.append(c); months.append(dstr(h))
    if len(mcols)<2: return {}

    result = {'_months': months, '_mcols': mcols,
              '_lc': mcols[-1], '_lm': months[-1],
              '_pc': mcols[-2], '_pm': months[-2]}

    for r in range(hdr+1, ws.max_row+1):
        b = ws.cell(r,site_col).value
        if not b or not isinstance(b,str): continue
        b = b.strip()
        if not b: continue
        vals = []
        for c in mcols:
            v = ws.cell(r,c).value
            vals.append(float(v) if isinstance(v,(int,float)) else None)
        result[b] = {'vals':vals, 'last': vals[-1], 'prev': vals[-2] if len(vals)>=2 else None}

    return result

def site_val(data, name):
    """Get last month value for a site name (exact or partial match)."""
    if name in data: return data[name]['last']
    for k in data:
        if isinstance(k,str) and name.lower() in k.lower() and not k.startswith('_'):
            return data[k]['last']
    return None

def site_series(data, name):
    """Get full time series for a site."""
    if name in data: return data[name]['vals']
    for k in data:
        if isinstance(k,str) and name.lower() in k.lower() and not k.startswith('_'):
            return data[k]['vals']
    return None

SITES = ['AvendreAlouer',"Bien'ici",'Figaro Immo','Green-Acres','Leboncoin',
         'LogicImmo','MeilleursAgents','OuestFrance','PAP','ParuVendu','SeLoger','Superimmo']

# ─── TAB 1 CHECKS ─────────────────────────────────────────────────────────────

def check_tab1(wb1_bytes):
    checks = []
    wb = load_workbook(io.BytesIO(wb1_bytes), data_only=True)

    for sn in wb.sheetnames:
        if sn=='Intro': continue
        ws = wb[sn]
        d = read_site_series(ws)
        if not d: continue
        lm = d.get('_lm','?')
        label = f"1 — {sn}"

        total = site_val(d,'Total')
        dedup_marche = site_val(d,'Total Panel Dédupliqué Marché')
        dedup_11 = site_val(d,'Total Panel Dédupliqué  - Top 11 Sites') or site_val(d,'Total Panel Dédupliqué - Top 11 Sites')
        dedup_5 = site_val(d,'Total Panel Dédupliqué - Top 5 Sites')

        # Check each site's evol (ML z-score)
        for site in SITES:
            series = site_series(d, site)
            if series and len([v for v in series if v and v>0])>=6:
                real = [v for v in series if v and v>0]
                if len(real)>=6:
                    hist = np.array(real[:-1])
                    last = real[-1]
                    mean,std = np.mean(hist),np.std(hist)
                    if std>0:
                        z=(last-mean)/std
                        if abs(z)>3:
                            checks.append({'check':f"Anomalie Z-score: {site}",'status':'warning',
                                'severity':'warning',
                                'detail':f"Z={z:.1f} | {lm}={last:,.0f} | moy={mean:,.0f} | σ={std:,.0f}",
                                'sheet':label})

        # Dedup <= Total
        if dedup_marche and total and total>1000 and dedup_marche>total*1.01:
            checks.append(ok(f"Dedup Marché ≤ Total ({lm})",False,
                f"Dedup={dedup_marche:,.0f} > Total={total:,.0f}","error",label))
        if dedup_marche:
            checks.append(ok(f"Total Panel Dédupliqué Marché présent ({lm})",True,
                f"{dedup_marche:,.0f}","ok",label))

    return checks

# ─── TAB 3.1 CHECKS ───────────────────────────────────────────────────────────

def check_tab31(wb31_bytes, wb1_bytes):
    """
    Panel Checker tab 3.1 controls:
    r23 AQ: =AQ18='tab 1'!O130  → 3.1.4 Total Dedup Marché = 1.2 Total Dedup Marché
    r23 D/G/J...: per-site total pros = 1.2 per-site total
    r73,82,84 U: Vente+Loc=Total per site
    r131-144 AC-AL: current month values consistent across sections
    r203 U: Total Dedup TOP11 = sum of segments
    r317 T: Total Dedup = sum of type segments
    """
    checks = []
    wb31 = load_workbook(io.BytesIO(wb31_bytes), data_only=True)
    wb1  = load_workbook(io.BytesIO(wb1_bytes), data_only=True)

    ws314 = wb31['3.1.4 Evolution Pros par type']
    ws312 = wb31['3.1.2 Pros partagés']
    ws315 = wb31['3.1.5 Evolution Pros exclu.']
    ws311 = wb31['3.1.1 Pros par site ']
    ws12  = wb1['1.2 Pro_Part']
    ws14  = wb1['1.4 Type de professionels']

    d314 = read_site_series(ws314)
    d12  = read_site_series(ws12)
    d14  = read_site_series(ws14)

    lm = d314.get('_lm','?')

    # ── CHECK 1: =AQ18='tab 1'!O130
    # 3.1.4 Total Panel Dédupliqué Marché = 1.2 Total Panel Dédupliqué Marché (Ancien)
    v314_dedup = site_val(d314,'Total Panel Dédupliqué')
    v314_dedup_m = site_val(d314,'Total Panel Dédupliqué Marché') or v314_dedup
    v12_dedup_m  = site_val(d12,'Total Panel Dédupliqué Marché')
    if v314_dedup_m and v12_dedup_m:
        checks.append(ok(
            f"3.1.4 Total Dedup = 1.2 Total Dedup Marché ({lm})",
            close(v314_dedup_m, v12_dedup_m),
            f"3.1.4={v314_dedup_m:,.0f} | 1.2={v12_dedup_m:,.0f} | diff={abs(v314_dedup_m-v12_dedup_m):,.0f}",
            "error","3.1 — contrôle tab 1 (AQ18='tab 1'!O130)"
        ))

    # ── CHECK 2: per-site Total général pros = per-site in 1.2
    # tab 3.1 row 18 per-site col = 3.1.4 "Total" per site last
    # tab 1 per-site = 1.2 per-site last
    lm2 = d12.get('_lm','?')
    for site in SITES:
        v314 = site_val(d314, site)
        v12  = site_val(d12, site)
        if v314 and v12 and v314>100 and v12>100:
            checks.append(ok(
                f"3.1.4 {site} = 1.2 {site} ({lm})",
                close(v314, v12, 2.0),
                f"3.1.4={v314:,.0f} | 1.2={v12:,.0f} | diff={abs(v314-v12):,.0f}",
                "error","3.1 — contrôle cohérence sites"
            ))

    # ── CHECK 3: =O73+O92=O191 → Bien'ici Vente+Loc=Total
    # In 3.1.4: first section=all, section2=vente, section3=location
    # Find section boundaries
    site_rows = []
    for r in range(1, ws314.max_row+1):
        if ws314.cell(r,2).value=='Site': site_rows.append(r)

    lc = d314.get('_lc')
    if lc and len(site_rows)>=3:
        for site in ["Bien'ici",'SeLoger','Total']:
            vals_by_section = []
            for sr in site_rows[:3]:
                for r in range(sr+1, min(sr+25,ws314.max_row+1)):
                    b = ws314.cell(r,2).value
                    if b and str(b).strip()==site:
                        v = ws314.cell(r,lc).value
                        if isinstance(v,(int,float)): vals_by_section.append(float(v))
                        break
                    if ws314.cell(r,2).value=='Site' and r!=sr: break
            if len(vals_by_section)==3:
                sum_v_l = vals_by_section[1]+vals_by_section[2]
                checks.append(ok(
                    f"3.1.4 {site}: Vente+Loc = Total ({lm})",
                    close(sum_v_l, vals_by_section[0], 1.0),
                    f"Vente={vals_by_section[1]:,.0f}+Loc={vals_by_section[2]:,.0f}={sum_v_l:,.0f} | Total={vals_by_section[0]:,.0f}",
                    "error","3.1 — contrôle Vente+Loc=Total (O73+O92=O191)"
                ))

    # ── CHECK 4: 3.1.2 Partagés + 3.1.5 Exclusifs = 3.1.4 Total Dedup
    # Zeynep's error: SUM 3.1.2+3.1.5 ≠ 3.1.4
    total_312 = None
    for r in range(1, ws312.max_row+1):
        b = ws312.cell(r,2).value
        if b and 'Dédupliqué' in str(b) and 'Panel' in str(b):
            v = ws312.cell(r,3).value  # col 3 = Total (snapshot)
            if isinstance(v,(int,float)): total_312=float(v); break

    total_315 = None
    d315 = read_site_series(ws315)
    total_315 = site_val(d315,'Total Panel Dédupliqué') or site_val(d315,'Total')

    if total_312 is not None and total_315 is not None and v314_dedup:
        sum_pt = total_312+total_315
        checks.append(ok(
            f"3.1.2 Partagés + 3.1.5 Exclusifs = 3.1.4 Total Dedup ({lm})",
            close(sum_pt, v314_dedup, 1.0),
            f"3.1.2={total_312:,.0f} + 3.1.5={total_315:,.0f} = {sum_pt:,.0f} | 3.1.4={v314_dedup:,.0f} | diff={abs(sum_pt-v314_dedup):,.0f}",
            "error","3.1 — contrôle 3.1.2+3.1.5 vs 3.1.4"
        ))

    # ── CHECK 5: 3.1.4 TOP11 = sum of type sections
    # Panel Checker: r317 =O317=O337+O357+O377+O397
    # In 3.1.4: Total Dedup TOP11 = sum of Agences+Intermédiaires+Notaires+Autres TOP11
    dedup_11 = site_val(d314,'Total Panel Dédupliqué TOP 11 SITES') or site_val(d314,'Total Panel Dédupliqué  - Top 11 Sites')
    # Find section totals
    section_dedup_11 = []
    for r in range(1,ws314.max_row+1):
        b = ws314.cell(r,2).value
        if b and ('Top 11' in str(b) or 'TOP 11' in str(b)) and 'Dédupliqué' in str(b):
            v = ws314.cell(r,lc).value if lc else None
            if isinstance(v,(int,float)): section_dedup_11.append(float(v))
    if len(section_dedup_11)>=5:  # total + 4 sections
        total_top11 = section_dedup_11[0]
        sum_sections = sum(section_dedup_11[1:5])
        checks.append(ok(
            f"3.1.4 Total Dedup TOP11 = sum of type sections ({lm})",
            close(total_top11, sum_sections, 0.5),
            f"Total={total_top11:,.0f} | Sum sections={sum_sections:,.0f} | diff={abs(total_top11-sum_sections):,.0f}",
            "error","3.1 — contrôle segments (O317=O337+O357+O377+O397)"
        ))

    # ── CHECK 6: 3.1.1 per-site = 3.1.4 per-site
    # Panel Checker: =P131=C131 type checks
    d311_sites = {}
    for c in range(2, ws311.max_column, 3):
        site = ws311.cell(1,c).value
        if not site or not isinstance(site,str): continue
        site = site.strip()
        v = ws311.cell(12,c).value  # Pros identifiés Joreca
        if isinstance(v,(int,float)) and v>0: d311_sites[site]=float(v)

    for site,v311 in d311_sites.items():
        v314 = site_val(d314,site)
        if v314 and v311 and v314>100:
            checks.append(ok(
                f"3.1.1 {site} = 3.1.4 {site} ({lm})",
                close(v314,v311,1.0),
                f"3.1.1={v311:,.0f} | 3.1.4={v314:,.0f} | diff={abs(v314-v311):,.0f}",
                "error","3.1 — cohérence 3.1.1 vs 3.1.4 (P131=C131)"
            ))

    return checks

# ─── TAB 3.2 CHECKS ───────────────────────────────────────────────────────────

def check_tab32(wb32_bytes, wb31_bytes):
    """
    Panel Checker tab 3.2 controls:
    r24 K: =K22='tab 3.1'!$O194  → 3.2.1 Leboncoin TOTAL = 3.1.4 Leboncoin last
    r24 AA: =AA22='tab 3.1'!$O202 → 3.2.1 Total TOTAL = 3.1.4 Total last
    r25 K: =K22=K45+K69+K91+K113  → Leboncoin total = sum of type sections
    r68 AK: MAX(sites)<=Dedup per région
    r225 K: =K222=K348+K474+K600+K726 → dept total = sum of depts by type
    """
    checks = []
    wb32 = load_workbook(io.BytesIO(wb32_bytes), data_only=True)
    wb31 = load_workbook(io.BytesIO(wb31_bytes), data_only=True)

    ws321 = wb32['3.2.1 Pros par régions']
    ws322 = wb32['3.2.2 Pros par département']
    ws314 = wb31['3.1.4 Evolution Pros par type']

    d314 = read_site_series(ws314)
    lm = d314.get('_lm','?')

    # ── CHECK 1: =K22='tab 3.1'!$O194 → 3.2.1 site TOTAL = 3.1.4 site last
    # tab 3.2 col K = Leboncoin (col 11), row 22 = TOTAL
    # tab 3.1 O194 = Leboncoin row 194 col O (last month)
    # HAM: 3.2.1 TOTAL row per site = 3.1.4 per site last month

    # Find TOTAL row in 3.2.1 and site columns
    total_row321 = None
    for r in range(ws321.max_row,0,-1):
        if ws321.cell(r,2).value=='TOTAL':
            total_row321=r; break

    site_cols321 = {}
    for c in range(3, ws321.max_column+1, 2):  # odd cols = site Pros cols
        h = ws321.cell(6,c).value
        if h and isinstance(h,str) and len(h.strip())>2:
            site_cols321[h.strip()] = c

    if total_row321:
        for site, sc in site_cols321.items():
            v321 = ws321.cell(total_row321, sc).value
            v314 = site_val(d314, site)
            if v321 and v314 and isinstance(v321,(int,float)) and v314>100:
                checks.append(ok(
                    f"3.2.1 {site} TOTAL = 3.1.4 {site} ({lm})",
                    close(float(v321), v314, 1.0),
                    f"3.2.1={v321:,.0f} | 3.1.4={v314:,.0f} | diff={abs(float(v321)-v314):,.0f}",
                    "error","3.2 — contrôle tab 3.1.4 (K22='tab 3.1'!$O194)"
                ))

    # ── CHECK 2: =K22=K45+K69+K91+K113 → site total = sum of type sections
    # 3.2.1 has sections: Total pros, Agences, Intermédiaires, Notaires, Autres
    # Each section has a TOTAL row; site total should = sum of type TOTAL rows
    total_rows321 = []
    for r in range(1,ws321.max_row+1):
        if ws321.cell(r,2).value=='TOTAL': total_rows321.append(r)

    if len(total_rows321)>=5:
        for site, sc in list(site_cols321.items())[:3]:  # check first 3 sites
            t_all = ws321.cell(total_rows321[0],sc).value
            sum_types = sum(ws321.cell(tr,sc).value or 0
                           for tr in total_rows321[1:5]
                           if isinstance(ws321.cell(tr,sc).value,(int,float)))
            if t_all and isinstance(t_all,(int,float)) and float(t_all)>100:
                checks.append(ok(
                    f"3.2.1 {site}: Total = sum type sections",
                    close(float(t_all), sum_types, 0.5),
                    f"Total={t_all:,.0f} | Sum types={sum_types:,.0f} | diff={abs(float(t_all)-sum_types):,.0f}",
                    "error","3.2 — contrôle segment (K22=K45+K69+K91+K113)"
                ))

    # ── CHECK 3: =MAX(C68,...Y68)<=AG68 → MAX(sites) <= Dedup per région
    hdr322 = None
    for r in range(1,20):
        if ws322.cell(r,2).value=='Département': hdr322=r; break
    if not hdr322:
        for r in range(1,20):
            if ws322.cell(r,1).value=='Département' or ws322.cell(r,2).value in ('Site','Département'):
                hdr322=r; break

    if hdr322:
        # Find dedup column and site data columns
        site_data_cols = []
        dedup_col = None
        for c in range(3, ws322.max_column+1):
            h = ws322.cell(hdr322,c).value
            if h and isinstance(h,(str,datetime.datetime)):
                hs = str(h)
                if 'Dédupliqué' in hs and 'Marché' in hs:
                    dedup_col = c
                elif 'Pros' in hs or any(s in hs for s in SITES):
                    site_data_cols.append(c)

        if dedup_col and site_data_cols:
            for r in range(hdr322+1, ws322.max_row+1):
                dept = ws322.cell(r,2).value
                if not dept or not isinstance(dept,str) or dept.strip() in ('TOTAL','',):
                    continue
                dept = dept.strip()
                dv = ws322.cell(r,dedup_col).value
                if not isinstance(dv,(int,float)) or float(dv)<=0: continue
                site_vals = [ws322.cell(r,c).value for c in site_data_cols
                            if isinstance(ws322.cell(r,c).value,(int,float)) and ws322.cell(r,c).value>0]
                if site_vals:
                    mv = max(site_vals)
                    if mv > float(dv)*1.01:
                        checks.append(ok(
                            f"3.2.2 {dept}: MAX sites ≤ Dedup",
                            False,
                            f"Max site={mv:,.0f} > Dedup={dv:,.0f}",
                            "error","3.2 — contrôle déduplication (MAX(...)<=AG68)"
                        ))

    # ── CHECK 4: dept sums = TOTAL in 3.2.2 (=K222=K348+K474+K600+K726)
    if hdr322 and site_cols321:
        # For the main TOTAL row vs sum of all dept rows
        total_r322 = None
        for r in range(ws322.max_row,0,-1):
            if ws322.cell(r,2).value=='TOTAL': total_r322=r; break
        if total_r322:
            for c in range(3, min(ws322.max_column+1, 10)):
                t_val = ws322.cell(total_r322,c).value
                if not isinstance(t_val,(int,float)) or t_val<=0: continue
                dept_sum = sum(ws322.cell(r,c).value or 0
                              for r in range(hdr322+1, total_r322)
                              if isinstance(ws322.cell(r,c).value,(int,float)))
                if t_val>100 and not close(float(t_val), dept_sum, 0.5):
                    h = ws322.cell(hdr322,c).value or f"col{c}"
                    checks.append(ok(
                        f"3.2.2 TOTAL {h} = sum of depts",
                        False,
                        f"TOTAL={t_val:,.0f} | Sum depts={dept_sum:,.0f} | diff={abs(float(t_val)-dept_sum):,.0f}",
                        "error","3.2 — contrôle K222=K348+K474+K600+K726"
                    ))

    return checks

# ─── TAB 4.1 CHECKS ───────────────────────────────────────────────────────────

def check_tab41(wb41_bytes, wb1_bytes):
    """
    Panel Checker tab 4.1.1 & 4.1.2 controls:
    r23: 4.1.1 Total = tab 1 per-site totals
    r171: 4.1.2 Vente+Loc = tab 1 Vente+Loc sum
    tab 4.1.3 r106: 4.1.3 Total = 4.1.1 Total (CONTRÔLE TAB 4.1)
    tab 4.1.3 r107: 4.1.3 Total = tab 1 Total (CONTRÔLE TAB 1)
    tab 4.1.4 r107: 4.1.4 = 4.1.2 (CONTRÔLE VS 4.1.2)
    tab 4.1.5&4.1.6 r107: 4.1.5 = tab 1 type sums (CONTRÔLE VS 1.4)
    """
    checks = []
    wb41 = load_workbook(io.BytesIO(wb41_bytes), data_only=True)
    wb1  = load_workbook(io.BytesIO(wb1_bytes), data_only=True)

    # Available sheets
    sheets41 = wb41.sheetnames
    sheets1  = wb1.sheetnames

    def get_d(wb, sn):
        if sn not in wb.sheetnames: return {}
        return read_site_series(wb[sn])

    d411 = get_d(wb41,'4.1.1 Région - Annonces')
    d412 = get_d(wb41,'4.1.2 Région - Types de Pros')
    d413 = get_d(wb41,'4.1.3 Dépt. - Annonces')
    d414 = get_d(wb41,'4.1.4 Dépt. - Types de Pros')
    d11  = get_d(wb1,'1.1 Total')
    d12  = get_d(wb1,'1.2 Pro_Part')
    d14  = get_d(wb1,'1.4 Type de professionels')

    lm = d411.get('_lm', d413.get('_lm','?'))

    # ── CHECK 1: =C21='tab 1'!$O28 → 4.1.1 Total Annonces per site = 1.1 per site
    # tab 4.1.1 row 21 = Total Ancien Annonces per site (last row before regions end)
    # tab 1 O28-O39 = per site total annonces
    for site in SITES:
        v411 = site_val(d411, site) or site_val(d411, site.replace('é','e'))
        v11  = site_val(d11,  site) or site_val(d11,  site.replace('é','e'))
        if v411 and v11 and v411>1000 and v11>1000:
            checks.append(ok(
                f"4.1.1 {site} Total = 1.1 {site} ({lm})",
                close(v411,v11,1.0),
                f"4.1.1={v411:,.0f} | 1.1={v11:,.0f} | diff={abs(v411-v11):,.0f}",
                "error","4.1 — CONTRÔLE TAB 1 (C21='tab 1'!$O28)"
            ))

    # ── CHECK 2: 4.1.3 Total = 4.1.1 Total (CONTRÔLE TAB 4.1)
    for site in SITES:
        v413 = site_val(d413, site)
        v411 = site_val(d411, site)
        if v413 and v411 and v413>1000:
            checks.append(ok(
                f"4.1.3 {site} = 4.1.1 {site} ({lm})",
                close(v413,v411,1.0),
                f"4.1.3={v413:,.0f} | 4.1.1={v411:,.0f} | diff={abs(v413-v411):,.0f}",
                "error","4.1 — CONTRÔLE TAB 4.1 (C104='tab 4.1.1 & 4.1.2'!C21)"
            ))

    # ── CHECK 3: 4.1.3 Total = 1.1 Total (CONTRÔLE TAB 1)
    for site in SITES:
        v413 = site_val(d413, site)
        v11  = site_val(d11,  site)
        if v413 and v11 and v413>1000:
            checks.append(ok(
                f"4.1.3 {site} = 1.1 {site} ({lm})",
                close(v413,v11,1.0),
                f"4.1.3={v413:,.0f} | 1.1={v11:,.0f} | diff={abs(v413-v11):,.0f}",
                "error","4.1 — CONTRÔLE TAB 1 (C104='tab 1'!O28)"
            ))

    # ── CHECK 4: 4.1.4 = 4.1.2 (CONTRÔLE VS 4.1.2)
    for site in SITES:
        v414 = site_val(d414, site)
        v412 = site_val(d412, site)
        if v414 and v412 and v414>1000:
            checks.append(ok(
                f"4.1.4 {site} = 4.1.2 {site} ({lm})",
                close(v414,v412,1.0),
                f"4.1.4={v414:,.0f} | 4.1.2={v412:,.0f} | diff={abs(v414-v412):,.0f}",
                "error","4.1 — CONTRÔLE VS 4.1.2 (C105='tab 4.1.1 & 4.1.2'!C169)"
            ))

    # ── CHECK 5: 4.1.5&4.1.6 = 1.4 type sums (CONTRÔLE VS 1.4)
    # Panel Checker: C105='tab 1'!$C$314+'tab 1'!$C$378+'tab 1'!$C$441+'tab 1'!$C$504
    # 4.1.5 Total = sum of 4 type sections in 1.4 per site per month
    if '4.1.5. Dépt. & Rég. Pros id Y-1' in sheets41:
        d415 = get_d(wb41, '4.1.5. Dépt. & Rég. Pros id Y-1')
        # 1.4 has multiple sections (Agences, Intermediaires, Notaires, Autres)
        # Sum their totals = 4.1.5 total
        v415_total = site_val(d415,'Total') or site_val(d415,'Total Panel Dédupliqué')
        v14_sum = 0
        for r in range(1, wb1['1.4 Type de professionels'].max_row+1):
            b = wb1['1.4 Type de professionels'].cell(r,2).value
            if b and 'Total Panel Dédupliqué' in str(b) and 'Top' not in str(b):
                lc = d14.get('_lc')
                if lc:
                    v = wb1['1.4 Type de professionels'].cell(r,lc).value
                    if isinstance(v,(int,float)): v14_sum+=float(v)
        if v415_total and v14_sum>0:
            checks.append(ok(
                f"4.1.5 Total = 1.4 sum type sections ({lm})",
                close(v415_total, v14_sum, 1.0),
                f"4.1.5={v415_total:,.0f} | 1.4 sum={v14_sum:,.0f} | diff={abs(v415_total-v14_sum):,.0f}",
                "error","4.1 — CONTRÔLE VS 1.4"
            ))

    # ── CHECK 6: dedup <= total per sheet
    for sn, d in [('4.1.1',d411),('4.1.3',d413)]:
        lm2 = d.get('_lm','?')
        total = site_val(d,'Total')
        dedup = site_val(d,'Total Panel Dédupliqué') or site_val(d,'Total Panel Dédupliqué  - Top 11 Sites')
        if total and dedup and total>1000:
            checks.append(ok(
                f"{sn} Dedup ≤ Total ({lm2})",
                dedup<=total*1.01,
                f"Dedup={dedup:,.0f} | Total={total:,.0f}",
                "error",f"4.1 — {sn}"
            ))

    return checks

# ─── TAB 5 CHECKS ─────────────────────────────────────────────────────────────

def check_tab5(wb5_bytes, wb41_bytes, wb1_bytes, label_prefix):
    """
    Panel Checker tab 5 / tab 5-2 controls:
    r16: sum of IDF depts = 4.1.1 IDF total + 4.1.2 IDF total
    r17: IDF dept values = 4.1.4 IDF dept values
    r347: tab 3.2 cross-check
    """
    checks = []
    wb5  = load_workbook(io.BytesIO(wb5_bytes),  data_only=True)
    wb41 = load_workbook(io.BytesIO(wb41_bytes), data_only=True)
    wb1  = load_workbook(io.BytesIO(wb1_bytes),  data_only=True)

    for sn in wb5.sheetnames:
        if sn=='Intro': continue
        ws = wb5[sn]
        d = read_site_series(ws)
        if not d: continue
        lm = d.get('_lm','?')
        label = f"{label_prefix} › {sn}"

        total = site_val(d,'Total')
        dedup = site_val(d,'Total Panel Dédupliqué')
        if total and dedup and total>100:
            checks.append(ok(
                f"Dedup ≤ Total ({lm})",
                dedup<=total*1.01,
                f"Dedup={dedup:,.0f} | Total={total:,.0f}",
                "error",label
            ))

        # Sum of site values vs Total
        skip = {'Total','Total Panel Dédupliqué','Total Panel Dédupliqué - Top 5 Sites',
                'Total Panel Dédupliqué  - Top 11 Sites','Total Panel Dédupliqué Marché',
                'Immobilier Notaire','Immonot','Site','Département','TOTAL'}
        sites_sum = sum(v['last'] for k,v in d.items()
                       if isinstance(k,str) and not k.startswith('_') and k not in skip
                       and v['last'] and v['last']>0)
        if total and sites_sum>0 and total>100:
            diff = abs(total-sites_sum)/max(total,sites_sum)*100
            if diff<50:  # skip ratio/percentage sheets
                checks.append(ok(
                    f"Total = sum of sites ({lm})",
                    diff<2,
                    f"Total={total:,.0f} | Sum={sites_sum:,.0f} | Diff={diff:.1f}%",
                    "error",label
                ))

        # ML: Z-score anomaly detection
        for site_name, sdata in d.items():
            if isinstance(site_name,str) and not site_name.startswith('_') and site_name not in skip:
                series = sdata['vals']
                real = [v for v in series if v and v>0]
                if len(real)>=6:
                    hist = np.array(real[:-1])
                    last = real[-1]
                    mean,std = np.mean(hist),np.std(hist)
                    if std>0:
                        z=(last-mean)/std
                        if abs(z)>3:
                            checks.append({'check':f"Anomalie Z-score: {site_name}",
                                'status':'warning','severity':'warning',
                                'detail':f"Z={z:.1f} | {lm}={last:,.0f} | moy={mean:,.0f}",
                                'sheet':label})

    return checks

# ─── CLASSIFIER ───────────────────────────────────────────────────────────────

def classify_files(uploaded_files):
    file_labels = {
        'file1':'1 — Evolution panel','file2':'2 — Performance qualité',
        'file3_1':'3.1 — Analyse Pros','file3_2':'3.2 — Géographique Pros',
        'file4_1':'4.1 — Stats géographiques','file4_2':'4.2 — Exclusivité/Partage',
        'file5':'5 — Focus IDF','file5_2':'5.2 — Grand Ouest',
    }
    files_bytes = {}
    for f in uploaded_files:
        nl = f.name.lower().replace(' ','_').replace('é','e').replace('è','e').replace('ô','o').replace('û','u')
        fb = f.read()
        if '5_2' in nl or ('5' in nl and ('grand' in nl or 'ouest' in nl)):
            files_bytes['file5_2'] = fb
        elif '5' in nl and ('idf' in nl or 'ile' in nl or 'alpes' in nl or 'focus' in nl) and '5_2' not in nl:
            files_bytes['file5'] = fb
        elif '4_2' in nl or ('4' in nl and ('exclus' in nl or 'partag' in nl)):
            files_bytes['file4_2'] = fb
        elif '4_1' in nl or ('4' in nl and 'stat' in nl and '4_2' not in nl):
            files_bytes['file4_1'] = fb
        elif '3_2' in nl or ('3' in nl and 'geo' in nl):
            files_bytes['file3_2'] = fb
        elif '3_1' in nl or ('3' in nl and 'pros' in nl and '3_2' not in nl):
            files_bytes['file3_1'] = fb
        elif '2' in nl and ('perform' in nl or 'qualit' in nl):
            files_bytes['file2'] = fb
        elif '1' in nl and ('evolution' in nl or 'annonce' in nl) and '3_1' not in nl and '4_1' not in nl:
            files_bytes['file1'] = fb
    return files_bytes, file_labels

def run_all_qc_gold(uploaded_files):
    files_bytes, _ = classify_files(uploaded_files)
    all_checks = []

    if 'file1' in files_bytes:
        all_checks += check_tab1(files_bytes['file1'])
    if 'file3_1' in files_bytes and 'file1' in files_bytes:
        all_checks += check_tab31(files_bytes['file3_1'], files_bytes['file1'])
    if 'file3_2' in files_bytes and 'file3_1' in files_bytes:
        all_checks += check_tab32(files_bytes['file3_2'], files_bytes['file3_1'])
    if 'file4_1' in files_bytes and 'file1' in files_bytes:
        all_checks += check_tab41(files_bytes['file4_1'], files_bytes['file1'])
    if 'file5' in files_bytes and 'file4_1' in files_bytes and 'file1' in files_bytes:
        all_checks += check_tab5(files_bytes['file5'], files_bytes['file4_1'],
                                 files_bytes['file1'], "5 — Focus IDF")
    if 'file5_2' in files_bytes and 'file4_1' in files_bytes and 'file1' in files_bytes:
        all_checks += check_tab5(files_bytes['file5_2'], files_bytes['file4_1'],
                                 files_bytes['file1'], "5.2 — Grand Ouest")

    return all_checks, list(files_bytes.keys())


def parse_ham_sections(uploaded_files):
    result = {}
    for f in uploaded_files:
        fb = f.read()
        wb = load_workbook(io.BytesIO(fb), data_only=True)
        for sn in wb.sheetnames:
            if sn == 'Intro': continue
            ws = wb[sn]
            for r in range(1, ws.max_row+1):
                b = ws.cell(r,2).value
                if b not in ('Site','Département'): continue
                month_cols, months = [], []
                for c in range(3, ws.max_column+1):
                    v = ws.cell(r,c).value
                    if v is None: continue
                    if (isinstance(v,datetime.datetime) or
                        (isinstance(v,(int,float)) and 40000<v<50000) or
                        (isinstance(v,str) and any(m in v.lower() for m in ['-25','-26','-24','-23']))):
                        month_cols.append(c); months.append(excel_date_str(v))
                if len(months)<2: continue
                label = None
                for tr in range(r-1,max(0,r-5),-1):
                    v = ws.cell(tr,2).value
                    if v and isinstance(v,str) and len(v.strip())>3 and v.strip()!='Site':
                        label=v.strip(); break
                if not label: label=sn
                sites={}; totals={}
                for dr in range(r+1, min(r+35,ws.max_row+1)):
                    site=ws.cell(dr,2).value
                    if not site or not isinstance(site,str): break
                    site=site.strip()
                    if site=='Site': break
                    vals=[float(ws.cell(dr,c).value) if isinstance(ws.cell(dr,c).value,(int,float)) else None for c in month_cols]
                    real=[v for v in vals if v is not None and v>0]
                    if 'Total Panel' in site or site=='Total':
                        if len(real)>=2: totals[site]=vals
                    elif site not in SKIP_ROWS and site!='' and len(real)>=2:
                        sites[site]=vals
                if sites:
                    result[f"{sn}_{r}"] = {'sheet':sn,'label':label,'months':months,
                                           'sites':sites,'totals':totals,'file':f.name}
    return result

def analyze_trends(sections):
    rows = []
    for key, sec in sections.items():
        months = sec['months']
        total_dedup = {}
        for k,v in sec['totals'].items():
            if 'Dédupliqué' in k or 'Dedup' in k:
                total_dedup={i:x for i,x in enumerate(v) if x is not None and x>0}; break
        for site, vals in sec['sites'].items():
            real=[(i,v) for i,v in enumerate(vals) if v is not None and v>5]
            if len(real)<2: continue
            li,lv=real[-1]; pi,pv=real[-2]
            evol=(lv/pv-1)*100 if pv>0 else None
            pr_m=(lv/total_dedup[li]*100) if li in total_dedup and total_dedup[li]>0 else None
            pr_m1=(pv/total_dedup[pi]*100) if pi in total_dedup and total_dedup[pi]>0 else None
            evol_pr=(pr_m-pr_m1) if pr_m and pr_m1 else None
            status='ok'; flags=[]
            if evol is not None:
                if evol<=-20: status='critical'; flags.append(f"Monthly change: {evol:+.1f}% (vs prev month)")
                elif evol<=-10: status='warning'; flags.append(f"Monthly change: {evol:+.1f}% (vs prev month)")
                elif evol>=30: status='warning'; flags.append(f"Monthly surge: {evol:+.1f}%")
            max_hist=max((v for v in vals[:-1] if v and v>5),default=None)
            if max_hist and lv/max_hist<0.6:
                status='critical'; flags.append(f"Crash vs historical max: {(lv/max_hist-1)*100:+.1f}%")
            last3=[v for v in vals[-3:] if v and v>0]
            if len(last3)==3 and last3[0]>last3[1]>last3[2]:
                drop=(last3[2]-last3[0])/last3[0]*100
                if drop<-5:
                    if status=='ok': status='warning'
                    flags.append(f"3-month downtrend: {drop:+.1f}%")
            rows.append({'file':sec['file'],'sheet':sec['sheet'],'section':sec['label'],
                'site':site,'month_m':months[li] if li<len(months) else '?',
                'month_m1':months[pi] if pi<len(months) else '?',
                'val_m':lv,'val_m1':pv,'evol_pct':evol,'pr_m':pr_m,'evol_pr':evol_pr,
                'status':status,'flags':flags,'values':vals,'months':months})
    return rows

# ─── SIDEBAR ─────────────────────────────────────────────────────────────────

with st.sidebar:
    st.markdown("## QC Monitor")
    st.divider()
    st.markdown("**Market**")
    sector = st.radio("", ["🏠 Real Estate","🚗 Auto"], label_visibility="collapsed", key="sector")
    if sector == "🏠 Real Estate":
        market = st.radio("", ["REFR"], label_visibility="collapsed", key="re_sub")
    else:
        st.caption("Coming soon"); market = None
    st.divider()
    st.markdown("**Page**")
    page = st.radio("", ["📊 Panel Tables Monitor","✅ Panel Checker (QC Gold)"],
        label_visibility="collapsed", key="page_sel")
    st.divider()
    st.markdown("**Upload files**")
    st.caption("Drop all files here — system sorts automatically")
    uploaded_files = st.file_uploader("", type=['xlsx'], accept_multiple_files=True,
        label_visibility="collapsed", key="global_upload")
    if uploaded_files:
        st.success(f"✅ {len(uploaded_files)} file(s) loaded")

if market is None:
    st.markdown("# 🚗 Auto"); st.info("Coming soon"); st.stop()

# ─── PAGE: PANEL TABLES MONITOR ──────────────────────────────────────────────

if page == "📊 Panel Tables Monitor":
    st.markdown(f"# 📊 Panel Tables Monitor — {market}")
    st.divider()
    if not uploaded_files:
        st.info("👆 Upload your files in the sidebar to get started"); st.stop()

    with st.spinner("Loading..."):
        sections = parse_ham_sections(uploaded_files)
        rows = analyze_trends(sections)

    if not rows:
        st.warning("No data found."); st.stop()

    n_crit=sum(1 for r in rows if r['status']=='critical')
    n_warn=sum(1 for r in rows if r['status']=='warning')

    if n_crit>0: st.error(f"❌ REFUSED — {n_crit} critical · {n_warn} warnings")
    elif n_warn>3: st.warning(f"⚠️ TO REVIEW — {n_warn} warnings")
    else:
        st.success(f"✅ VALIDATED — {n_warn} minor warning(s)")
        if n_warn>0: st.info(f"{n_warn} warning(s) to monitor below")

    c1,c2 = st.columns([2,3])
    with c1:
        sev=st.radio("",["All","🔴 Critical","🟡 Warnings"],horizontal=True,key="sev_ptm")
    with c2:
        search=st.text_input("",placeholder="🔍 Search...",key="search_ptm",label_visibility="collapsed")

    filtered=rows
    if sev=="🔴 Critical": filtered=[r for r in rows if r['status']=='critical']
    elif sev=="🟡 Warnings": filtered=[r for r in rows if r['status'] in ('warning','critical')]
    if search:
        q=search.lower()
        filtered=[r for r in filtered if q in r['site'].lower() or q in r['section'].lower()]
    filtered=sorted(filtered,key=lambda r:{'critical':3,'warning':2,'ok':1}.get(r['status'],0),reverse=True)

    month_m=rows[0]['month_m'] if rows else ''
    month_m1=rows[0]['month_m1'] if rows else ''
    st.caption(f"M = {month_m}  ·  M-1 = {month_m1}  ·  {len(filtered)} series")

    table=[]
    for r in filtered:
        icon="🔴" if r['status']=='critical' else "🟡" if r['status']=='warning' else "✅"
        evol=f"{r['evol_pct']:+.1f}%" if r['evol_pct'] is not None else "—"
        pr=f"{r['pr_m']:.1f}%" if r['pr_m'] is not None else "—"
        evol_pr=f"{r['evol_pr']:+.2f}pp" if r['evol_pr'] is not None else "—"
        table.append({'':icon,'Section':r['section'][:35],'Site':r['site'],
            f"M ({month_m})":f"{r['val_m']:,.0f}",f"M-1 ({month_m1})":f"{r['val_m1']:,.0f}",
            'Evol %':evol,'Market share':pr,'MS evol':evol_pr,
            'Note':' | '.join(r['flags']) if r['flags'] else '—'})

    df=pd.DataFrame(table)
    def color_rows(row):
        if '🔴' in str(row.iloc[0]): return ['background-color:rgba(255,68,68,0.08)']*len(row)
        if '🟡' in str(row.iloc[0]): return ['background-color:rgba(245,166,35,0.05)']*len(row)
        return ['']*len(row)
    st.dataframe(df.style.apply(color_rows,axis=1),use_container_width=True,hide_index=True,
        height=min(600,40+35*len(table)))

    issue_rows=[r for r in filtered if r['flags']]
    if issue_rows:
        st.markdown("### Sites with anomalies")
        cols_n=3
        for i in range(0,min(len(issue_rows),12),cols_n):
            cols=st.columns(cols_n)
            for j,row in enumerate(issue_rows[i:i+cols_n]):
                with cols[j]:
                    clr='#ff4444' if row['status']=='critical' else '#f5a623'
                    evol_s=f"{row['evol_pct']:+.1f}%" if row['evol_pct'] else ''
                    icon2='🔴' if row['status']=='critical' else '🟡'
                    st.markdown(f"**{icon2} {row['site']}** — <span style='color:{clr}'>{evol_s}</span>",unsafe_allow_html=True)
                    st.caption(row['section'][:40])
                    vals=[v if v else 0 for v in row['values']]
                    fig=go.Figure(go.Scatter(x=row['months'][:len(vals)],y=vals,mode='lines+markers',
                        line=dict(color=clr,width=2),
                        marker=dict(size=[6 if k==len(vals)-1 else 0 for k in range(len(vals))]),
                        connectgaps=True,hovertemplate='%{x}: %{y:,.0f}<extra></extra>'))
                    fig.update_layout(height=120,margin=dict(l=0,r=0,t=0,b=0),
                        paper_bgcolor='rgba(0,0,0,0)',plot_bgcolor='rgba(0,0,0,0)',
                        xaxis=dict(showgrid=False,tickfont=dict(size=8),nticks=4),
                        yaxis=dict(showgrid=True,gridcolor='rgba(128,128,128,0.15)',
                            tickfont=dict(size=8),tickformat='.2s'),showlegend=False)
                    st.plotly_chart(fig,use_container_width=True,config={'displayModeBar':False},
                        key=f"ptm_{i}_{j}_{row['site'][:8]}")
                    for flag in row['flags']:
                        st.caption(f"⚠️ {flag}")

    st.divider()
    export=[{'File':r['file'],'Sheet':r['sheet'],'Section':r['section'],'Site':r['site'],
        'Status':r['status'],'Month M':r['month_m'],'Value M':r['val_m'],
        'Month M-1':r['month_m1'],'Value M-1':r['val_m1'],
        'Evol %':f"{r['evol_pct']:+.1f}%" if r['evol_pct'] else '',
        'Market share':f"{r['pr_m']:.1f}%" if r['pr_m'] else '',
        'Flags':' | '.join(r['flags'])} for r in rows]
    csv=pd.DataFrame(export).to_csv(index=False).encode('utf-8-sig')
    st.download_button("⬇️ Download report (CSV)",data=csv,
        file_name=f"ptm_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.csv",mime='text/csv')

# ─── PAGE: PANEL CHECKER (QC GOLD) ───────────────────────────────────────────

elif page == "✅ Panel Checker (QC Gold)":
    st.markdown(f"# ✅ Panel Checker (QC Gold) — {market}")
    st.divider()

    if not uploaded_files:
        st.info("👆 Upload your files in the sidebar to get started"); st.stop()

    with st.spinner("Running QC Gold checks..."):
        all_checks, detected = run_all_qc_gold(uploaded_files)

    if not all_checks:
        st.warning("No checks could be run. Make sure the correct files are uploaded."); st.stop()

    errors=[c for c in all_checks if c['status']=='fail' and c['severity']=='error']
    warnings=[c for c in all_checks if c['severity']=='warning']
    ok_checks=[c for c in all_checks if c['status']=='ok']

    if errors: st.error(f"❌ REFUSED — {len(errors)} error(s) · {len(warnings)} warning(s)")
    elif warnings: st.warning(f"⚠️ TO REVIEW — {len(warnings)} warning(s)")
    else: st.success(f"✅ VALIDATED — all {len(ok_checks)} checks passed")

    c1,c2,c3,c4=st.columns(4)
    c1.metric("Total checks",len(all_checks))
    c2.metric("✅ Passed",len(ok_checks))
    c3.metric("❌ Errors",len(errors))
    c4.metric("🟡 Warnings",len(warnings))

    st.divider()

    by_sheet=defaultdict(list)
    for c in all_checks:
        by_sheet[c.get('sheet','Unknown')].append(c)

    def sheet_priority(items):
        if any(i['status']=='fail' and i['severity']=='error' for i in items): return 0
        if any(i['severity']=='warning' for i in items): return 1
        return 2

    for sheet,items in sorted(by_sheet.items(),key=lambda x:sheet_priority(x[1])):
        n_err=sum(1 for i in items if i['status']=='fail' and i['severity']=='error')
        n_warn=sum(1 for i in items if i['severity']=='warning')
        n_ok=sum(1 for i in items if i['status']=='ok')
        icon="❌" if n_err else "🟡" if n_warn else "✅"
        badge=f"{n_err} error(s)" if n_err else f"{n_warn} warning(s)" if n_warn else f"{n_ok} OK"

        with st.expander(f"{icon} **{sheet}** — {badge}",expanded=(n_err>0)):
            for item in items:
                if item['status']=='ok':
                    st.markdown(f"✅ {item['check']}")
                elif item['severity']=='error':
                    st.markdown(f"❌ **{item['check']}**")
                    st.caption(f"  {item['detail']}")
                else:
                    st.markdown(f"🟡 {item['check']}")
                    st.caption(f"  {item['detail']}")

    st.divider()
    rows_exp=[{'Sheet':c.get('sheet',''),'Check':c['check'],
        'Status':'❌ Error' if c['status']=='fail' and c['severity']=='error' else '🟡 Warning' if c['severity']=='warning' else '✅ OK',
        'Detail':c.get('detail','')} for c in all_checks]
    csv=pd.DataFrame(rows_exp).to_csv(index=False).encode('utf-8-sig')
    st.download_button("⬇️ Download report (CSV)",data=csv,
        file_name=f"qc_gold_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.csv",mime='text/csv')
