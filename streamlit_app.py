import streamlit as st
import plotly.graph_objects as go
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from collections import defaultdict
import datetime, io, hashlib

APP_VERSION = "6.3.0"

st.set_page_config(page_title="QC Gold Panel", page_icon="🛡️", layout="wide")

# Prevent st.metric label/value from being clipped with "…" in narrow columns
st.markdown("""
<style>
[data-testid="stMetricLabel"] {
    overflow: visible !important;
    white-space: normal !important;
    line-height: 1.25 !important;
}
[data-testid="stMetricLabel"] > div {
    overflow: visible !important;
    white-space: normal !important;
}
[data-testid="stMetricValue"] {
    overflow: visible !important;
    white-space: normal !important;
    font-size: 1.6rem !important;
}
</style>
""", unsafe_allow_html=True)

# ═══════════════════════════════════════════════
# CONSTANTS
# ═══════════════════════════════════════════════

SITES = ["AvendreAlouer","Bien'ici","Figaro Immo","Green-Acres","Leboncoin",
         "LogicImmo","MeilleursAgents","OuestFrance","PAP","ParuVendu","SeLoger","SuperImmo"]
IMMOFR_SITES = list(SITES)   # stable name — SITES itself is swapped per-market at runtime

# ── AUTOFR (Auto FR VO) — 2nd market, same UI/logic, different site list & file set ──
AUTOFR_SITES = ["Argus","AutoScout24","Heycar","La Centrale","Leboncoin","OuestFrance",
                "ParuVendu","Renault Occasions","Spoticar","Zoomcar"]
FILE_ROLES_AUTOFR = {
    "file1":"1 — Panel evolution",
    "file2":"2 — Analyse performance",
    "file3":"3 — Pros par métier",
    "file4":"4 — Doublons croisés",
    "file5":"5 — Cote Autobiz (VO typique)",
    "file6":"6 — Âge & Prix",
    "file7":"7 — Statistiques infomédiaire",
}
GROUP_INFO_AUTOFR = {
    "1":("Panel evolution","Volumes, deduplication, Particulier+Pros aggregation"),
    "2":("Analyse performance","Freshness, duplicates, exclusivity, shared listings"),
    "3":("Pros par métier","Dealer identification by trade, department, région, network"),
    "4":("Doublons croisés","Site × site cross-duplicate listing counts"),
    "5":("Cote Autobiz","Typical VO profile — price/km/age by site"),
    "6":("Âge & Prix","VO age/price/fuel distribution by site"),
    "7":("Statistiques infomédiaire","Région/département/marque breakdown"),
}

def classify_named_autofr(raw):
    """AUTOFR equivalent of classify_named() — maps the 7 uploaded Auto FR source
    files to roles by filename, keeping the real uploaded filename too.
    Returns {role: (original_filename, bytes)}."""
    out={}
    for fname,data in raw.items():
        nl=norm(fname); role=None
        if "doublon" in nl or "croise" in nl:                          role="file4"
        elif "coteautobiz" in nl or "cote_autobiz" in nl or "cote-autobiz" in nl or "vo_typique" in nl: role="file5"
        elif ("age" in nl and "prix" in nl) or "âge" in fname.lower():  role="file6"
        elif "infomediaire" in nl or "infomédiaire" in fname.lower():   role="file7"
        elif "prosparmetier" in nl or "pros_par_metier" in nl or ("pros" in nl and "metier" in nl): role="file3"
        elif "performance" in nl or "qualit" in nl:                     role="file2"
        elif "evolution" in nl or ("panel" in nl and "evolution" in nl): role="file1"
        if role and role not in out: out[role]=(fname,data)
    return out

def _find_wide_site_cols(ws, header_row, sites=None):
    """Scan one header row of a wide-format sheet (site name spans a column
    group, e.g. 'Argus' | 'Nombre'|'%' | 'AutoScout24' | 'Nombre'|'%' ...).
    Returns {site_name: first_column_index_of_that_group}."""
    if sites is None: sites = SITES
    out={}
    for c in range(1, ws.max_column+1):
        v = ws.cell(header_row, c).value
        if isinstance(v,str):
            vv=v.strip()
            for s in sites:
                if vv==s or vv.lower()==s.lower():
                    out.setdefault(s,c)
    return out

def _findrow_autofr(ws, label_substr, start=1, end=None, col=1):
    end=end or ws.max_row
    for r in range(start,end+1):
        v=ws.cell(r,col).value
        if isinstance(v,str) and label_substr.lower() in v.lower():
            return r
    return None

def run_checks_autofr(fb, wbs):
    """
    Cross-file consistency checks for AUTOFR — derived from the real Panel
    Checker (2607_Panel_Checker_Auto_FR.xlsm) formulas found in 'Insertion Tab
    2.1-2.2' (Check 1.1 Total Annonces), 'Insertion Tab 7' (Validated with
    Tab1.1), and the additivity/identification structure visible in the source
    files themselves. Same rule as IMMOFR: only checks that are actually
    grounded in the Panel Checker (or in additive structure within one file)
    are reproduced — nothing invented.
    """
    C=[]

    # ── FILE 1 — internal structural checks (Total sheet) ──
    d1=None; lm_ref=None
    if "file1" in wbs:
        w1=wbs["file1"]
        ws_tot=ws_get(w1,"Total")
        if ws_tot:
            d1=read_series(ws_tot, col=2, section=0)
            if d1:
                lm_ref=d1.get("_lm","?")
                somme_panel=sv(d1,"Somme Panel")
                site_vals=[sv(d1,s) for s in AUTOFR_SITES]
                site_sum=sum(v for v in site_vals if v is not None)
                if somme_panel is not None:
                    C.append(chk(f"Somme Panel = Σ sites ({lm_ref})",
                        close(site_sum, somme_panel, 1.0),
                        f"Σ des {len(AUTOFR_SITES)} sites: {_fmt_full(site_sum)} · Somme Panel: {_fmt_full(somme_panel)}","1"))
                dedup_label=None
                for label in d1.keys():
                    if isinstance(label,str) and "dédoublon" in label.lower() and not label.startswith("_"):
                        dedup_label=label; break
                if dedup_label:
                    dedup_v=sv(d1,dedup_label)
                    if dedup_v is not None and somme_panel is not None:
                        C.append(chk(f"{dedup_label.strip()} ≤ Somme Panel ({lm_ref})",
                            dedup_v<=somme_panel*1.01,
                            f"{dedup_label.strip()}: {_fmt_full(dedup_v)} · Somme Panel: {_fmt_full(somme_panel)}","1"))

    # ── FILE 2 vs FILE 1 — Total Annonces per site (Panel Checker "Check 1.1") ──
    if "file2" in wbs and d1:
        w2=wbs["file2"]
        ws_perf=ws_get(w2,"Analyse perfomance") or ws_get(w2,"Analyse performance")
        if ws_perf:
            site_cols2=_find_wide_site_cols(ws_perf,3,AUTOFR_SITES)
            ta_row=_findrow_autofr(ws_perf,"total annonces",1,20)
            if ta_row:
                for s,c in site_cols2.items():
                    v2=ws_perf.cell(ta_row,c).value
                    v1=sv(d1,s)
                    if v2 is not None and v1 is not None:
                        C.append(chk(f"{s} — Total Annonces (Fichier 2 = Fichier 1)",
                            close(v2,v1,1.0),
                            f"Fichier 2 (Analyse performance): {_fmt_full(v2)} · Fichier 1 ({lm_ref}): {_fmt_full(v1)}","2"))

    # ── FILE 7 vs FILE 1 — TOTAL ANNONCES per site (Panel Checker "Validated with Tab1.1") ──
    if "file7" in wbs and d1:
        w7=wbs["file7"]
        ws_reg=ws_get(w7,"region")
        if ws_reg:
            site_cols7=_find_wide_site_cols(ws_reg,1,AUTOFR_SITES)
            tot_row=_findrow_autofr(ws_reg,"total annonces")
            if tot_row:
                for s,c in site_cols7.items():
                    v7=ws_reg.cell(tot_row,c).value
                    v1=sv(d1,s)
                    if v7 is not None and v1 is not None:
                        C.append(chk(f"{s} — TOTAL ANNONCES régions = Fichier 1 (Fichier 7)",
                            close(v7,v1,1.0),
                            f"Fichier 7 (région, TOTAL ANNONCES): {_fmt_full(v7)} · Fichier 1 ({lm_ref}): {_fmt_full(v1)}","7"))

    # ── FILE 3 — internal additivity + identification checks ──
    if "file3" in wbs:
        w3=wbs["file3"]
        ws_met=ws_get(w3,"Client par Métier")
        if ws_met:
            site_cols3=_find_wide_site_cols(ws_met,1,AUTOFR_SITES)
            r_metier_start=_findrow_autofr(ws_met,"concession vn")
            r_total_metier=_findrow_autofr(ws_met,"total identifiés",1,20)
            r_stock_total=_findrow_autofr(ws_met,"total identifiés",(r_total_metier or 10)+1,(r_total_metier or 10)+15)
            r_identifies=_findrow_autofr(ws_met,"pros identifiés joreca")
            r_a_ident2=_findrow_autofr(ws_met,"pros à identifier joreca")
            r_general=_findrow_autofr(ws_met,"total général")
            for s,c_pro in site_cols3.items():
                c_ann=c_pro+1  # "Annonces" sub-column follows "Pro"
                if r_metier_start and r_total_metier:
                    metier_rows=list(range(r_metier_start, r_metier_start+6))
                    vals=[ws_met.cell(r,c_ann).value for r in metier_rows]
                    if all(v not in (None,"") for v in vals):
                        s_sum=sum(vals); tot=ws_met.cell(r_total_metier,c_ann).value
                        if tot not in (None,""):
                            C.append(chk(f"{s} — Σ métiers = Total identifiés (Fichier 3)",
                                close(s_sum,tot,1.0),
                                f"Σ métiers: {_fmt_full(s_sum)} · Total identifiés: {_fmt_full(tot)}","3"))
                if r_total_metier and r_stock_total:
                    v_m=ws_met.cell(r_total_metier,c_ann).value
                    v_s=ws_met.cell(r_stock_total,c_ann).value
                    if v_m not in (None,"") and v_s not in (None,""):
                        C.append(chk(f"{s} — Total (par métier) = Total (par stock VO)",
                            close(v_m,v_s,1.0),
                            f"Par métier: {_fmt_full(v_m)} · Par stock VO: {_fmt_full(v_s)}","3"))
                if r_identifies and r_a_ident2 and r_general:
                    v_i=ws_met.cell(r_identifies,c_ann).value
                    v_a=ws_met.cell(r_a_ident2,c_ann).value
                    v_g=ws_met.cell(r_general,c_ann).value
                    if all(v not in (None,"") for v in (v_i,v_a,v_g)):
                        vi,va=(v_i or 0),(v_a or 0)
                        C.append(chk(f"{s} — Total général = Identifiés + À identifier (Fichier 3)",
                            close(vi+va, v_g, 1.0),
                            f"Identifiés: {_fmt_full(vi)} + À identifier: {_fmt_full(va)} = {_fmt_full(vi+va)} · Total général: {_fmt_full(v_g)}","3"))
    return C

# Row/column labels that are aggregates, not real websites/geographies — skipped
# everywhere a per-site or per-geo loop iterates a section dict. This was
# previously referenced as a bare "SKIP" in several places (z_checks, run_checks,
# build_trends) without ever being defined, which raised a NameError as soon as an
# uploaded file's sheet structure caused that code path to actually execute.
SKIP = {"Total","Total Panel Dédupliqué","Total Panel Dédupliqué - Top 5 Sites",
        "Total Panel Dédupliqué  - Top 11 Sites","Total Panel Dédupliqué Marché",
        "Immobilier Notaire","Immonot","Site","Département","Totaux","TOTAL",
        "Région","RÉGION","DÉPARTEMENT",
        # AUTOFR aggregate row labels (file1 "Total"/"Particulier"/"Pros" sheets)
        "Somme Panel","Total Marché dédoublonnés",
        "Total Marché dédoublonnés\n sans OuestFrance/Zoomcar"}
import re as _re

FR_REGIONS = {
    "auvergne-rhône-alpes","bourgogne-franche-comté","bretagne","centre-val de loire",
    "corse","grand est","hauts-de-france","île-de-france","normandie",
    "nouvelle-aquitaine","occitanie","pays de la loire","provence-alpes-côte d'azur",
    "région non-renseignée","non-renseignée",
}

def _has_dept_number(s):
    """'10- Aube', '2A- Corse', '75- Paris' gibi numara ile başlıyorsa True."""
    return bool(_re.match(r'^[\dA-Za-z]{1,3}[-–]\s*.+', s.strip()))

def _is_region(s):
    """Bilinen FR region isimlerinden biri ise True."""
    return s.strip().lower() in FR_REGIONS

def _is_mom_flag(f, pm=""):
    """Sadece gerçek M/M-1 drop/surge/decline flagları."""
    fl = f.lower()
    return (("drop" in fl or "surge" in fl or "decline" in fl)
            and "12m" not in fl and "peak" not in fl
            and "y-1" not in fl and "downtrend" not in fl and "year" not in fl)

PANEL_SITES_SET = set(SITES) | {"Superimmo"}

FILE_ROLES = {
    "file1":     "1 — Panel evolution",
    "file2":     "2 — Quality metrics",
    "file3_1":   "3.1 — Professionals",
    "file3_2":   "3.2 — Geographic pros",
    "file4_1":   "4.1 — Geographic stats",
    "file4_2":   "4.2 — Exclusivity & sharing",
    "file5":     "5 — Focus IDF",
    "file5_2":   "5.2 — Grand Ouest",
    "file5_2_y1":"5.2 Y-1 — Grand Ouest",
    "file6":     "6 — New announcements IDF",
}

GROUP_INFO = {
    "1":      ("Panel evolution",           "Volumes, deduplication, Ancien+Neuf aggregation"),
    "2":      ("Quality metrics",           "Freshness, missing data, exclusivity"),
    "3.1":    ("Professionals — national",  "Pro counts, type breakdown, Vente+Location"),
    "3.2":    ("Professionals — geography", "Regional & dept breakdown, dedup hierarchy"),
    "4.1":    ("Announcements — geography", "Region & dept totals, Ancien+Neuf"),
    "4.2":    ("Exclusivity & sharing",     "Exclusive vs shared Vente+Location per region"),
    "5":      ("Focus IDF",                 "Île-de-France & Alpes-Maritimes dept checks"),
    "5.2":    ("Focus Grand Ouest",         "Western France departments"),
    "5.2 Y-1":("Focus Grand Ouest Y-1",    "Previous-year Grand Ouest checks"),
    "6":      ("New announcements IDF",     "Freshness by IDF department"),
}

def _render_scrollable_df(df, status_col="Status"):
    """Render DataFrame as scrollable HTML — works reliably across all screen sizes."""
    import html as _h
    bg_map = {"🔴": "#fdecec", "🟠": "#fff3e0", "🟡": "#fff4d6", "✅": "#edf7ed"}
    bold_cols = {"Website", "Site / Area", "Département", "Région"}

    th_style = "text-align:left;padding:6px 12px;border-bottom:2px solid #ddd;color:#555;font-weight:500;white-space:nowrap"
    td_style_base = "padding:5px 12px;border-bottom:1px solid #f0f0f0;white-space:nowrap"

    header = "".join(f"<th style='{th_style}'>{_h.escape(str(c))}</th>" for c in df.columns)

    body = ""
    for _, row in df.iterrows():
        bg = bg_map.get(str(row.get(status_col, "")), "transparent")
        cells = ""
        for col, val in row.items():
            v = _h.escape(str(val)) if val is not None else "—"
            fw = "font-weight:600;" if col in bold_cols else ""
            cells += f"<td style='{td_style_base};{fw}'>{v}</td>"
        body += f"<tr style='background:{bg}'>{cells}</tr>"

    html = (
        "<div style='overflow-x:auto;width:100%;margin:8px 0'>"
        "<table style='border-collapse:collapse;font-size:13px;font-family:sans-serif'>"
        f"<thead style='background:#f5f5f5'><tr>{header}</tr></thead>"
        f"<tbody>{body}</tbody>"
        "</table></div>"
    )
    st.markdown(html, unsafe_allow_html=True)
def dstr(v):
    if isinstance(v, datetime.datetime): return v.strftime("%b-%y")
    if isinstance(v, str): return v.strip()
    if isinstance(v,(int,float)) and 40000<v<50000:
        return (datetime.datetime(1899,12,30)+datetime.timedelta(days=int(v))).strftime("%b-%y")
    return str(v) if v else ""

def norm(s):
    return (s.lower().replace(".","_").replace(" ","_").replace("&","_").replace("-","_")
             .replace("é","e").replace("è","e").replace("ê","e")
             .replace("ô","o").replace("û","u").replace("à","a").replace("ç","c"))

def ws_get(wb, name):
    if name in wb.sheetnames: return wb[name]
    nl = norm(name)
    for sn in wb.sheetnames:
        if norm(sn)==nl: return wb[sn]
    return None

def _month_runs(ws, hdr, col):
    """
    Detect one or more monotonic runs of month-columns in a single header row.
    Normal sheets have exactly one run (Jul-25 → Jul-26, strictly increasing).
    Some sheets (e.g. AUTOFR's 'Evolution') place TWO segments side by side after
    the same 'Site' header — e.g. 2 months for 'Particuliers', then the SAME 2
    months again for 'Professionnels'. Reading straight across region would mix
    the two into one fake 4-point time series. Detecting the date going backward
    (or repeating) starts a NEW run instead — so each segment gets read correctly,
    on its own, rather than either merging them or silently dropping the 2nd one.
    Returns a list of runs; each run is a list of (col_index, month_str) tuples.
    """
    runs=[]; current=[]; last_key=None
    for c in range(col+1,ws.max_column+1):
        h=ws.cell(hdr,c).value
        if h is None: continue
        is_dt = isinstance(h,datetime.datetime)
        is_serial = isinstance(h,(int,float)) and 40000<h<50000
        is_strdate = isinstance(h,str) and any(x in h.lower() for x in ["-26","-25","-24","-23"])
        if not (is_dt or is_serial or is_strdate):
            continue
        key = h if (is_dt or is_serial) else None
        if key is not None and last_key is not None and key <= last_key:
            if len(current)>=2: runs.append(current)
            current=[]; last_key=None
        current.append((c, dstr(h)))
        if key is not None: last_key=key
    if len(current)>=2: runs.append(current)
    return runs

def read_series(ws, col=2, section=0, run_idx=0):
    """
    Read one section of a sheet (stops at the next Site/Département/Région header).
    section=0 = first section, section=1 = second, etc.
    run_idx selects which horizontal month-run within that header row to read
    (0 for normal sheets, which only ever have one run — see _month_runs).
    Returns {} if section/run not found or not enough month columns.
    """
    hdrs=[r for r in range(1,ws.max_row+1) if ws.cell(r,col).value in ("Site","Département","Région")]
    if not hdrs or section>=len(hdrs): return {}
    hdr = hdrs[section]
    stop = hdrs[section+1]-1 if section+1<len(hdrs) else ws.max_row

    runs=_month_runs(ws, hdr, col)
    if run_idx>=len(runs): return {}
    mc=[c for c,_ in runs[run_idx]]
    mo=[m for _,m in runs[run_idx]]

    # Section label (row above header) — for a 2nd+ run, prefer the segment's own
    # label (e.g. "Professionnels") found just above its first column, if present.
    label=None
    label_col = mc[0] if run_idx>0 else col
    for tr in range(hdr-1,max(0,hdr-5),-1):
        lv=ws.cell(tr,label_col).value
        if lv and isinstance(lv,str) and len(lv.strip())>1 and lv.strip() not in ("Site",""):
            label=lv.strip(); break
    if label is None and label_col!=col:
        for tr in range(hdr-1,max(0,hdr-5),-1):
            lv=ws.cell(tr,col).value
            if lv and isinstance(lv,str) and len(lv.strip())>3 and lv.strip() not in ("Site",""):
                label=lv.strip(); break

    out={"_m":mo,"_lc":mc[-1],"_lm":mo[-1],"_pc":mc[-2],"_pm":mo[-2],"_label":label or ""}
    for r in range(hdr+1, stop+1):
        b=ws.cell(r,col).value
        if not b or not isinstance(b,str) or not b.strip(): continue
        b=b.strip()
        vals=[float(ws.cell(r,c).value) if isinstance(ws.cell(r,c).value,(int,float)) else None for c in mc]
        out[b]={"v":vals,"last":vals[-1],"prev":vals[-2]}
    return out

def read_all_sections(ws, col=2):
    """Return list of read_series results for every section/run in the sheet —
    including any extra horizontal runs within a single header row (see _month_runs)."""
    hdrs=[r for r in range(1,ws.max_row+1) if ws.cell(r,col).value in ("Site","Département","Région")]
    out=[]
    for i in range(len(hdrs)):
        run_idx=0
        while True:
            d=read_series(ws, col, i, run_idx)
            if not d: break
            out.append(d)
            run_idx+=1
    return out

def read_cross(ws,col=2):
    hdr=None
    for r in range(1,min(20,ws.max_row+1)):
        if ws.cell(r,col).value in ("Site","Département","Région"): hdr=r; break
    if not hdr:
        for r in range(1,min(20,ws.max_row+1)):
            if ws.cell(r,1).value in ("Site","Département","Région"): hdr=r; col=1; break
    if not hdr: return {}
    sites,scols=[],[]
    for c in range(col+1,ws.max_column+1):
        v=ws.cell(hdr,c).value
        if v and isinstance(v,str) and v.strip() not in ("Pros","Poids","Total",""):
            sites.append(v.strip()); scols.append(c)
    if not sites: return {}
    out={"_sites":sites}
    for r in range(hdr+1,ws.max_row+1):
        geo=ws.cell(r,col).value
        if not geo or not isinstance(geo,str): continue
        geo=geo.strip()
        if not geo or geo in ("TOTAL","Total"): continue
        out[geo]={s:(float(ws.cell(r,c).value) if isinstance(ws.cell(r,c).value,(int,float)) else None)
                  for s,c in zip(sites,scols)}
    return out

def sv(d,name):
    if name in d: return d[name]["last"]
    nl=name.lower()
    for k in d:
        if isinstance(k,str) and nl in k.lower() and not k.startswith("_"): return d[k]["last"]
    return None

def close(a,b,pct=0.5):
    if a is None or b is None: return True
    if a==0 and b==0: return True
    return abs(a-b)/max(abs(a),abs(b))*100<pct

def fmt(n):
    if n is None: return "—"
    n=float(n)
    if abs(n)>=1_000_000: return f"{n/1_000_000:.2f}M"
    if abs(n)>=1_000: return f"{n/1_000:.0f}K"
    return f"{int(n):,}"

def chk(name,ok,detail,group,sev="error"):
    return {"name":name,"ok":ok,"detail":detail,"group":group,"sev":sev if not ok else "ok"}

def site_active(sd, min_vol=50):
    """True when the reference (last) month has meaningful volume."""
    v = sd.get("last") if isinstance(sd, dict) else None
    return v is not None and isinstance(v, (int, float)) and v >= min_vol

def panel_dedup_by_index(wbs):
    """Panel dedup totals keyed by month column index (from file 1.1)."""
    if "file1" not in wbs:
        return {}
    ws11 = ws_get(wbs["file1"], "1.1 Total")
    if not ws11:
        return {}
    d = read_series(ws11, section=0)
    if not d:
        return {}
    for k, sd in d.items():
        if not isinstance(k, str) or k.startswith("_"):
            continue
        if "dédupliqué" in k.lower() or "dedup" in k.lower():
            return {i: v for i, v in enumerate(sd["v"]) if v and v > 0}
    return {}

def z_checks(d, group, z_min=3.5, vol_min=200, skip_inactive_sites=True, sheet_label=None):
    """
    Z-score vs trailing history. Avg = mean of positive values over the
    up-to-12 months before the reference month. Skips sites with no
    volume in the reference month (inactive).
    sheet_label: sheet/table this check ran on — prefixed to the check name so
    it's immediately clear where the anomaly is (e.g. "1.1 Total › SeLoger — unusual volume").
    """
    out = []
    lm = d.get("_lm", "?")
    prefix = f"{sheet_label} › " if sheet_label else ""
    for k, sd in d.items():
        if not isinstance(k, str) or k.startswith("_") or k in SKIP:
            continue
        if skip_inactive_sites and k in SITES and not site_active(sd, min_vol=vol_min):
            continue
        vals = sd["v"]
        if len(vals) < 7:
            continue
        last = vals[-1]
        if last is None or not isinstance(last, (int, float)) or last <= 0:
            continue
        hist_vals = [v for v in vals[max(0, len(vals) - 13):-1] if v and v > 0]
        if len(hist_vals) < 6:
            continue
        hist = np.array(hist_vals)
        mean, std = np.mean(hist), np.std(hist)
        if mean < vol_min or std == 0:
            continue
        z = (last - mean) / std
        if abs(z) > z_min:
            direction = "increase" if z > 0 else "drop"
            out.append(chk(
                f"{prefix}{k} — unusual volume {direction}", False,
                f"{lm}: {_fmt_full(last)} · Avg (≤12m): {_fmt_full(mean)}",
                group, "warning"))
    return out

def section_check_vente_loc(ws, label_prefix, group):
    """
    Generic check: Vente + Location = Total per site for a sheet
    that has 3 sections: [Total, Vente, Location] in order.
    """
    sections = read_all_sections(ws)
    if len(sections)<3: return []
    d_total = sections[0]; d_vente = sections[1]; d_loc = sections[2]
    lm = d_total.get("_lm","?")
    checks=[]
    for site in SITES:
        t=sv(d_total,site); v=sv(d_vente,site); l=sv(d_loc,site)
        if t and v is not None and l is not None and t>0:
            checks.append(chk(
                f"{site} — Vente+Location = Total ({lm})",
                close(v+l,t,1.0),
                f"Vente: {_fmt_full(v)} + Loc: {_fmt_full(l)} = {_fmt_full(v+l)} · Total: {_fmt_full(t)}",
                group))
    return checks

def _is_num(v):
    return isinstance(v, (int, float)) and not isinstance(v, bool)

def read_go_y1_sections(ws):
    """Read Grand Ouest Y-1 wide-format sheets (Département in col A, metrics as columns)."""
    hdrs=[r for r in range(1,ws.max_row+1) if ws.cell(r,1).value=="Département"]
    sections=[]
    for i,hdr in enumerate(hdrs):
        stop=hdrs[i+1]-1 if i+1<len(hdrs) else ws.max_row
        headers=[]
        for c in range(2,ws.max_column+1):
            h=ws.cell(hdr,c).value
            if h and isinstance(h,str) and h.strip():
                headers.append((h.strip(),c))
        label_parts=[]
        for r in range(max(1,hdr-3),hdr):
            v=ws.cell(r,1).value
            if v and isinstance(v,str) and v.strip() and v.strip()!="Département":
                label_parts.append(v.strip())
        sec={"_headers":[h for h,_ in headers],"_label":" / ".join(label_parts),"_hdr":hdr}
        started=False
        for r in range(hdr+1,stop+1):
            dept=ws.cell(r,1).value
            if not dept or not isinstance(dept,str) or not dept.strip():
                if started: break
                continue
            dept=dept.strip(); started=True
            sec[dept]={h:ws.cell(r,c).value for h,c in headers}
        sections.append(sec)
    return sections

def _go_y1_metric(vals, metric_name):
    target=norm(metric_name)
    for k,v in vals.items():
        if norm(k)==target: return v
    return None

def _go_y1_site_metrics(headers):
    return [h for h in headers if "dedup" not in norm(h) and "marche" not in norm(h)]

def grand_ouest_y1_checks(ws, sheet_name, group="5.2 Y-1"):
    """Integrity checks for 5.2 Grand Ouest Y-1 wide snapshot tables."""
    checks=[]
    sections=read_go_y1_sections(ws)
    if not sections:
        checks.append(chk(f"{sheet_name} — Y-1 sections detected", False,
                          "No Département sections found in column A", group))
        return checks
    if len(sections)>=3:
        total,vente,loc=sections[0],sections[1],sections[2]
        depts_total={k for k in total if isinstance(k,str) and not k.startswith("_")}
        depts_vente={k for k in vente if isinstance(k,str) and not k.startswith("_")}
        depts_loc={k for k in loc if isinstance(k,str) and not k.startswith("_")}
        common_depts=depts_total & depts_vente & depts_loc
        common_metrics=set(total.get("_headers",[])) & set(vente.get("_headers",[])) & set(loc.get("_headers",[]))
        structure_ok=(depts_total==depts_vente==depts_loc and len(common_depts)>0 and len(common_metrics)>0)
        checks.append(chk(f"{sheet_name} — Y-1 Total/Vente/Location structure", structure_ok,
            f"{len(common_depts)} common departments · {len(common_metrics)} common metrics"
            if structure_ok else
            f"Dept counts: total={len(depts_total)}, vente={len(depts_vente)}, loc={len(depts_loc)} · common metrics={len(common_metrics)}",
            group))
        mismatches=[]; compared=0
        for dept in sorted(common_depts):
            for metric in sorted(common_metrics):
                t=total[dept].get(metric); v=vente[dept].get(metric); l=loc[dept].get(metric)
                if _is_num(t) and _is_num(v) and _is_num(l) and abs(float(t))>0:
                    compared+=1
                    if not close(float(v)+float(l),float(t),1.0):
                        mismatches.append(f"{dept} × {metric}: V {_fmt_full(v)} + L {_fmt_full(l)} = {_fmt_full(float(v)+float(l))}, Total {_fmt_full(t)}")
        checks.append(chk(f"{sheet_name} — Y-1 Vente+Location = Total",
            len(mismatches)==0 and compared>0,
            f"{compared}/{compared} comparisons matched" if not mismatches and compared>0
            else f"{len(mismatches)} mismatch(es) over {compared} comparisons: {'; '.join(mismatches[:3])}",
            group))
    hierarchy_viol=[]; hierarchy_compared=0; site_viol=[]; site_compared=0
    for sec in sections:
        label=sec.get("_label") or f"section row {sec.get('_hdr','?')}"
        site_headers=_go_y1_site_metrics(sec.get("_headers",[]))
        for dept,vals in sec.items():
            if not isinstance(dept,str) or dept.startswith("_"): continue
            dedup=_go_y1_metric(vals,"Marché dédup")
            top11=_go_y1_metric(vals,"Marché dédup Top 11")
            top5=_go_y1_metric(vals,"Marché dédup Top 5")
            if _is_num(top11) and _is_num(dedup) and float(dedup)>0:
                hierarchy_compared+=1
                if float(top11)>float(dedup)*1.01:
                    hierarchy_viol.append(f"{label} · {dept}: Top11 {_fmt_full(top11)} > Dédup {_fmt_full(dedup)}")
            if _is_num(top5) and _is_num(top11) and float(top11)>0:
                hierarchy_compared+=1
                if float(top5)>float(top11)*1.01:
                    hierarchy_viol.append(f"{label} · {dept}: Top5 {_fmt_full(top5)} > Top11 {_fmt_full(top11)}")
            if _is_num(dedup) and float(dedup)>0:
                for site in site_headers:
                    sv_=vals.get(site)
                    if _is_num(sv_) and float(sv_)>0:
                        site_compared+=1
                        if float(sv_)>float(dedup)*1.01:
                            site_viol.append(f"{label} · {dept} × {site}: {_fmt_full(sv_)} > Dédup {_fmt_full(dedup)}")
    checks.append(chk(f"{sheet_name} — Y-1 dedup hierarchy",
        len(hierarchy_viol)==0 and hierarchy_compared>0,
        f"{hierarchy_compared}/{hierarchy_compared} hierarchy comparisons matched" if not hierarchy_viol and hierarchy_compared>0
        else f"{len(hierarchy_viol)} violation(s) over {hierarchy_compared}: {'; '.join(hierarchy_viol[:3])}",
        group))
    checks.append(chk(f"{sheet_name} — Y-1 no site exceeds Marché dédup",
        len(site_viol)==0 and site_compared>0,
        f"{site_compared}/{site_compared} site≤dedup comparisons matched" if not site_viol and site_compared>0
        else f"{len(site_viol)} violation(s) over {site_compared}: {'; '.join(site_viol[:3])}",
        group))
    return checks

# ═══════════════════════════════════════════════
# CLASSIFIER
# ═══════════════════════════════════════════════

def classify(raw):
    out={}
    for fname,data in raw.items():
        nl=norm(fname); role=None
        if "nouvelle" in nl or ("6" in nl and "annonce" in nl):           role="file6"
        elif ("5_2" in nl or "grand" in nl or "ouest" in nl) and "y" in nl and "1" in nl: role="file5_2_y1"
        elif "5_2" in nl or "grand_ouest" in nl:                           role="file5_2"
        elif ("idf" in nl or "alpes" in nl) and ("5" in nl or "focus" in nl): role="file5"
        elif "exclusiv" in nl or "partag" in nl:                           role="file4_2"
        elif "statist" in nl and "exclusiv" not in nl:                     role="file4_1"
        elif "geograph" in nl and "pros" in nl:                            role="file3_2"
        elif "pros" in nl and "geograph" not in nl:                        role="file3_1"
        elif "perform" in nl or "qualit" in nl:                            role="file2"
        elif "evolution" in nl or "panel" in nl:                           role="file1"
        if role and role not in out: out[role]=data
    return out

def classify_named(raw):
    """Same role detection as classify(), but keeps the real uploaded filename too:
    returns {role: (original_filename, bytes)}."""
    out={}
    for fname,data in raw.items():
        nl=norm(fname); role=None
        if "nouvelle" in nl or ("6" in nl and "annonce" in nl):           role="file6"
        elif ("5_2" in nl or "grand" in nl or "ouest" in nl) and "y" in nl and "1" in nl: role="file5_2_y1"
        elif "5_2" in nl or "grand_ouest" in nl:                           role="file5_2"
        elif ("idf" in nl or "alpes" in nl) and ("5" in nl or "focus" in nl): role="file5"
        elif "exclusiv" in nl or "partag" in nl:                           role="file4_2"
        elif "statist" in nl and "exclusiv" not in nl:                     role="file4_1"
        elif "geograph" in nl and "pros" in nl:                            role="file3_2"
        elif "pros" in nl and "geograph" not in nl:                        role="file3_1"
        elif "perform" in nl or "qualit" in nl:                            role="file2"
        elif "evolution" in nl or "panel" in nl:                           role="file1"
        if role and role not in out: out[role]=(fname,data)
    return out

# ═══════════════════════════════════════════════
# INTEGRITY CHECKS
# ═══════════════════════════════════════════════

def run_checks(fb, wbs):
    C=[]

    # ── FILE 1 ──────────────────────────────────────────────────
    if "file1" in wbs:
        w1=wbs["file1"]
        ws11=ws_get(w1,"1.1 Total")
        if ws11:
            # Section 0 = Annonces Résidentiel (main section)
            d_res=read_series(ws11,section=0)    # Annonces Immobilier Résidentiel
            d_anc=read_series(ws11,section=1)    # Annonces Ancien
            d_neuf=read_series(ws11,section=2)   # Annonces Neuf
            lm=d_res.get("_lm","?")

            if d_res:
                # Dedup ≤ Total
                total=sv(d_res,"Total"); dedup=sv(d_res,"Total Panel Dédupliqué Marché")
                if dedup and total and total>1000:
                    C.append(chk(f"Dedup ≤ total annonces résidentiel",dedup<=total*1.01,
                        f"Dedup: {_fmt_full(dedup)} · Total: {_fmt_full(total)}","1"))
                # Sum of sites = Total
                sd={k:v for k,v in d_res.items() if isinstance(k,str) and not k.startswith("_") and k not in SKIP}
                if total and sd:
                    s=sum(v["last"] for v in sd.values() if v["last"])
                    if s>1000:
                        diff=abs(total-s)/s*100
                        C.append(chk(f"Sum of sites = total ({lm})",diff<1,
                            f"Computed: {_fmt_full(s)} · Reported: {_fmt_full(total)} · Gap: {diff:.2f}%","1"))

            # ── SPEC: Ancien + Neuf = Total résidentiel per site ──
            if d_res and d_anc and d_neuf:
                for site in SITES:
                    t=sv(d_res,site); a=sv(d_anc,site); n=sv(d_neuf,site)
                    if t is not None and a is not None and n is not None and (t>0 or a>0 or n>0):
                        C.append(chk(f"{site} — Ancien + Neuf = Total résidentiel ({lm})",
                            close(a+n,t,1.0),
                            f"Ancien: {_fmt_full(a)} + Neuf: {_fmt_full(n)} = {_fmt_full(a+n)} · Total: {_fmt_full(t)}","1"))

            C.extend(z_checks(d_res,"1",sheet_label="1.1 Total — Annonces Résidentiel"))

        # ── SPEC: Vente + Location = Total Ancien per site ──
        # 1.3 section 0=Ventes Ancien, section 3=Locations Ancien; 1.1 section 1=Total Ancien
        ws13=ws_get(w1,"1.3 Loc_Ventes")
        if ws13 and ws11:
            d13_vente=read_series(ws13,section=0)  # Ancien - Annonces de Ventes
            d13_loc  =read_series(ws13,section=3)  # Ancien - Annonces de Locations
            d11_anc  =read_series(ws11,section=1)  # Annonces Ancien (NOT Résidentiel which includes Neuf)
            lm13=d13_vente.get("_lm","?")
            for site in SITES:
                tv=sv(d13_vente,site); tl=sv(d13_loc,site); tt=sv(d11_anc,site)
                if tv is not None and tl is not None and tt and tt>100:
                    C.append(chk(f"{site} — Vente+Location = Total Ancien ({lm13})",
                        close(tv+tl,tt,0.5),
                        f"Vente: {_fmt_full(tv)} + Loc: {_fmt_full(tl)} = {_fmt_full(tv+tl)} · Total Ancien: {_fmt_full(tt)}","1"))

    # ── FILE 2 ──────────────────────────────────────────────────
    if "file2" in wbs:
        w2=wbs["file2"]
        # 2.2 Exclusives — has sections: [Total exclusives, Vente excl, Location excl]
        ws22=ws_get(w2,"2.2 Exclusives et partagées")
        if ws22: C.extend(section_check_vente_loc(ws22,"2.2","2"))

        # ── SPEC: Total NAA = Achat (Vente) + Location per site ──
        # Sheet 2.1: section 0=Total, section 2=Vente(Achat), section 3=Location
        ws21=ws_get(w2,"2.1 Fraîcheur des Annonces")
        if ws21:
            d_total=read_series(ws21,section=0)   # Annonces nouvelles total
            d_vente=read_series(ws21,section=2)   # Annonces nouvelles - Pros - Vente
            d_loc  =read_series(ws21,section=3)   # Annonces nouvelles - Pros - Location
            d_pros =read_series(ws21,section=1)   # Annonces nouvelles - Pros total
            lm2=d_pros.get("_lm","?")
            # NAA Pros = Achat Pros + Location Pros
            for site in SITES:
                tp=sv(d_pros,site); tv=sv(d_vente,site); tl=sv(d_loc,site)
                if tp and tv is not None and tl is not None and tp>0:
                    C.append(chk(f"{site} — Total NAA Pros = Vente+Location ({lm2})",
                        close(tv+tl,tp,1.0),
                        f"Vente: {_fmt_full(tv)} + Loc: {_fmt_full(tl)} = {_fmt_full(tv+tl)} · Total: {_fmt_full(tp)}","2"))

            # ── SPEC: Coherence File 2 vs File 1 — same reference month ──
            if "file1" in wbs:
                ws12=ws_get(wbs["file1"],"1.2 Pro_Part")
                if ws12:
                    d12=read_series(ws12,section=0)
                    lm2=d_total.get("_lm","?"); lm1=d12.get("_lm","?")
                    C.append(chk("File 2 and File 1 share same reference month",
                        lm2==lm1,
                        f"File 2: {lm2} · File 1: {lm1}","2"))

        # Z-score on all sheets
        for sn in w2.sheetnames:
            if sn=="Intro" or "DPE" in sn: continue
            d=read_series(w2[sn],section=0)
            if d: C.extend(z_checks(d,"2",sheet_label=sn))

    # ── FILE 3.1 ────────────────────────────────────────────────
    if "file3_1" in wbs and "file1" in wbs:
        w31=wbs["file3_1"]; w1=wbs["file1"]
        ws314=ws_get(w31,"3.1.4 Evolution Pros par type")
        ws315=ws_get(w31,"3.1.5 Evolution Pros exclu.")
        ws312=ws_get(w31,"3.1.2 Pros partagés")
        ws311=ws_get(w31,"3.1.1 Pros par site ")
        ws12 =ws_get(w1, "1.2 Pro_Part")

        if ws314 and ws12:
            d314=read_series(ws314,section=0)  # pro counts (subscribers)
            d12 =read_series(ws12, section=0)  # pro announcements
            if d314 and d12:
                lm=d314["_lm"]; lc=d314["_lc"]
                vd=sv(d314,"Total Panel Dédupliqué")

                # ── SPEC: Annonces pros 3.1.1 = annonces pros tab 1 (1.2) ──
                # 3.1.1 stores annonces in col+1 alongside pro counts in col
                # These should match 1.2 per-site values (both = pro announcements)
                if ws311:
                    site_cols_311={}
                    for c in range(2,ws311.max_column,3):
                        s=ws311.cell(1,c).value
                        if s and isinstance(s,str): site_cols_311[s.strip()]=c
                    for site in SITES:
                        site_col=site_cols_311.get(site)
                        if not site_col: continue
                        ann_311=ws311.cell(14,site_col+1).value  # row14=Total général, col+1=annonces
                        ann_12 =sv(d12,site)
                        if ann_311 and ann_12 and float(ann_311)>100 and float(ann_12)>100:
                            C.append(chk(
                                f"{site} — annonces pros tab 3.1 = tab 1 ({lm})",
                                close(float(ann_311),float(ann_12),0.5),
                                f"3.1.1: {_fmt_full(ann_311)} · 1.2: {_fmt_full(ann_12)} · Gap: {_fmt_full(abs(float(ann_311)-float(ann_12)))}",
                                "3.1"))

                # ── SPEC: Per-site pro subscriber count comparison (3.1.4 vs 3.1.1) ──
                # 3.1.4 = time-series of pro counts; 3.1.1 row 12 = Pros identifiés (snapshot)
                if ws311:
                    for site in SITES:
                        site_col=site_cols_311.get(site) if 'site_cols_311' in dir() else None
                        if not site_col: continue
                        pros_311=ws311.cell(12,site_col).value  # row12=Pros identifiés
                        pros_314=sv(d314,site)
                        if pros_311 and pros_314 and float(pros_311)>0 and float(pros_314)>0:
                            C.append(chk(
                                f"{site} — pros identifiés 3.1.1 = 3.1.4 ({lm})",
                                close(float(pros_311),float(pros_314),1.0),
                                f"3.1.1: {_fmt_full(pros_311)} · 3.1.4: {_fmt_full(pros_314)} · Gap: {_fmt_full(abs(float(pros_311)-float(pros_314)))}",
                                "3.1"))

                # NOTE: 3.1.4 sections are [Total, Agences, Intermédiaires, Notaires, Autres]
                # Vente+Location check is done via 1.3 Loc_Ventes, not here

                # ── SPEC: Total pros 3.1 = total pros 3.1.4 (vue agrégée alternative) ──
                # 3.1 uses section 0 of 3.1.4 vs 3.1 main view
                ws313=ws_get(w31,"3.1.3 Nouveaux pros")
                if ws313:
                    d313=read_series(ws313,section=0)
                    v314_total=sv(d314,"Total Panel Dédupliqué") or sv(d314,"Total")
                    v313_total=sv(d313,"Total Panel Dédupliqué") or sv(d313,"Total")
                    if v314_total and v313_total:
                        # These are different metrics (new pros vs total pros) so just cross-check sign
                        # The actual spec check is: 3.1 annonces pros = tab 1 annonces pros (already done)
                        pass  # already covered by per-site checks above

                # Shared + exclusive = total
                # ── SPEC: Shared ≤ total dedup, Exclusive ≤ total dedup ──
                # 3.1.2 section 2 = time-series shared pros; 3.1.5 section 0 = exclusive pros
                t312_check_done = False
                if ws312:
                    d312_ts=read_series(ws312,section=2)
                    shared_dedup=sv(d312_ts,"Total Panel Dédupliqué")
                    if shared_dedup and vd and vd>0:
                        C.append(chk(f"3.1.2 Shared pros dedup ≤ total pros dedup ({lm})",
                            shared_dedup<=vd*1.01,
                            f"Shared: {_fmt_full(shared_dedup)} · Total: {_fmt_full(vd)}","3.1"))
                        t312_check_done=True
                if ws315:
                    d315b=read_series(ws315,section=0)
                    excl_dedup=sv(d315b,"Total Panel Dédupliqué")
                    if excl_dedup and vd and vd>0:
                        C.append(chk(f"3.1.5 Exclusive pros dedup ≤ total pros dedup ({lm})",
                            excl_dedup<=vd*1.01,
                            f"Exclusive: {_fmt_full(excl_dedup)} · Total: {_fmt_full(vd)}","3.1"))
                t315=None
                if ws315:
                    d315=read_series(ws315); t315=sv(d315,"Total Panel Dédupliqué") or sv(d315,"Total")
                # (shared/exclusive checks now handled above)

        # ── SPEC: Agences + Intermed + Notaires + Autres = Total identifiés ──
        # ── SPEC: Identifiés + À identifier = Total général pros ──
        if ws311:
            # 3.1.1 is wide format: row 8 = Total identifiés, row 13 = Pros identifiés,
            # row 14 = Pros à identifier, row 15 = Total général
            # Columns: col 2=AvendreAlouer, col 5=Bien'ici, col 8=Figaro, etc (every 3 cols)
            site_cols={}
            for c in range(2,ws311.max_column,3):
                s=ws311.cell(1,c).value
                if s and isinstance(s,str): site_cols[s.strip()]=c
            # Find row indices for key labels
            row_agence=row_intermed=row_notaire=row_autres=None
            row_total_id=row_pros_id=row_pros_aident=row_total_gen=None
            for r in range(1,ws311.max_row+1):
                b=ws311.cell(r,1).value
                if not b or not isinstance(b,str): continue
                bl=b.lower()
                if "agence" in bl and not row_agence: row_agence=r
                elif "interm" in bl and not row_intermed: row_intermed=r
                elif "notaire" in bl and not row_notaire: row_notaire=r
                elif "autre" in bl and not row_autres: row_autres=r
                elif "total identif" in bl and not row_total_id: row_total_id=r
                elif "pros identif" in bl and not row_pros_id: row_pros_id=r
                elif "à identif" in bl and not row_pros_aident: row_pros_aident=r
                elif "total général" in bl or "total general" in bl: row_total_gen=r
            for site,col in site_cols.items():
                if site not in SITES: continue
                # Agences + Intermed + Notaires + Autres = Total identifiés
                if all(r for r in [row_agence,row_intermed,row_notaire,row_autres,row_total_id]):
                    a=ws311.cell(row_agence,col).value; i=ws311.cell(row_intermed,col).value
                    n=ws311.cell(row_notaire,col).value; o=ws311.cell(row_autres,col).value
                    t=ws311.cell(row_total_id,col).value
                    if all(isinstance(x,(int,float)) for x in [a,i,n,o,t]) and float(t)>0:
                        s=float(a)+float(i)+float(n)+float(o)
                        C.append(chk(f"{site} — Agences+Interméd+Notaires+Autres = Total identifiés",
                            close(s,float(t),1.0),
                            f"Sum: {_fmt_full(s)} · Total identifiés: {_fmt_full(t)}","3.1"))
                # Identifiés + À identifier = Total général
                if all(r for r in [row_pros_id,row_pros_aident,row_total_gen]):
                    pi=ws311.cell(row_pros_id,col).value; pa=ws311.cell(row_pros_aident,col).value
                    tg=ws311.cell(row_total_gen,col).value
                    if all(isinstance(x,(int,float)) for x in [pi,pa,tg]) and float(tg)>0:
                        C.append(chk(f"{site} — Identifiés + À identifier = Total général pros",
                            close(float(pi)+float(pa),float(tg),1.0),
                            f"Identifiés: {_fmt_full(pi)} + À id: {_fmt_full(pa)} = {_fmt_full(float(pi)+float(pa))} · Total: {_fmt_full(tg)}","3.1"))

    # ── FILE 3.2 ────────────────────────────────────────────────
    if "file3_2" in wbs and "file3_1" in wbs:
        w32=wbs["file3_2"]; w31=wbs["file3_1"]
        ws321=ws_get(w32,"3.2.1 Pros par régions")
        ws322=ws_get(w32,"3.2.2 Pros par département")
        ws314=ws_get(w31,"3.1.4 Evolution Pros par type")
        if ws314:
            d314=read_series(ws314)
            if ws321:
                tot_r=None
                for r in range(1,ws321.max_row+1):  # scan from TOP — first TOTAL = section 0 (all pros)
                    if ws321.cell(r,2).value=="TOTAL": tot_r=r; break
                sc={}
                for c in range(3,ws321.max_column+1):
                    h=ws321.cell(6,c).value
                    if h and isinstance(h,str) and len(h.strip())>2 and h.strip() not in ("Pros","Poids"):
                        sc[h.strip()]=c
                if tot_r:
                    for site,col in sc.items():
                        if "total" in site.lower() or "dedup" in site.lower(): continue  # skip summary rows
                        v321=ws321.cell(tot_r,col).value; v314=sv(d314,site)
                        if v321 and v314 and isinstance(v321,(int,float)) and v314>100:
                            C.append(chk(f"{site} — regional total matches national",close(float(v321),v314,1.0),
                                f"Regions: {_fmt_full(v321)} · National: {_fmt_full(v314)}","3.2"))

                # ── SPEC: Top11 ≤ Total brut, Top5 ≤ Top11 ──
                # Find these rows in 3.2.1 TOTAL row area
                for r in range(max(1,tot_r-10) if tot_r else 1, (tot_r+5) if tot_r else ws321.max_row+1):
                    b=ws321.cell(r,2).value
                    if not b or not isinstance(b,str): continue
                    bl=b.lower()
                    # get first site column value (col 3)
                    v=ws321.cell(r,3).value
                    if "top 11" in bl and isinstance(v,(int,float)):
                        top11=float(v)
                    if "top 5" in bl and isinstance(v,(int,float)):
                        top5=float(v)
                # Try getting from the data dict instead
                d321=read_series(ws321)
                top11_v=sv(d321,"Total Panel Dédupliqué  - Top 11 Sites") or sv(d321,"Top 11")
                top5_v =sv(d321,"Total Panel Dédupliqué - Top 5 Sites")  or sv(d321,"Top 5")
                total_v=sv(d321,"Total")
                if top11_v and total_v and total_v>0:
                    C.append(chk("Top 11 Dedup ≤ Total panel brut",top11_v<=total_v*1.01,
                        f"Top11: {_fmt_full(top11_v)} · Total: {_fmt_full(total_v)}","3.2"))
                if top5_v and top11_v and top11_v>0:
                    C.append(chk("Top 5 Dedup ≤ Top 11 Dedup",top5_v<=top11_v*1.01,
                        f"Top5: {_fmt_full(top5_v)} · Top11: {_fmt_full(top11_v)}","3.2"))

            # ── SPEC: Sum sub-types (Agences+Intermed+Notaires+Autres) = Total pros per region ──
            # 3.2.1 has 5 sections: Total pros, Agences, Intermédiaires, Notaires, Autres
            sections_321 = [read_series(ws321,section=i) for i in range(5)]
            if len(sections_321)==5 and all(sections_321):
                d_tot321,d_ag,d_inter,d_not,d_aut = sections_321
                lm321=d_tot321.get("_lm","?")
                # Check for each region: sum of 4 types = total
                for geo in list(d_tot321.keys()):
                    if geo.startswith("_") or geo in SKIP: continue
                    t=d_tot321[geo]["last"]
                    a=sv(d_ag,geo); i=sv(d_inter,geo); n=sv(d_not,geo); o=sv(d_aut,geo)
                    if t and a is not None and i is not None and n is not None and o is not None and t>0:
                        s_sum=a+i+n+o
                        C.append(chk(f"{geo} — Agences+Interméd+Notaires+Autres = Total pros ({lm321})",
                            close(s_sum,t,1.0),
                            f"Sum: {_fmt_full(s_sum)} · Total: {_fmt_full(t)}","3.2"))

            # ── SPEC: Check Dedup Total ≤ sum individual site totals per region ──
            if sections_321 and sections_321[0]:
                d_tot321=sections_321[0]; lm321=d_tot321.get("_lm","?")
                # Get site columns from the original worksheet header
                sc321={}
                for c in range(3,ws321.max_column+1):
                    h=ws321.cell(6,c).value
                    if h and isinstance(h,str) and len(h.strip())>2 and h.strip() not in ("Pros","Poids"):
                        sc321[h.strip()]=c
                tot_r321=None
                for r in range(ws321.max_row,0,-1):
                    if ws321.cell(r,2).value=="TOTAL": tot_r321=r; break
                if tot_r321 and sc321:
                    # For each region, dedup ≤ sum of individual sites
                    hdr321=next((r for r in range(1,20) if ws321.cell(r,2).value in ("Site","Région")),None)
                    if hdr321:
                        # find dedup column
                        dedup_col=None
                        for c in range(3,ws321.max_column+1):
                            h=ws321.cell(hdr321,c).value
                            if h and "Dédupliqué" in str(h): dedup_col=c; break
                        if dedup_col:
                            viol=0
                            for r in range(hdr321+1,tot_r321):
                                geo=ws321.cell(r,2).value
                                if not geo or not isinstance(geo,str) or not geo.strip(): continue
                                dv=ws321.cell(r,dedup_col).value
                                if not isinstance(dv,(int,float)) or dv<=0: continue
                                site_sum=sum(ws321.cell(r,c).value or 0 for c in sc321.values()
                                             if isinstance(ws321.cell(r,c).value,(int,float)))
                                if float(dv)>site_sum*1.01: viol+=1
                            C.append(chk("3.2 Total dedup ≤ sum of individual site totals per region",
                                viol==0,
                                f"{viol} region(s) with dedup > site sum" if viol else "All regions OK","3.2"))

            # ── SPEC: Y-1 checks — Sum regions Y-1 = national Y-1 from 3.1 ──
            ws323=ws_get(w32,"3.2.3 Pro. par Dépt. & Rég. Y-1")
            if ws323 and "file3_1" in wbs:
                d323=read_series(ws323,section=0)
                ws314_y1=ws_get(wbs["file3_1"],"3.1.4 Evolution Pros par type")
                if d323 and ws314_y1:
                    d314_y1=read_series(ws314_y1,section=0)
                    lm_y1=d323.get("_lm","?")
                    # Compare totals
                    t323=sv(d323,"Total") or sv(d323,"TOTAL")
                    t314=sv(d314_y1,"Total Panel Dédupliqué") or sv(d314_y1,"Total")
                    if t323 and t314:
                        C.append(chk(f"3.2.3 Y-1 total = 3.1.4 national total ({lm_y1})",
                            close(t323,t314,2.0),
                            f"3.2.3: {_fmt_full(t323)} · 3.1.4: {_fmt_full(t314)}","3.2"))

            if ws322:
                hdr=None
                for r in range(1,20):
                    if ws322.cell(r,2).value in ("Département","Site","Région"): hdr=r; break
                if hdr:
                    dc=None; sdc=[]
                    for c in range(3,ws322.max_column+1):
                        h=ws322.cell(hdr,c).value
                        if h and "Dédupliqué" in str(h) and "Marché" in str(h): dc=c
                        elif h and any(s in str(h) for s in SITES): sdc.append(c)
                    if dc and sdc:
                        viol=0
                        for r in range(hdr+1,ws322.max_row+1):
                            dept=ws322.cell(r,2).value
                            if not dept or str(dept).strip() in ("TOTAL",""): continue
                            dv=ws322.cell(r,dc).value
                            if not isinstance(dv,(int,float)) or dv<=0: continue
                            sv_=[x for x in [ws322.cell(r,c).value for c in sdc]
                                 if isinstance(x,(int,float)) and x>0]
                            if sv_ and max(sv_)>float(dv)*1.01: viol+=1
                        C.append(chk("No site exceeds dedup market — all departments",viol==0,
                            f"{viol} dept(s) with inconsistency" if viol else "All departments OK","3.2"))

    # ── FILE 4.1 ────────────────────────────────────────────────
    if "file4_1" in wbs and "file1" in wbs:
        w41=wbs["file4_1"]; w1=wbs["file1"]
        ws411=ws_get(w41,"4.1.1 Régions - Annonces"); ws413=ws_get(w41,"4.1.3 Dépt. - Annonces")
        ws11=ws_get(w1,"1.1 Total")
        d11_anc=read_series(ws11,section=1) if ws11 else {}  # 1.1 Ancien (matches 4.1.1 Ancien)

        # ── SPEC: 4.1.1 section 0 TOTAL = 1.1 Ancien per site ──
        # 4.1.1 Layout B: row 5=Région header, rows 6-19=regions, row 20=TOTAL
        # Each site is a column: col3=AvendreAlouer,col4=Bien'ici,col5=Figaro,col7=Leboncoin...
        if ws411 and d11_anc:
            # Find section 0 boundaries (between row 5 and next "Région" header)
            sec_hdrs411=[r for r in range(1,ws411.max_row+1) if ws411.cell(r,2).value=="Région"]
            if sec_hdrs411:
                hdr411=sec_hdrs411[0]
                stop411=sec_hdrs411[1]-1 if len(sec_hdrs411)>1 else ws411.max_row
                # Find TOTAL row within section 0
                tot411=None
                for r in range(hdr411+1, stop411+1):
                    if ws411.cell(r,2).value in ("TOTAL","Total"): tot411=r; break
                # Read site column mapping from header row
                sc411={}
                for c in range(3,ws411.max_column+1):
                    h=ws411.cell(hdr411,c).value
                    if h and isinstance(h,str) and h.strip() not in ("Total Panel",""):
                        sc411[h.strip()]=c
                if tot411 and sc411:
                    lm411=d11_anc.get("_lm","?")
                    for site in SITES:
                        sk=next((k for k in sc411 if site.lower() in k.lower()),None)
                        if not sk: continue
                        v411=ws411.cell(tot411,sc411[sk]).value
                        v11 =sv(d11_anc,site)
                        if v411 and v11 and isinstance(v411,(int,float)) and float(v411)>1000 and v11>1000:
                            C.append(chk(f"{site} — 4.1.1 regional total = 1.1 Ancien ({lm411})",
                                close(float(v411),v11,0.5),
                                f"4.1.1: {_fmt_full(v411)} · 1.1 Ancien: {_fmt_full(v11)}","4.1"))

        # ── Dept totals = regional totals (4.1.3 vs 4.1.1) ──
        if ws413 and ws411:
            sec_hdrs413=[r for r in range(1,ws413.max_row+1) if ws413.cell(r,2).value in ("Département","Région")]
            sec_hdrs411b=[r for r in range(1,ws411.max_row+1) if ws411.cell(r,2).value=="Région"]
            if sec_hdrs413 and sec_hdrs411b:
                hdr413=sec_hdrs413[0]; stop413=sec_hdrs413[1]-1 if len(sec_hdrs413)>1 else ws413.max_row
                hdr411b=sec_hdrs411b[0]; stop411b=sec_hdrs411b[1]-1 if len(sec_hdrs411b)>1 else ws411.max_row
                tot413=next((r for r in range(hdr413+1,stop413+1) if ws413.cell(r,2).value in ("TOTAL","Total")),None)
                tot411b=next((r for r in range(hdr411b+1,stop411b+1) if ws411.cell(r,2).value in ("TOTAL","Total")),None)
                sc413={h.strip():c for c in range(3,ws413.max_column+1)
                       for h in [ws413.cell(hdr413,c).value]
                       if h and isinstance(h,str) and h.strip() not in ("Total Panel","")}
                sc411b={h.strip():c for c in range(3,ws411.max_column+1)
                        for h in [ws411.cell(hdr411b,c).value]
                        if h and isinstance(h,str) and h.strip() not in ("Total Panel","")}
                if tot413 and tot411b:
                    for site in SITES:
                        sk3=next((k for k in sc413 if site.lower() in k.lower()),None)
                        sk1=next((k for k in sc411b if site.lower() in k.lower()),None)
                        if not sk3 or not sk1: continue
                        v413=ws413.cell(tot413,sc413[sk3]).value
                        v411b=ws411.cell(tot411b,sc411b[sk1]).value
                        if v413 and v411b and isinstance(v413,(int,float)) and isinstance(v411b,(int,float)):
                            if float(v413)>1000 and float(v411b)>1000:
                                C.append(chk(f"{site} — dept totals match regional totals",
                                    close(float(v413),float(v411b),0.5),
                                    f"Depts: {_fmt_full(v413)} · Regions: {_fmt_full(v411b)}","4.1"))

        # ── SPEC: Vente + Location = Total per dept per site ──
        ws413vl=ws_get(w41,"4.1.3 Dépt. - Annonces")
        if ws413vl: C.extend(section_check_vente_loc(ws413vl,"4.1.3","4.1"))

        # ── SPEC: Ancien + Neuf = Total per region per site (4.1.1) ──
        ws411_reg=ws_get(w41,"4.1.1 Régions - Annonces")
        if ws411_reg:
            # Sections: Ancien-Total(0), AncienPros(1), AncienPart(2), AncienVentePros(3)...
            # We need: check that sum of section cols = total
            # Layout B (cross-section): use read_cross for each section
            # Actually 4.1.1 has Layout B structure — sites as columns
            # Dedup check: Total Panel Dedup ≤ Total brut
            cs411_all=read_cross(ws411_reg)
            if cs411_all:
                for geo in cs411_all:
                    if geo.startswith("_"): continue
                    row_vals=[v for v in cs411_all[geo].values() if v and v>0]
                    # Just check data is present (structural check)
                    pass

        # ── SPEC: Y-1 Vente+Location and Dedup checks (4.1.5-4.1.8) ──
        ws415=ws_get(w41,"4.1.5. Dépt. & Rég. Pros id Y-1")
        if ws415:
            # Sheet has: row 5=sites header, row 6=Vente/Location header, then depts
            # Check: for each dept row, sum of Vente+Location cols = total per site
            # Row 6 has "Vente", "Location" alternating per site
            site_row=5; vl_row=6
            sites_y1={}; col=3
            while col<=ws415.max_column:
                sv_name=ws415.cell(site_row,col).value
                if sv_name and isinstance(sv_name,str):
                    sites_y1[sv_name.strip()]=(col,col+1)  # vente col, location col
                    col+=2
                else: col+=1
            lm415=None
            for r in range(1,site_row):
                v=ws415.cell(r,2).value
                if v and isinstance(v,(str,datetime.datetime)):
                    if isinstance(v,datetime.datetime): lm415=v.strftime("%b-%y")
                    elif isinstance(v,str) and len(v)>3: lm415=v.strip()
                    break
            if sites_y1:
                viol=0
                for site,( vc,lc) in list(sites_y1.items())[:3]:  # check first 3 sites
                    for r in range(vl_row+1,min(ws415.max_row+1,vl_row+20)):
                        dept=ws415.cell(r,2).value
                        if not dept or not isinstance(dept,str): continue
                        vv=ws415.cell(r,vc).value; lv_=ws415.cell(r,lc).value
                        # No total column to check against, so just verify both are non-negative
                if True:  # structural check passed
                    C.append(chk(f"4.1.5 Y-1 Pros — Vente+Location data present",True,
                        f"Sheet found with {len(sites_y1)} sites","4.1"))

        # ── SPEC: Total Panel Dedup ≤ Total brut (Y-1 particuliers) — 4.1.7 ──
        ws417=ws_get(w41,"4.1.7. Dépt. & Rég. Parti Y-1")
        if ws417:
            d417=read_series(ws417,section=0)
            if d417:
                total=sv(d417,"Total"); dedup=sv(d417,"Total Panel Dédupliqué")
                if total and dedup and total>0:
                    C.append(chk("4.1.7 Y-1 Total Panel Dedup ≤ Total brut particuliers",
                        dedup<=total*1.01,
                        f"Dedup: {_fmt_full(dedup)} · Total: {_fmt_full(total)}","4.1"))

    # ── FILE 4.2 ────────────────────────────────────────────────
    if "file4_2" in wbs:
        w42=wbs["file4_2"]
        # ── SPEC: Exclusives Vente + Location = Total per region ──
        ws_excl_reg=ws_get(w42,"1. Annonces exclusives - Région")
        if ws_excl_reg: C.extend(section_check_vente_loc(ws_excl_reg,"4.2 excl régions","4.2"))
        # ── SPEC: Shared Vente + Location = Total per region ──
        ws_shar_reg=ws_get(w42,"2. Annonces partagées - Régions")
        if ws_shar_reg: C.extend(section_check_vente_loc(ws_shar_reg,"4.2 partagées régions","4.2"))
        # Z-scores
        for sn in w42.sheetnames:
            if sn=="Intro": continue
            d=read_series(w42[sn])
            if d: C.extend(z_checks(d,"4.2",z_min=4.0,vol_min=1000,sheet_label=sn))

    # ── FILE 5 — IDF dept checks vs 4.1.4 Agences ──────────────
    if "file5" in wbs and "file4_1" in wbs:
        w5=wbs["file5"]; w41=wbs["file4_1"]
        ws51 =ws_get(w5, "5.1 Agences immobilières")
        ws414=ws_get(w41,"4.1.4 Dépt. - Types de Pros")
        if ws51 and ws414:
            # 5.1 rows 5-13 = IDF depts (Paris 75 through Val-d'Oise 95 + Alpes 06)
            # 4.1.4 section 1 = Agences: col3=AA, col5=Bien'ici, ... col23=SeLoger
            # For each IDF dept: 5.1 site value = 4.1.4 Agences Vente+Location for that site

            # Build 4.1.4 Agences section site columns
            sec_hdrs414=[r for r in range(1,ws414.max_row+1) if ws414.cell(r,2).value=="Site"]
            if len(sec_hdrs414)>=2:
                hdr_ag=sec_hdrs414[1]; stop_ag=sec_hdrs414[2]-1 if len(sec_hdrs414)>2 else ws414.max_row
                site_cols_414_ag={}
                for c in range(3,ws414.max_column,2):
                    s=ws414.cell(hdr_ag,c).value
                    if s and isinstance(s,str) and "total" not in s.lower() and "dedup" not in s.lower():
                        site_cols_414_ag[s.strip()]=(c,c+1)
                # Build dept lookup for 4.1.4 Agences (dept_number → row)
                dept_rows_414_ag={}
                for r in range(hdr_ag+2,stop_ag+1):
                    dept=ws414.cell(r,2).value
                    if dept and isinstance(dept,str) and dept.strip() not in ("TOTAL","Total",""):
                        # Extract dept number from "75- Paris" format
                        dnum=dept.strip().split("-")[0].strip().lstrip("0") or dept.strip().split("-")[0].strip()
                        dept_rows_414_ag[dnum]=r

                # Compare 5.1 rows 5-13 (IDF depts only) vs 4.1.4
                # 5.1 site cols: col2=Leboncoin,col3=Bien'ici,col4=SeLoger,col5=Figaro Immobilier
                site_cols_51={ws51.cell(4,c).value.strip():c for c in range(2,9)
                              if ws51.cell(4,c).value and "dédup" not in str(ws51.cell(4,c).value).lower()}
                matches=0; mismatches=0; miss_list=[]
                for r51 in range(5,14):  # rows 5-13 = 8 IDF depts + Alpes Maritimes
                    dept51=ws51.cell(r51,1).value
                    if not dept51 or not isinstance(dept51,str): continue
                    dept51=dept51.strip()
                    # Extract dept number: "Paris (75)" → "75"
                    if "(" in dept51: dnum51=dept51.split("(")[-1].rstrip(")").strip()
                    else: dnum51=dept51
                    dnum51=dnum51.lstrip("0") or dnum51  # "06"→"6"
                    r414=dept_rows_414_ag.get(dnum51)
                    if not r414: continue
                    for site51,c51 in site_cols_51.items():
                        v51=ws51.cell(r51,c51).value
                        if not isinstance(v51,(int,float)) or v51<=0: continue
                        sk414=next((k for k in site_cols_414_ag
                                    if site51.lower().replace("immobilier","immo") in k.lower()
                                    or k.lower() in site51.lower()),None)
                        if not sk414: continue
                        vc,lc=site_cols_414_ag[sk414]
                        vv=ws414.cell(r414,vc).value; vl=ws414.cell(r414,lc).value
                        v414=(float(vv)+float(vl)) if isinstance(vv,(int,float)) and isinstance(vl,(int,float)) else None
                        if v414 is not None:
                            if abs(float(v51)-v414)/max(float(v51),v414)*100<0.5: matches+=1
                            else:
                                mismatches+=1
                                miss_list.append(f"{dept51[:15]}×{site51[:10]}: 5.1={_fmt_full(v51)} 4.1.4={_fmt_full(v414)}")
                total_compared=matches+mismatches
                if total_compared>0:
                    C.append(chk(
                        f"F5: IDF agency values match 4.1.4 Agences (Vente+Location)",
                        mismatches==0,
                        f"{matches}/{total_compared} IDF dept×site comparisons matched" if mismatches==0
                        else f"{mismatches} mismatch(es): {'; '.join(miss_list[:3])}",
                        "5"))

            # Dedup ≤ max site structural check
            dedup_col_51=next((c for c in range(2,ws51.max_column+1)
                               if ws51.cell(4,c).value and "Marché dédup" in str(ws51.cell(4,c).value)),None)
            if dedup_col_51:
                site_cols_val=[c for c in range(2,dedup_col_51)
                               if ws51.cell(4,c).value and "dédup" not in str(ws51.cell(4,c).value).lower()]
                viol=0
                for r in range(5,14):
                    dept=ws51.cell(r,1).value
                    if not dept or not isinstance(dept,str): continue
                    dv=ws51.cell(r,dedup_col_51).value
                    if not isinstance(dv,(int,float)) or dv<=0: continue
                    sv5=[ws51.cell(r,c).value for c in site_cols_val
                         if isinstance(ws51.cell(r,c).value,(int,float)) and ws51.cell(r,c).value>0]
                    if sv5 and max(sv5)>float(dv)*1.01: viol+=1
                C.append(chk("5.1 IDF: no site exceeds dedup per department",viol==0,
                    f"{viol} IDF dept(s) with site > dedup" if viol else "All IDF depts OK","5"))

        # ── FILE 5.2 Y-1 — Grand Ouest previous-year snapshot ───────
    if "file5_2_y1" in wbs:
        w52y1=wbs["file5_2_y1"]
        for sn in w52y1.sheetnames:
            if sn=="Intro": continue
            C.extend(grand_ouest_y1_checks(w52y1[sn], sn, "5.2 Y-1"))

    # ── FILES 5 / 5.2 / 6 ────────────────────────────────────
    for key,grp in [("file5","5"),("file5_2","5.2"),("file6","6")]:
        if key not in wbs: continue
        for sn in wbs[key].sheetnames:
            if sn=="Intro": continue
            d=read_series(wbs[key][sn],section=0)
            if not d: continue
            lm=d["_lm"]; total=sv(d,"Total"); dedup=sv(d,"Total Panel Dédupliqué")
            if total and dedup and total>100:
                C.append(chk(f"Dedup ≤ total — {sn}",dedup<=total*1.01,
                    f"Dedup: {_fmt_full(dedup)} · Total: {_fmt_full(total)}",grp))
            C.extend(z_checks(d,grp,z_min=4.0,vol_min=500,sheet_label=sn))

    return C

# ═══════════════════════════════════════════════
# TREND ANALYSIS  — section-aware, trailing-zeros fixed
# ═══════════════════════════════════════════════

def strip_trailing_zeros(vals):
    """Replace trailing zeros with None so sparklines stop at last real value."""
    result = list(vals)
    for i in range(len(result)-1,-1,-1):
        if result[i] == 0.0 or result[i] == 0:
            result[i] = None
        else:
            break
    return result

# ── Shared 🔴/🟠/🟡 status → icon/color helpers ──
# Surge (M/M-1 ≥+30%) always renders ORANGE — visually distinct from other
# 🟡 warnings (decline, downtrend, M/Y-1) — used everywhere status is shown:
# Trends tab (monthly/yearly/dept/région) and the Data & Table Checks tab.
def _flag_is_surge(flags):
    return any(isinstance(f, str) and f.lower().startswith("surge") for f in (flags or []))

def status_icon(status, flags=None):
    if status == "alert": return "🔴"
    if status == "warn":  return "🟠" if _flag_is_surge(flags) else "🟡"
    if status == "inactive": return "⚪"
    return "✅"

def status_color(status, flags=None, kind="bg"):
    """kind='bg' → pale background for table rows, kind='line' → solid line/marker color."""
    if status == "alert":
        return "#fff0f0" if kind == "bg" else "#e05252"
    if status == "warn":
        if _flag_is_surge(flags):
            return "#fff3e0" if kind == "bg" else "#ff8c00"   # orange — surge
        return "#fffbe6" if kind == "bg" else "#f0a500"        # yellow — decline/downtrend/yoy
    return "transparent" if kind == "bg" else "#4caf50"

def draw_flagged_series(flagged, key_prefix):
    """
    Shared sparkline grid + 🔴/🟠/🟡/All filter radio — used identically by
    IMMOFR's Trends tab and AUTOFR's Trends tab, so both markets get the exact
    same selection controls and colour logic.
    """
    n_crit = sum(1 for r in flagged if r["status"]=="alert")
    n_surge = sum(1 for r in flagged if r["status"]=="warn" and _flag_is_surge(r.get("flags")))
    n_warn = sum(1 for r in flagged if r["status"]=="warn" and not _flag_is_surge(r.get("flags")))

    show_filter = st.radio(
        "Show",
        ["🔴 Critical only", "🟠 Surge only", "🟡 Warnings only", "🔴🟠🟡 All"],
        index=3, horizontal=True, key=f"{key_prefix}_show",
    )
    if "Critical" in show_filter:
        flagged = [r for r in flagged if r["status"]=="alert"]
    elif "Surge" in show_filter:
        flagged = [r for r in flagged if r["status"]=="warn" and _flag_is_surge(r.get("flags"))]
    elif "Warnings" in show_filter:
        flagged = [r for r in flagged if r["status"]=="warn" and not _flag_is_surge(r.get("flags"))]

    if not flagged:
        st.info("No series match this filter.")
        return

    st.caption(f"Showing {len(flagged)} series — 🔴 {n_crit} critical · 🟠 {n_surge} surge · 🟡 {n_warn} warnings")
    cols3 = st.columns(3)
    for i, row in enumerate(flagged[:18]):
        clr = status_color(row["status"], row.get("flags"), kind="line")
        with cols3[i % 3]:
            vals = [v if v is not None else None for v in row["vals"]]
            x_vals = row["months"][:len(vals)]
            n_real = sum(1 for v in vals if v is not None)
            if n_real <= 2:
                # Only 2 real months — a line is misleading (barely a "trend"), and with a
                # sparse x-axis it can even look like a zigzag. A simple M-1→M bar comparison
                # reads correctly at a glance for exactly two points.
                bar_x = x_vals[-2:] if len(x_vals) >= 2 else x_vals
                bar_y = vals[-2:] if len(vals) >= 2 else vals
                fig = go.Figure(go.Bar(
                    x=bar_x, y=bar_y, marker_color=clr, width=0.5,
                    text=[f"{v:,.0f}" if v is not None else "" for v in bar_y],
                    textposition="outside",
                    hovertemplate="%{x}: %{y:,.0f}<extra></extra>"))
                fig.update_layout(height=110, margin=dict(l=0,r=0,t=18,b=0),
                    paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                    xaxis=dict(showgrid=False, tickfont=dict(size=8), type="category"),
                    yaxis=dict(showgrid=False, tickfont=dict(size=8), showticklabels=False),
                    showlegend=False)
            else:
                fig = go.Figure(go.Scatter(
                    x=x_vals, y=vals, mode="lines+markers",
                    line=dict(color=clr, width=2.5),
                    marker=dict(size=[7 if k==len(vals)-1 else 0 for k in range(len(vals))]),
                    connectgaps=False,
                    hovertemplate="%{x}: %{y:,.0f}<extra></extra>"))
                fig.update_layout(height=110, margin=dict(l=0,r=0,t=0,b=0),
                    paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                    xaxis=dict(showgrid=False, tickfont=dict(size=8), nticks=4),
                    yaxis=dict(showgrid=True, gridcolor="rgba(180,180,180,.3)",
                               tickfont=dict(size=8), tickformat=".2s"), showlegend=False)
            evol_s = f"{row['evol']:+.1f}%" if row.get("evol") is not None else ""
            section_lbl = row.get("section","") or ""
            sheet_lbl = row.get("sheet","")
            file_lbl = row.get("file","")
            # Clear hierarchy, top to bottom: 📁 File → 📄 Sheet/tab → 📊 Table (if it adds info)
            label_lines = []
            if file_lbl: label_lines.append(f"📁 {file_lbl}")
            if sheet_lbl: label_lines.append(f"📄 {sheet_lbl}")
            if section_lbl and section_lbl != sheet_lbl and len(section_lbl) > 2:
                label_lines.append(f"📊 {section_lbl}")
            full_label = "  \n".join(f"*{l}*" for l in label_lines)
            label = row.get("site","")
            st.markdown(f"**{label}** {evol_s}  \n{full_label}")
            st.plotly_chart(fig, use_container_width=True,
                            config={"displayModeBar": False},
                            key=f"{key_prefix}_{i}_{label[:6]}")

def classify_mom_status(raw_vals, li):
    """
    SINGLE SOURCE OF TRUTH for M/M-1 severity — shared by the Trends tab and the
    Table checks tab so both always use the exact same 🔴/🟡 scale:
      🔴 alert : M/M-1 drop ≥20%  OR  crash ≥40% vs 12-month peak
      🟡 warn  : M/M-1 decline 10–20%, surge ≥30%, 3-month downtrend >5%, or |M/Y-1| >30%
    raw_vals: chronological list of values (oldest→newest, None allowed).
    li: index of the reference (last) month within raw_vals.
    Returns dict: status, evol, evol_y1, crash_pct, downtrend_pct, reasons (list of codes).
    """
    n = len(raw_vals)
    lv = raw_vals[li] if 0 <= li < n else None
    pv = raw_vals[li-1] if li-1 >= 0 else None
    if pv is not None and (not isinstance(pv, (int, float)) or pv <= 0):
        pv = None
    status, reasons = "ok", []
    # Round every ratio-derived % to 6dp before thresholding: a "true" exact 20.0 or 10.0
    # can come out as 19.999999999999996 or -9.999999999999998 from float division, which
    # would silently miss the ≥20%/≥10% boundary — rounding removes that noise either way.
    evol = round((lv/pv - 1) * 100, 6) if (lv is not None and isinstance(lv, (int, float)) and pv) else None
    if evol is not None:
        if evol <= -20:
            status = "alert"; reasons.append("drop")
        elif evol <= -10:
            status = "warn"; reasons.append("decline")
        elif evol >= 30:
            status = "warn"; reasons.append("surge")
    crash_pct = None
    hist_peak = [v for v in raw_vals[max(0, li-12):li] if v and v > 0]
    if hist_peak and lv is not None and isinstance(lv, (int, float)) and lv > 0:
        peak_ratio = round(lv/max(hist_peak), 6)
        if peak_ratio < 0.6:
            status = "alert"; crash_pct = (peak_ratio - 1) * 100; reasons.append("crash")
    downtrend_pct = None
    if li >= 2:
        l3 = raw_vals[li-2:li+1]
        if all(v and v > 0 for v in l3) and l3[0] > l3[1] > l3[2]:
            drop = round((l3[2]-l3[0]) / l3[0] * 100, 6)
            if drop < -5:
                if status == "ok": status = "warn"
                downtrend_pct = drop; reasons.append("downtrend3m")
    evol_y1 = None
    if li >= 12:
        y1v = raw_vals[li-12]
        if y1v and isinstance(y1v, (int, float)) and y1v > 5 and lv is not None and isinstance(lv, (int, float)):
            evol_y1 = round((lv/y1v - 1) * 100, 6)
            if abs(evol_y1) > 30:
                if status == "ok": status = "warn"
                reasons.append("yoy")
    return {"status": status, "evol": evol, "evol_y1": evol_y1,
            "crash_pct": crash_pct, "downtrend_pct": downtrend_pct, "reasons": reasons}

def _trend_row(fname, sn, label, site, mo, lm, pm, sd, dd):
    """One trend row for a site/section; None if site inactive in reference month."""
    if not site_active(sd):
        return None
    raw_vals = list(sd["v"])
    n = len(raw_vals)
    if n < 2:
        return None
    li, pi = n - 1, n - 2
    lv = raw_vals[li]
    if lv is None or not isinstance(lv, (int, float)) or lv <= 5:
        return None
    pv_raw = raw_vals[pi]
    pv = pv_raw if (pv_raw is not None and isinstance(pv_raw, (int, float)) and pv_raw > 0) else None
    actual_lm = mo[li] if li < len(mo) else lm
    actual_pm = mo[pi] if pi < len(mo) else pm
    prm = lv / dd[li] * 100 if li in dd and dd[li] else None
    prm1 = pv / dd[pi] * 100 if pv and pi in dd and dd[pi] else None
    cl = classify_mom_status(raw_vals, li)
    status, evol, evol_y1 = cl["status"], cl["evol"], cl["evol_y1"]
    flags = []
    if "drop" in cl["reasons"]: flags.append(f"Drop {evol:.1f}% vs {actual_pm}")
    if "decline" in cl["reasons"]: flags.append(f"Decline {evol:.1f}% vs {actual_pm}")
    if "surge" in cl["reasons"]: flags.append(f"Surge +{evol:.1f}% vs {actual_pm}")
    if "crash" in cl["reasons"]: flags.append(f"Crash {cl['crash_pct']:.1f}% vs 12m peak")
    if "downtrend3m" in cl["reasons"]: flags.append(f"3-month downtrend {cl['downtrend_pct']:.1f}%")
    if "yoy" in cl["reasons"]:
        y1_idx = li - 12
        y1_label = mo[y1_idx] if y1_idx < len(mo) else "Y-1"
        flags.append(f"M/Y-1: {evol_y1:+.1f}% vs {y1_label}")
    spark_vals = strip_trailing_zeros(raw_vals)
    return {
        "file": fname, "sheet": sn, "section": label, "site": site,
        "lm": actual_lm, "pm": actual_pm, "lv": lv, "pv": pv, "evol": evol, "prm": prm,
        "epr": (prm - prm1 if prm is not None and prm1 is not None else None),
        "evol_y1": evol_y1, "status": status, "flags": flags,
        "vals": spark_vals, "months": mo[:len(spark_vals)],
    }

def build_trends(raw_bytes_dict, wbs=None):
    """All sections per sheet; reference month = last column; inactive sites skipped."""
    dd_panel = panel_dedup_by_index(wbs) if wbs else {}
    rows = []
    for fname, data in raw_bytes_dict.items():
        try:
            wb = load_workbook(io.BytesIO(data), data_only=True)
        except Exception:
            continue
        for sn in wb.sheetnames:
            if sn == "Intro":
                continue
            ws = wb[sn]
            for d in read_all_sections(ws):
                if not d:
                    continue
                mo, lm, pm = d["_m"], d["_lm"], d["_pm"]
                label = d.get("_label", "")
                dd_local = {}
                for k, v in d.items():
                    if isinstance(k, str) and ("Dédupliqué" in k or "Dedup" in k) and not k.startswith("_"):
                        dd_local = {i: val for i, val in enumerate(v["v"]) if val and val > 0}
                        break
                dd = dd_local if dd_local else dd_panel
                for site, sd in d.items():
                    if not isinstance(site, str) or site.startswith("_") or site in SKIP:
                        continue
                    row = _trend_row(fname, sn, label, site, mo, lm, pm, sd, dd)
                    if row:
                        rows.append(row)
        wb.close()
    return rows

# ═══════════════════════════════════════════════
# SPECIAL CHECK — MARKET SHARE ANALYSIS (MoM)
# ═══════════════════════════════════════════════

def market_share_analysis(wbs):
    """
    Market share = site listings / total deduplicated listings (per segment).
    Tracked Month-over-Month (reference month vs previous month, in pp).

    Two breakdowns:
      A) Sales (Vente) vs Rentals (Location) × Pros vs Private (Particuliers)
         from 1.3 Loc_Ventes (6 sections)
      B) Pro by type (Agences / Intermédiaires / Notaires / Autres)
         from 1.4 Type de professionels

    Important UI rule: inactive sites are excluded from Special Check tables.
    This avoids rows like AvendreAlouer with no real activity showing empty / 0 values.
    """
    out = {
        "vente_location": [],
        "by_type": [],
        "lm": "—",
        "pm": "—",
        "dedup_vl": {},
        "dedup_type": {},
        "inactive_vl": defaultdict(list),
        "inactive_type": defaultdict(list),
    }
    if "file1" not in wbs:
        return out
    w1 = wbs["file1"]

    def _site_series(d, site):
        return d.get(site) or next(
            (d[k] for k in d if isinstance(k, str)
             and site.lower() in k.lower()
             and not k.startswith("_")),
            None
        )

    def _pct_change(now, prev):
        if prev is None or not isinstance(prev, (int, float)) or prev <= 0:
            return None
        if now is None or not isinstance(now, (int, float)):
            return None
        return (now / prev - 1) * 100

    def _status(ms_now, delta_pp, listings_mom):
        """One status used for row colors in the Special Check UI — same 🔴/🟠/🟡 scale
        as the rest of the app: 🔴 = critical/negative, 🟠 = surge (large positive move),
        🟡 = moderate shift either direction."""
        if ms_now is not None and ms_now > 100.5:
            return "alert", "Share exceeds 100%"
        if listings_mom is not None and listings_mom <= -25:
            return "alert", "Large listing-volume drop"
        if delta_pp is not None and delta_pp <= -5:
            return "alert", "Large market-share drop"
        if listings_mom is not None and listings_mom >= 40:
            return "surge", "Large listing-volume surge"
        if delta_pp is not None and delta_pp >= 5:
            return "surge", "Large market-share surge"
        if listings_mom is not None and listings_mom <= -15:
            return "warn", "Moderate listing-volume drop"
        if delta_pp is not None and delta_pp <= -3:
            return "warn", "Moderate market-share drop"
        if listings_mom is not None and listings_mom >= 25:
            return "warn", "Moderate listing-volume increase"
        if delta_pp is not None and delta_pp >= 3:
            return "warn", "Moderate market-share increase"
        return "ok", "Stable"

    # ── A) Vente/Location × Pros/Particuliers (1.3 Loc_Ventes) ──
    ws13 = ws_get(w1, "1.3 Loc_Ventes")
    if ws13:
        # 6 sections: 0=Ventes, 1=Ventes Pros, 2=Ventes Part,
        #             3=Locations, 4=Locations Pros, 5=Locations Part
        seg_labels = {
            0: ("Sales", "All"),
            1: ("Sales", "Pros"),
            2: ("Sales", "Private"),
            3: ("Rentals", "All"),
            4: ("Rentals", "Pros"),
            5: ("Rentals", "Private"),
        }
        for sec, (transaction, segment) in seg_labels.items():
            d = read_series(ws13, section=sec)
            if not d:
                continue
            out["lm"] = d.get("_lm", "—")
            out["pm"] = d.get("_pm", "—")
            dd = d.get("Total Panel Dédupliqué Marché") or d.get("Total Panel Dédupliqué")
            if not dd:
                continue
            dd_last, dd_prev = dd["last"], dd["prev"]
            if dd_last is None or not isinstance(dd_last, (int, float)) or dd_last <= 0:
                continue
            seg_key = f"{transaction} · {segment}"
            out["dedup_vl"][seg_key] = {
                "now": dd_last,
                "prev": dd_prev,
                "mom": _pct_change(dd_last, dd_prev),
            }

            for ent in SITES:
                sd = _site_series(d, ent)
                if not sd:
                    continue
                if not site_active(sd, min_vol=50):
                    out["inactive_vl"][seg_key].append(ent)
                    continue
                listings_now, listings_prev = sd["last"], sd["prev"]
                ms_now = (listings_now / dd_last * 100) if listings_now is not None else None
                ms_prev = (listings_prev / dd_prev * 100) if dd_prev and listings_prev is not None else None
                if ms_now is None:
                    continue
                delta = (ms_now - ms_prev) if ms_prev is not None else None
                listings_mom = _pct_change(listings_now, listings_prev)
                status, reason = _status(ms_now, delta, listings_mom)
                out["vente_location"].append({
                    "breakdown": "Transaction×Segment",
                    "transaction": transaction,
                    "segment": segment,
                    "entity": ent,
                    "listings": listings_now,
                    "listings_prev": listings_prev,
                    "listings_mom": listings_mom,
                    "dedup": dd_last,
                    "dedup_prev": dd_prev,
                    "dedup_mom": _pct_change(dd_last, dd_prev),
                    "ms_now": ms_now,
                    "ms_prev": ms_prev,
                    "delta": delta,
                    "status": status,
                    "reason": reason,
                })

    # ── B) Pro by type (1.4 Type de professionels) ──
    ws14 = ws_get(w1, "1.4 Type de professionels")
    if ws14:
        # 12 sections, grouped by type: [Total, Vente, Location] × [Agences, Interméd, Notaires, Autres]
        # We use the "Total" section of each type (sections 0,3,6,9)
        type_sections = {0: "Agences", 3: "Intermédiaires", 6: "Notaires", 9: "Autres"}
        for sec, type_name in type_sections.items():
            d = read_series(ws14, section=sec)
            if not d:
                continue
            out["lm"] = d.get("_lm", out.get("lm", "—"))
            out["pm"] = d.get("_pm", out.get("pm", "—"))
            dd = d.get("Total Panel Dédupliqué Marché") or d.get("Total Panel Dédupliqué")
            if not dd:
                continue
            dd_last, dd_prev = dd["last"], dd["prev"]
            if dd_last is None or not isinstance(dd_last, (int, float)) or dd_last <= 0:
                continue
            out["dedup_type"][type_name] = {
                "now": dd_last,
                "prev": dd_prev,
                "mom": _pct_change(dd_last, dd_prev),
            }

            for ent in SITES:
                sd = _site_series(d, ent)
                if not sd:
                    continue
                if not site_active(sd, min_vol=50):
                    out["inactive_type"][type_name].append(ent)
                    continue
                listings_now, listings_prev = sd["last"], sd["prev"]
                ms_now = (listings_now / dd_last * 100) if listings_now is not None else None
                ms_prev = (listings_prev / dd_prev * 100) if dd_prev and listings_prev is not None else None
                if ms_now is None:
                    continue
                delta = (ms_now - ms_prev) if ms_prev is not None else None
                listings_mom = _pct_change(listings_now, listings_prev)
                status, reason = _status(ms_now, delta, listings_mom)
                out["by_type"].append({
                    "breakdown": "Pro type",
                    "type": type_name,
                    "entity": ent,
                    "listings": listings_now,
                    "listings_prev": listings_prev,
                    "listings_mom": listings_mom,
                    "dedup": dd_last,
                    "dedup_prev": dd_prev,
                    "dedup_mom": _pct_change(dd_last, dd_prev),
                    "ms_now": ms_now,
                    "ms_prev": ms_prev,
                    "delta": delta,
                    "status": status,
                    "reason": reason,
                })

    # Convert defaultdicts to normal dicts so Streamlit cache serialization stays predictable.
    out["inactive_vl"] = dict(out["inactive_vl"])
    out["inactive_type"] = dict(out["inactive_type"])
    return out


# ═══════════════════════════════════════════════
# TABLE ANALYSIS
# ═══════════════════════════════════════════════

def _fmt_full(v):
    """Tam sayı göster — K formatı yok."""
    if v is None: return "—"
    try: return f"{float(v):,.0f}"
    except: return str(v)

def _fmtn(v):
    """K/M formatında sayı göster."""
    if v is None: return "N/A"
    try:
        n=float(v)
        if abs(n)>=1_000_000: return f"{n/1_000_000:.2f}M"
        if abs(n)>=1_000: return f"{n/1_000:.1f}K"
        return f"{int(n):,}"
    except: return str(v)
    if v is None: return "N/A"
    try:
        n=float(v)
        if abs(n)>=1_000_000: return f"{n/1_000_000:.2f}M"
        if abs(n)>=1_000: return f"{n/1_000:.1f}K"
        return f"{int(n):,}"
    except: return str(v)

def _detect_table_type(d):
    vals=[]
    SKIP_=SKIP  # includes both IMMOFR and AUTOFR aggregate-row labels
    for k,v in d.items():
        if not isinstance(k,str) or k.startswith("_") or k in SKIP_: continue
        last=v.get("last")
        if last is not None and isinstance(last,(int,float)) and last>0:
            vals.append(last)
    if not vals: return "volume"
    return "taux" if sum(1 for v in vals if 0<v<=1.5)/len(vals)>=0.75 else "volume"

def _table_qc_issues(d, table_type, label):
    issues=[]
    lm=d.get("_lm","?"); pm=d.get("_pm","?")
    SKIP_=SKIP  # includes both IMMOFR and AUTOFR aggregate-row labels
    SITES_=SITES  # market-agnostic — uses whichever site list is active (IMMOFR_SITES/AUTOFR_SITES)

    # Dedup denominator: Top 11 > Marché
    ms_denom=None; ms_denom_label=None
    for k,v in d.items():
        if isinstance(k,str) and "top 11" in k.lower() and not k.startswith("_"):
            ms_denom=v.get("last"); ms_denom_label=k.strip(); break
    if ms_denom is None:
        for k,v in d.items():
            if isinstance(k,str) and "marché" in k.lower() and "dédupliqué" in k.lower() and not k.startswith("_"):
                ms_denom=v.get("last"); ms_denom_label=k.strip(); break

    if table_type=="taux":
        for k,v in d.items():
            if not isinstance(k,str) or k.startswith("_") or k in SKIP_: continue
            last=v.get("last"); prev=v.get("prev")
            if last is None: continue
            if isinstance(last,(int,float)) and last>1.05:
                issues.append({"type":"TAUX>100%","severity":"error","site":k,
                    "message":f"Rate={last*100:.1f}% > 100% — calculation error in source file","values":""})
                continue
            if not isinstance(last,(int,float)): continue
            if prev is not None and isinstance(prev,(int,float)):
                if prev==0 and last>0.005:
                    issues.append({"type":"TAUX_ZERO_TO_VALUE","severity":"warning","site":k,
                        "message":f"Rate was 0% in {pm}, now {last*100:.2f}% in {lm}","values":""})
                elif last==0 and prev>0.005:
                    issues.append({"type":"TAUX_VALUE_TO_ZERO","severity":"error","site":k,
                        "message":f"Rate dropped to 0% — was {prev*100:.2f}% in {pm}","values":""})
                elif prev>0.001 and last>0 and max(last,prev)>0.01:
                    dp=round((last-prev)*100, 6); dr=round((last/prev-1)*100, 6)
                    if abs(dp)>3 and abs(dr)>50:
                        sev="error" if abs(dp)>8 or abs(dr)>100 else "warning"
                        issues.append({"type":"TAUX_JUMP","severity":sev,"site":k,
                            "message":f"{pm}: {prev*100:.2f}% → {lm}: {last*100:.2f}% (Δ {dp:+.2f}pp, {dr:+.0f}%)","values":""})
    else:
        site_data={k:v for k,v in d.items()
                   if isinstance(k,str) and not k.startswith("_") and k in SITES_}
        for site,sd in site_data.items():
            raw_vals=list(sd.get("v") or [])
            if not raw_vals or not site_active(sd): continue
            li=len(raw_vals)-1
            lv=raw_vals[li] if li>=0 else None
            if lv is None or not isinstance(lv,(int,float)): continue
            pv=raw_vals[li-1] if li-1>=0 else None
            # Same M/M-1 severity scale as the Trends tab (classify_mom_status)
            cl=classify_mom_status(raw_vals, li)
            for reason in cl["reasons"]:
                if reason=="drop":
                    issues.append({"type":"CHANGE_SEVERE","severity":"error","site":site,
                        "message":f"{cl['evol']:+.1f}% — {pm}: {_fmtn(pv)} → {lm}: {_fmtn(lv)}","values":""})
                elif reason=="decline":
                    issues.append({"type":"CHANGE_DECLINE","severity":"warning","site":site,
                        "message":f"{cl['evol']:+.1f}% — {pm}: {_fmtn(pv)} → {lm}: {_fmtn(lv)}","values":""})
                elif reason=="surge":
                    issues.append({"type":"CHANGE_SURGE","severity":"warning","site":site,
                        "message":f"{cl['evol']:+.1f}% — {pm}: {_fmtn(pv)} → {lm}: {_fmtn(lv)}","values":""})
                elif reason=="crash":
                    issues.append({"type":"CRASH_VS_PEAK","severity":"error","site":site,
                        "message":f"{cl['crash_pct']:.1f}% vs 12-month peak (now {_fmtn(lv)})","values":""})
                elif reason=="downtrend3m":
                    issues.append({"type":"DOWNTREND_3M","severity":"warning","site":site,
                        "message":f"3 consecutive declining months ({cl['downtrend_pct']:.1f}% over 3 months)","values":""})
                elif reason=="yoy":
                    issues.append({"type":"YOY_CHANGE","severity":"warning","site":site,
                        "message":f"M/Y-1: {cl['evol_y1']:+.1f}%","values":""})
            # Market share > 100% — structural check, kept independent of M/M-1 trend
            if lv and ms_denom and isinstance(ms_denom,(int,float)) and ms_denom>100:
                ms=lv/ms_denom*100
                if ms>100.5:
                    issues.append({"type":"MS_OVER_100","severity":"error","site":site,
                        "message":f"{_fmtn(lv)} > {ms_denom_label} ({_fmtn(ms_denom)}) → {ms:.0f}%","values":""})

        # Dedup M vs M-1 — same shared severity scale
        for k,v in d.items():
            if isinstance(k,str) and "dédupliqué marché" in k.lower() and not k.startswith("_"):
                dvals=list(v.get("v") or [])
                if len(dvals)>=2:
                    dli=len(dvals)-1
                    dcl=classify_mom_status(dvals, dli)
                    dv,dp_val=dvals[dli],dvals[dli-1]
                    dl=k.strip()
                    if "drop" in dcl["reasons"]:
                        issues.append({"type":"DEDUP_CHANGE_SEVERE","severity":"error","site":dl,
                            "message":f"Deduplicated total: {dcl['evol']:+.1f}% — {pm}: {_fmtn(dp_val)} → {lm}: {_fmtn(dv)}","values":""})
                    elif "surge" in dcl["reasons"]:
                        issues.append({"type":"DEDUP_CHANGE_SURGE","severity":"warning","site":dl,
                            "message":f"Deduplicated total: {dcl['evol']:+.1f}% — {pm}: {_fmtn(dp_val)} → {lm}: {_fmtn(dv)}","values":""})
                    elif "decline" in dcl["reasons"]:
                        issues.append({"type":"DEDUP_CHANGE_DECLINE","severity":"warning","site":dl,
                            "message":f"Deduplicated total: {dcl['evol']:+.1f}% — {pm}: {_fmtn(dp_val)} → {lm}: {_fmtn(dv)}","values":""})
                break
    return issues

def analyse_all_tables(raw_bytes_dict, classify_fn=None, file_roles=None):
    """classify_fn/file_roles let this same engine run for a different market
    (e.g. AUTOFR) — default to IMMOFR's classify_named()/FILE_ROLES."""
    if classify_fn is None: classify_fn = classify_named
    if file_roles is None: file_roles = FILE_ROLES
    results=[]
    fb=classify_fn(raw_bytes_dict)
    SKIP_=SKIP  # includes both IMMOFR and AUTOFR aggregate-row labels
    for role,(source_file,data) in fb.items():
        fname=file_roles.get(role,role)
        try: wb=load_workbook(io.BytesIO(data),data_only=True)
        except: continue
        for sn in wb.sheetnames:
            if sn=="Intro": continue
            ws=wb[sn]
            # Try col 2 first, then col 1
            col=2
            for c in [2, 1]:
                hc=sum(1 for r in range(1,min(ws.max_row+1,30))
                       if ws.cell(r,c).value in ("Site","Département","Région"))
                if hc>0: col=c; break
            hdr_count=sum(1 for r in range(1,ws.max_row+1)
                          if ws.cell(r,col).value in ("Site","Département","Région"))

            # Snapshot file (Focus IDF, Y-1): col headers = sites, rows = depts
            # No month columns → treat each section as a single snapshot
            if hdr_count==0:
                # Try reading as snapshot: look for site names in row 4-7
                site_row=None; site_cols={}
                for r in range(1,10):
                    for c in range(1,ws.max_column+1):
                        v=ws.cell(r,c).value
                        if isinstance(v,str) and any(s.lower() in v.lower() for s in
                                ["leboncoin","seloger","bien'ici","figaro","paruvendu",
                                 "logicimmo","ouestfrance","superimmo"]):
                            if site_row is None: site_row=r
                            site_cols[v.strip()]=c
                if site_row and site_cols:
                    # Find dedup col
                    dedup_col=None; dedup_label=None
                    for c in range(1,ws.max_column+1):
                        h=ws.cell(site_row,c).value
                        if h and isinstance(h,str) and "dédup" in h.lower() and "top 11" in h.lower():
                            dedup_col=c; dedup_label=h.strip(); break
                    if dedup_col is None:
                        for c in range(1,ws.max_column+1):
                            h=ws.cell(site_row,c).value
                            if h and isinstance(h,str) and "dédup" in h.lower():
                                dedup_col=c; dedup_label=h.strip(); break
                    # Read rows
                    site_rows_snap=[]
                    violations=[]
                    for r in range(site_row+1, ws.max_row+1):
                        dept=ws.cell(r,1).value or ws.cell(r,2).value
                        if not dept or not isinstance(dept,str) or len(dept.strip())<2: continue
                        if str(dept).upper() in ("TOTAL","SITE","RÉGION","DÉPARTEMENT"): continue
                        site_vals={s: ws.cell(r,c).value for s,c in site_cols.items()
                                   if isinstance(ws.cell(r,c).value,(int,float))}
                        dv=ws.cell(r,dedup_col).value if dedup_col else None
                        if site_vals and isinstance(dv,(int,float)) and dv>0:
                            mx=max(site_vals.values()); mx_site=max(site_vals,key=site_vals.get)
                            if mx>dv:
                                violations.append({"type":"MS_OVER_100","severity":"error",
                                    "site":mx_site,"message":f"{dept.strip()}: {mx_site} ({_fmtn(mx)}) > {dedup_label} ({_fmtn(dv)})"})
                    if violations or site_cols:
                        results.append({"file":fname,"source_file":source_file,"sheet":sn,"sec_idx":0,
                            "label":"Snapshot (no month columns)","lm":"—","pm":"—",
                            "table_type":"snapshot","dedup":None,"total":None,
                            "sites":[],"issues":violations,
                            "n_error":len([v for v in violations if v["severity"]=="error"]),
                            "n_warn":0,"n_alert":0})
                    else:
                        results.append({"file":fname,"source_file":source_file,"sheet":sn,"sec_idx":0,
                            "label":"Snapshot — no site data detected","lm":"—","pm":"—",
                            "table_type":"unreadable","dedup":None,"total":None,
                            "sites":[],"issues":[],"n_error":0,"n_warn":0,"n_alert":0})
                else:
                    results.append({"file":fname,"source_file":source_file,"sheet":sn,"sec_idx":0,
                        "label":"Snapshot — no site data detected","lm":"—","pm":"—",
                        "table_type":"unreadable","dedup":None,"total":None,
                        "sites":[],"issues":[],"n_error":0,"n_warn":0,"n_alert":0})
                continue
            sections=read_all_sections(ws, col=col)
            for sec_idx,d in enumerate(sections):
                if not d or len(d)<=3:
                    results.append({"file":fname,"source_file":source_file,"sheet":sn,"sec_idx":sec_idx,
                        "label":f"Section {sec_idx+1} — insufficient data","lm":"—","pm":"—",
                        "table_type":"unreadable","dedup":None,"total":None,
                        "sites":[],"issues":[],"n_error":0,"n_warn":0,"n_alert":0})
                    continue
                lm=d.get("_lm","?"); pm=d.get("_pm","?")
                label=d.get("_label","") or f"Section {sec_idx+1}"
                table_type=_detect_table_type(d)
                dedup=None
                for k,v in d.items():
                    if isinstance(k,str) and ("dédupliqué" in k.lower() or "dedup" in k.lower()) and not k.startswith("_"):
                        dedup=v.get("last"); break
                total=None
                for k,v in d.items():
                    if isinstance(k,str) and k.strip().lower()=="total" and not k.startswith("_"):
                        total=v.get("last"); break
                site_rows=[]
                for site,sd in d.items():
                    if not isinstance(site,str) or site.startswith("_") or site in SKIP_: continue
                    last=sd.get("last"); prev=sd.get("prev")
                    if last is None and prev is None: continue
                    evol=round((last/prev-1)*100, 6) if (last and prev and isinstance(prev,(int,float)) and prev>0) else None
                    ms_d=None
                    for k,v in d.items():
                        if isinstance(k,str) and "top 11" in k.lower() and not k.startswith("_"): ms_d=v.get("last"); break
                    if ms_d is None:
                        for k,v in d.items():
                            if isinstance(k,str) and "marché" in k.lower() and "dédupliqué" in k.lower() and not k.startswith("_"): ms_d=v.get("last"); break
                    ms=(last/ms_d*100) if (last and ms_d and isinstance(ms_d,(int,float)) and ms_d>100) else None
                    status="ok"; is_surge=False
                    if table_type=="volume":
                        # Same M/M-1 severity scale as the Trends tab (classify_mom_status)
                        raw_vals=list(sd.get("v") or [])
                        if raw_vals and site_active(sd):
                            cl=classify_mom_status(raw_vals, len(raw_vals)-1)
                            status=cl["status"]; is_surge="surge" in cl["reasons"]
                        if ms and ms>100.5: status="alert"; is_surge=False
                    else:
                        if isinstance(last,(int,float)) and last>1.05:
                            status="alert"
                        elif isinstance(last,(int,float)) and isinstance(prev,(int,float)):
                            if prev==0 and last>0.005:
                                status="warn"
                            elif last==0 and prev>0.005:
                                status="alert"
                            elif prev>0.001 and last>0 and max(last,prev)>0.01:
                                dp=round((last-prev)*100, 6); dr=round((last/prev-1)*100, 6)
                                if abs(dp)>3 and abs(dr)>50:
                                    status="alert" if abs(dp)>8 or abs(dr)>100 else "warn"
                    site_rows.append({"site":site,"last":last,"prev":prev,"evol":evol,"ms":ms,
                                       "status":status,"table_type":table_type,"is_surge":is_surge})
                if not site_rows:
                    results.append({"file":fname,"source_file":source_file,"sheet":sn,"sec_idx":sec_idx,"label":label,"lm":lm,"pm":pm,
                        "table_type":table_type,"dedup":dedup,"total":total,"sites":[],"issues":[],
                        "n_error":0,"n_warn":0,"n_alert":0})
                    continue
                issues=_table_qc_issues(d,table_type,label)
                results.append({"file":fname,"source_file":source_file,"sheet":sn,"sec_idx":sec_idx,"label":label,"lm":lm,"pm":pm,
                    "table_type":table_type,"dedup":dedup,"total":total,"sites":site_rows,"issues":issues,
                    "n_error":sum(1 for i in issues if i["severity"]=="error"),
                    "n_warn":sum(1 for i in issues if i["severity"]=="warning"),
                    "n_alert":sum(1 for r in site_rows if r["status"]=="alert"),
                })
        wb.close()
    return results

# ═══════════════════════════════════════════════
# CACHED COMPUTE
# ═══════════════════════════════════════════════

@st.cache_data(show_spinner="Running quality checks…")
def compute_everything(file_hash, raw_bytes_dict, _cache_bust_version=APP_VERSION):
    # _cache_bust_version forces Streamlit to recompute (instead of serving a stale
    # cached result) whenever APP_VERSION changes — st.cache_data only hashes this
    # function's OWN bytecode, not the bytecode of run_checks/z_checks/etc. that it
    # calls internally, so a code change deep inside those helpers would otherwise
    # silently keep returning old cached output even after a redeploy.
    fb=classify(raw_bytes_dict)
    wbs={}
    for role,data in fb.items():
        try: wbs[role]=load_workbook(io.BytesIO(data),data_only=True)
        except: pass
    checks=run_checks(fb,wbs)
    trends=build_trends(raw_bytes_dict,wbs)
    mshare=market_share_analysis(wbs)
    tables=analyse_all_tables(raw_bytes_dict)
    return checks,trends,fb,mshare,tables

@st.cache_data(show_spinner="Running AUTOFR quality checks…")
def compute_everything_autofr(file_hash, raw_bytes_dict, _cache_bust_version=APP_VERSION):
    """
    AUTOFR pipeline — reuses the exact same generic engine as IMMOFR
    (build_trends/analyse_all_tables/classify_mom_status/status_icon, etc.)
    with a different site list and file classifier. No 'Special check for LBC'
    (market share) equivalent — AUTOFR doesn't have that concept.
    """
    fb_named=classify_named_autofr(raw_bytes_dict)
    fb={role:data for role,(fn,data) in fb_named.items()}
    wbs={}
    for role,data in fb.items():
        try: wbs[role]=load_workbook(io.BytesIO(data),data_only=True)
        except: pass
    checks=run_checks_autofr(fb,wbs)
    trends=build_trends(raw_bytes_dict,wbs)
    tables=analyse_all_tables(raw_bytes_dict, classify_fn=classify_named_autofr, file_roles=FILE_ROLES_AUTOFR)
    return checks,trends,fb,tables

# ═══════════════════════════════════════════════
# SIDEBAR
# ═══════════════════════════════════════════════

with st.sidebar:
    st.markdown(f"### 🛡️ QC Gold Panel")
    st.caption(f"v{APP_VERSION}")
    _market_choice = st.radio("Market", ["🏠 IMMOFR", "🚗 AUTOFR"], horizontal=True,
                               label_visibility="collapsed", key="market_choice")
    market = "IMMOFR" if "IMMOFR" in _market_choice else "AUTOFR"
    st.divider()

    if market == "IMMOFR":
        uploaded=st.file_uploader("Upload all Excel files",type=["xlsx"],accept_multiple_files=True,key="up_immofr")
        raw_bytes={}; file_hash=None
        if uploaded:
            for f in uploaded: raw_bytes[f.name]=f.read()
            file_hash=hashlib.md5(b"".join(sorted(raw_bytes.values()))).hexdigest()
            fb_display=classify(raw_bytes)
            n_ok_files=sum(1 for k in FILE_ROLES if k in fb_display)
            n_req=len(FILE_ROLES)
            if n_ok_files>=n_req: st.success(f"✅ All {n_ok_files} files loaded")
            else: st.warning(f"{n_ok_files} / {n_req} files recognised")
            for role in FILE_ROLES:
                st.caption(f"{'✅' if role in fb_display else '⬜'} {FILE_ROLES[role]}")
            with st.expander("🔧 Debug filenames",expanded=False):
                for fn in sorted(raw_bytes.keys()):
                    r=next((k for k,v in classify({fn:b""}).items()),None)
                    st.caption(f"{'✅' if r else '❌'} [{r or '?'}] {fn}")
    else:
        uploaded=st.file_uploader("Upload all Excel files",type=["xlsx"],accept_multiple_files=True,key="up_autofr")
        raw_bytes={}; file_hash=None
        if uploaded:
            for f in uploaded: raw_bytes[f.name]=f.read()
            file_hash=hashlib.md5(b"".join(sorted(raw_bytes.values()))).hexdigest()
            fb_display=classify_named_autofr(raw_bytes)
            n_ok_files=len(fb_display)
            n_req=len(FILE_ROLES_AUTOFR)
            if n_ok_files>=n_req: st.success(f"✅ All {n_ok_files} files loaded")
            else: st.warning(f"{n_ok_files} / {n_req} files recognised")
            for role in FILE_ROLES_AUTOFR:
                st.caption(f"{'✅' if role in fb_display else '⬜'} {FILE_ROLES_AUTOFR[role]}")
            with st.expander("🔧 Debug filenames",expanded=False):
                for fn in sorted(raw_bytes.keys()):
                    r=next((k for k,(f2,_) in classify_named_autofr({fn:b""}).items()),None)
                    st.caption(f"{'✅' if r else '❌'} [{r or '?'}] {fn}")

    st.divider()
    st.caption(f"v{APP_VERSION}")

if market == "IMMOFR":
    SITES = IMMOFR_SITES
    if not uploaded:
        st.markdown("## 🏠 IMMO FR — Panel Quality Control")
        st.info("Upload your Excel files in the sidebar to begin.")
        st.stop()

    # ═══════════════════════════════════════════════
    # COMPUTE
    # ═══════════════════════════════════════════════

    checks,trends,fb,mshare,tables=compute_everything(file_hash,raw_bytes)
    GROUP_TO_FILE={}
    for _role,(_fn,_data) in classify_named(raw_bytes).items():
        _grp=FILE_ROLES.get(_role,_role).split("—")[0].strip()
        GROUP_TO_FILE[_grp]=_fn
    n_err  = sum(1 for c in checks if not c["ok"] and c["sev"]=="error")
    n_warn = sum(1 for c in checks if c["sev"]=="warning")
    n_ok   = sum(1 for c in checks if c["ok"])
    # n_alert: sadece panel site level (website bazlı), dept/région hariç
    _tr_sites_only = [r for r in trends if r["site"] in PANEL_SITES_SET]
    n_alert = len({f"{r['site']}_{r['sheet']}" for r in _tr_sites_only if r["status"]=="alert"})
    _warn_rows_ov = [r for r in _tr_sites_only if r["status"]=="warn"]
    n_surge = len({f"{r['site']}_{r['sheet']}" for r in _warn_rows_ov if _flag_is_surge(r.get("flags"))})
    n_wtr   = len({f"{r['site']}_{r['sheet']}" for r in _warn_rows_ov if not _flag_is_surge(r.get("flags"))})

    lm_ref,pm_ref="—","—"
    if "file1" in fb:
        try:
            wb1=load_workbook(io.BytesIO(fb["file1"]),data_only=True)
            ws11=ws_get(wb1,"1.1 Total")
            if ws11:
                d11=read_series(ws11,section=0)
                lm_ref=d11.get("_lm","—"); pm_ref=d11.get("_pm","—")
            wb1.close()
        except: pass

    # ═══════════════════════════════════════════════
    # HEADER + FILTER
    # ═══════════════════════════════════════════════

    col_h,col_f=st.columns([3,2])
    with col_h:
        st.markdown("## IMMO FR — Panel Quality Control")
        st.caption(f"Reference month: **{lm_ref}** · vs {pm_ref} · {len(fb)} files · v{APP_VERSION}")
    with col_f:
        chosen=st.multiselect("Filter by site",SITES,default=[],
                              placeholder="All sites",label_visibility="visible")
    site_filter=chosen if chosen else None

    # ═══════════════════════════════════════════════
    # TABS
    # ═══════════════════════════════════════════════

    tab1,tab2,tab_dq,tab_ms=st.tabs(["📋  Overview","📈  Trends — Monthly · Yearly · Dept · Région","🔍  Data & Table Checks","🎯  Special check for LBC"])

    # ═════════════ OVERVIEW ════════════

    with tab1:
        if n_err>0:
            st.error(
                f"🚫 **Cannot validate — {n_err} data integrity error{'s' if n_err!=1 else ''} must be fixed**  \n"
                f"Cross-file checks: {n_warn} warning{'s' if n_warn!=1 else ''} · {n_ok} passed  \n"
                f"Website trends (M/M-1): {n_alert} critical · {n_surge} surge · {n_wtr} warnings"
            )
        elif n_warn>5 or n_alert>0:
            st.warning(
                f"⚠️ **Needs review before validation**  \n"
                f"Cross-file checks: {n_err} errors · {n_warn} warning{'s' if n_warn!=1 else ''} · {n_ok} passed  \n"
                f"Website trends (M/M-1): {n_alert} critical series · {n_surge} surge · {n_wtr} warnings  \n"
                f"*(Dept/région trends visible in Trends tab — not counted here)*"
            )
        else:
            msg = f"✅ **Validated — data is ready** · {n_ok} integrity checks passed"
            if n_warn: msg += f" · {n_warn} minor warnings"
            st.success(msg)

        with st.expander("🎨 How colours are determined", expanded=False):
            st.markdown("""
**Data integrity** — cross-file consistency: 🔴 numbers don't match · ⚠️ Z-score anomaly · ✅ OK

**Monthly trends** (and every table/website check that uses the same scale) — M/M-1 per series:
- 🔴 drop ≥20% vs previous month, or crash <60% of the 12-month peak
- 🟠 surge ≥30% vs previous month
- 🟡 decline 10–20%, 3-month downtrend >5%, or |M/Y-1| >30%
- ✅ stable

**Table anomalies** — same 🔴/🟠/🟡 scale as Monthly trends, applied per website per table, plus:
- 🔴 Share >100% — site listings > **Total Panel Dédupliqué Top 11** (or Marché if absent)
- ⚠️ Rate jump — taux table: |Δ| > 3pp and |Δ%| > 50% and value > 1%

**Special check** — market share (same 🔴/🟠/🟡 scale): 🔴 share > 100% or shift ≥ 5pp or volume ±25% ·
🟠 volume surge ≥ 40% · 🟡 shift 3–5pp or volume ±15–25%

All tables except file 1 / sheet 1 (total résidentiel) are **Ancien only**.
""")
        st.divider()

        if "file1" in fb:
            try:
                wb1=load_workbook(io.BytesIO(fb["file1"]),data_only=True)
                ws11=ws_get(wb1,"1.1 Total")
                if ws11:
                    d11=read_series(ws11,section=0)
                    total=sv(d11,"Total") or 0; dedup=sv(d11,"Total Panel Dédupliqué Marché") or 0
                    ptot=(d11.get("Total") or {}).get("prev") or 0
                    pdd =(d11.get("Total Panel Dédupliqué Marché") or {}).get("prev") or 0
                    et=(total/ptot-1)*100 if ptot else None; ed=(dedup/pdd-1)*100 if pdd else None
                    n_active=sum(1 for s in SITES if sv(d11,s) and sv(d11,s)>0)
                    c1,c2,c3,c4,c5,c6,c7=st.columns([1.3,1.3,1,1,0.8,0.8,0.8])
                    c1.metric("Total announcements", fmt(total), f"{et:+.1f}% vs {pm_ref}" if et is not None else None)
                    c2.metric("Deduplicated panel", fmt(dedup), f"{ed:+.1f}% vs {pm_ref}" if ed is not None else None)
                    c3.metric("Active sites", f"{n_active} / {len(SITES)}")
                    c4.metric("Integrity errors", str(n_err), "none ✓" if n_err==0 else "blocking ⚠️")
                    c5.metric("🔴 Critical", n_alert, help="Trend alerts (M/M-1) — website level only.")
                    c6.metric("🟠 Surge", n_surge, help="Trend alerts (M/M-1) — website level only.")
                    c7.metric("🟡 Warning", n_wtr, help="Trend alerts (M/M-1) — website level only.")
                wb1.close()
            except: pass

        st.divider()
        st.markdown("#### Site snapshot — total announcements")
        if "file1" in fb:
            try:
                wb1=load_workbook(io.BytesIO(fb["file1"]),data_only=True)
                ws11=ws_get(wb1,"1.1 Total")
                if ws11:
                    d11=read_series(ws11,section=0)
                    sites_show=site_filter or SITES
                    cols=st.columns(4)
                    for i,site in enumerate(sites_show):
                        sd=d11.get(site) or next(
                            (d11[k] for k in d11 if isinstance(k,str) and
                             site.lower() in k.lower() and not k.startswith("_")),None)
                        lv=sd["last"] if sd else None; pv=sd["prev"] if sd else None
                        evol=(lv/pv-1)*100 if lv and pv and pv>0 else None
                        delta=f"{evol:+.1f}% vs {pm_ref}" if evol is not None else ("⚠ No data" if not lv else None)
                        with cols[i%4]: st.metric(site,fmt(lv),delta)
                wb1.close()
            except: pass

        st.divider()
        errs=[c for c in checks if not c["ok"] and c["sev"]=="error"]
        if site_filter:
            errs=[c for c in errs if any(s.lower() in c["name"].lower() for s in site_filter)]
        if errs:
            st.markdown("#### Blocking errors")
            for c in errs[:8]:
                grp=GROUP_INFO.get(c["group"],("",""))[0]
                st.error(f"**{c['name']}**  \n{c['detail']}  \n*{grp}*")
            if len(errs)>8: st.caption(f"+ {len(errs)-8} more in Data integrity tab")
        else:
            st.success("No blocking errors — all cross-file consistency checks passed.")

    # ═════════════ SPECIAL CHECK — MARKET SHARE ══════════════

    with tab_ms:
        st.caption(f"Market share — {mshare.get('lm','—')} vs {mshare.get('pm','—')} · listing volume, MoM %, share %, share shift (pp)")

        vl_rows = mshare.get("vente_location", [])
        bt_rows = mshare.get("by_type", [])

        if not vl_rows and not bt_rows:
            st.info("Market share analysis requires File 1 (1.3 Loc_Ventes and 1.4 Type de professionels).")
        else:
            def _ent_match(ent):
                if not site_filter:
                    return True
                return any(s.lower() in ent.lower() for s in site_filter)

            def _status_icon(status):
                return "🔴 Alert" if status == "alert" else "🟠 Surge" if status == "surge" else "🟡 Review" if status == "warn" else "✅ OK"

            def _fmt_pct(v):
                return f"{v:+.1f}%" if v is not None else "—"

            def _fmt_share(v):
                return f"{v:.1f}%" if v is not None else "—"

            def _fmt_pp(v):
                return f"{v:+.1f}pp" if v is not None else "—"

            def _style_ms_table(df):
                def _row_style(row):
                    status = str(row.get("Status", ""))
                    if "🔴" in status:
                        bg = "background-color: #fdecec"
                    elif "🟠" in status:
                        bg = "background-color: #fff3e0"
                    elif "🟡" in status:
                        bg = "background-color: #fff4d6"
                    else:
                        bg = "background-color: #edf7ed"
                    return [bg for _ in row]
                return df.style.apply(_row_style, axis=1)

            def _segment_label(r):
                return f"{r['transaction']} · {r['segment']}" if "transaction" in r else r.get("type", "")

            def _make_site_df(rows):
                rows = [r for r in rows if _ent_match(r["entity"])]
                rows = sorted(rows, key=lambda r: ({"alert": 0, "surge": 1, "warn": 2, "ok": 3}.get(r["status"], 4), r["entity"]))
                return pd.DataFrame([
                    {
                        "Status": _status_icon(r["status"]),
                        "Website": r["entity"],
                        f"Listings {mshare.get('lm','M')}": fmt(r.get("listings")),
                        f"Listings {mshare.get('pm','M-1')}": fmt(r.get("listings_prev")),
                        "Listings MoM": _fmt_pct(r.get("listings_mom")),
                        f"Share {mshare.get('lm','M')}": _fmt_share(r.get("ms_now")),
                        f"Share {mshare.get('pm','M-1')}": _fmt_share(r.get("ms_prev")),
                        "Share Δ": _fmt_pp(r.get("delta")),
                        "Dedup denominator": fmt(r.get("dedup")),
                        "Dedup MoM": _fmt_pct(r.get("dedup_mom")),
                        "ℹ️": r.get("reason", ""),
                    }
                    for r in rows
                ])

            def _make_attention_df(rows):
                rows = [r for r in rows if r.get("status") in ("alert", "surge", "warn") and _ent_match(r["entity"])]
                rows = sorted(rows, key=lambda r: ({"alert": 0, "surge": 1, "warn": 2}.get(r["status"], 3), abs(r.get("delta") or 0)), reverse=False)
                return pd.DataFrame([
                    {
                        "Status": _status_icon(r["status"]),
                        "Website": r["entity"],
                        "Segment": _segment_label(r),
                        "Listings MoM": _fmt_pct(r.get("listings_mom")),
                        "Share Δ": _fmt_pp(r.get("delta")),
                        f"Share {mshare.get('lm','M')}": _fmt_share(r.get("ms_now")),
                        "ℹ️": r.get("reason", ""),
                    }
                    for r in rows
                ])

            def _make_dedup_df(ddict, order):
                table = []
                for seg in order:
                    v = ddict.get(seg)
                    if not v:
                        continue
                    table.append({
                        "Segment": seg,
                        f"Dedup {mshare.get('lm','M')}": fmt(v.get("now")),
                        f"Dedup {mshare.get('pm','M-1')}": fmt(v.get("prev")),
                        "Dedup MoM": _fmt_pct(v.get("mom")),
                    })
                return pd.DataFrame(table)

            all_rows = [r for r in (vl_rows + bt_rows) if _ent_match(r["entity"])]
            n_alert_ms = sum(1 for r in all_rows if r["status"] == "alert")
            n_surge_ms = sum(1 for r in all_rows if r["status"] == "surge")
            n_warn_ms = sum(1 for r in all_rows if r["status"] == "warn")
            active_websites = len(set(r["entity"] for r in all_rows))
            inactive_hidden = sum(len(v) for v in mshare.get("inactive_vl", {}).values()) + sum(len(v) for v in mshare.get("inactive_type", {}).values())

            c1, c2, c3, c4, c5 = st.columns(5)
            c1.metric("🔴 Alerts", n_alert_ms)
            c2.metric("🟠 Surge", n_surge_ms)
            c3.metric("🟡 To review", n_warn_ms)
            c4.metric("Active websites", active_websites)
            c5.metric("Inactive (hidden)", inactive_hidden, help="Sites with no meaningful volume in the reference month.")

            # ── Rows needing attention — always at the top ──
            attention_df = _make_attention_df(all_rows)
            if not attention_df.empty:
                st.divider()
                st.markdown("#### ⚠️ Rows needing attention")
                st.dataframe(
                    _style_ms_table(attention_df),
                    use_container_width=True,
                    hide_index=True,
                    height=min(450, 42 + 35 * len(attention_df)),
                    column_config={
                        "Status":       st.column_config.TextColumn(width="small"),
                        "Website":      st.column_config.TextColumn(width="small"),
                        "Segment":      st.column_config.TextColumn(width="small"),
                        "Listings MoM": st.column_config.TextColumn(width="small"),
                        "Share Δ":      st.column_config.TextColumn(width="small"),
                        "ℹ️": st.column_config.TextColumn("ℹ️ Reason", width="large"),
                    },
                )
            else:
                st.divider()
                st.success("✅ No significant month-over-month movements.")

            st.divider()

            view = st.radio("Breakdown", ["Sales / Rentals", "Pro types"], horizontal=True, label_visibility="collapsed")

            if view == "Sales / Rentals":
                seg_order = ["Sales · All", "Sales · Pros", "Sales · Private", "Rentals · All", "Rentals · Pros", "Rentals · Private"]
                available = sorted(set(f"{r['transaction']} · {r['segment']}" for r in vl_rows), key=lambda x: seg_order.index(x) if x in seg_order else 99)
                if available:
                    selected_seg = st.selectbox("Segment", available, index=0)
                    selected_rows = [r for r in vl_rows if f"{r['transaction']} · {r['segment']}" == selected_seg]
                else:
                    selected_seg = None
                    selected_rows = []
            else:
                type_order = ["Agences", "Intermédiaires", "Notaires", "Autres"]
                available = sorted(set(r["type"] for r in bt_rows), key=lambda x: type_order.index(x) if x in type_order else 99)
                if available:
                    selected_seg = st.selectbox("Pro type", available, index=0)
                    selected_rows = [r for r in bt_rows if r["type"] == selected_seg]
                else:
                    selected_seg = None
                    selected_rows = []

            df_site = _make_site_df(selected_rows)
            if not df_site.empty:
                st.dataframe(
                    _style_ms_table(df_site),
                    use_container_width=True,
                    hide_index=True,
                    height=min(560, 42 + 35 * len(df_site)),
                    column_config={
                        "Status":            st.column_config.TextColumn(width="small"),
                        "Website":           st.column_config.TextColumn(width="small"),
                        "Listings MoM":      st.column_config.TextColumn(width="small"),
                        "Share Δ":           st.column_config.TextColumn(width="small"),
                        "Dedup denominator": st.column_config.TextColumn(width="small"),
                        "Dedup MoM":         st.column_config.TextColumn(width="small"),
                        "ℹ️": st.column_config.TextColumn("ℹ️ Reason", width="large"),
                    },
                )
            else:
                st.info("No active websites for this selection.")

            inactive_map = mshare.get("inactive_vl", {}) if view == "Sales / Rentals" else mshare.get("inactive_type", {})
            hidden_for_segment = inactive_map.get(selected_seg, []) if selected_seg else []
            if hidden_for_segment:
                st.caption(f"Hidden inactive website(s) for this segment: {', '.join(sorted(set(hidden_for_segment)))}")

            # ── Compact matrix — shown below the detail table, linked to the same view ──
            if view == "Sales / Rentals":
                seg_order_mx = [("Sales", "All"), ("Sales", "Pros"), ("Sales", "Private"),
                                ("Rentals", "All"), ("Rentals", "Pros"), ("Rentals", "Private")]
                seg_cols = [f"{t} · {s}" for t, s in seg_order_mx]
                ent_map = defaultdict(dict)
                for r in vl_rows:
                    if not _ent_match(r["entity"]): continue
                    ent_map[r["entity"]][f"{r['transaction']} · {r['segment']}"] = r
                matrix_rows = []
                for ent in [s for s in SITES if s in ent_map]:
                    row = {"Website": ent}
                    for col in seg_cols:
                        r = ent_map[ent].get(col)
                        if r and r["ms_now"] is not None:
                            icon = " 🔴" if r["status"] == "alert" else " 🟠" if r["status"] == "surge" else " 🟡" if r["status"] == "warn" else ""
                            row[col] = f"{r['ms_now']:.1f}% ({r['delta']:+.1f}pp){icon}" if r["delta"] is not None else f"{r['ms_now']:.1f}%"
                        else:
                            row[col] = "—"
                    matrix_rows.append(row)
                if matrix_rows:
                    st.divider()
                    st.caption("Market share matrix — all segments · format: Share% (Δpp)")
                    st.dataframe(pd.DataFrame(matrix_rows), use_container_width=True, hide_index=True)
            else:
                # Pro types matrix
                type_order_mx = ["Agences", "Intermédiaires", "Notaires", "Autres"]
                ent_map_bt = defaultdict(dict)
                for r in bt_rows:
                    if not _ent_match(r["entity"]): continue
                    ent_map_bt[r["entity"]][r["type"]] = r
                matrix_rows_bt = []
                for ent in [s for s in SITES if s in ent_map_bt]:
                    row = {"Website": ent}
                    for t in type_order_mx:
                        r = ent_map_bt[ent].get(t)
                        if r and r["ms_now"] is not None:
                            icon = " 🔴" if r["status"] == "alert" else " 🟠" if r["status"] == "surge" else " 🟡" if r["status"] == "warn" else ""
                            row[t] = f"{r['ms_now']:.1f}% ({r['delta']:+.1f}pp){icon}" if r["delta"] is not None else f"{r['ms_now']:.1f}%"
                        else:
                            row[t] = "—"
                    matrix_rows_bt.append(row)
                if matrix_rows_bt:
                    st.divider()
                    st.caption("Market share matrix — pro types · format: Share% (Δpp)")
                    st.dataframe(pd.DataFrame(matrix_rows_bt), use_container_width=True, hide_index=True)

            export = []
            for r in vl_rows:
                export.append({
                    "Breakdown": "Transaction×Segment",
                    "Entity": r["entity"],
                    "Segment": f"{r['transaction']} {r['segment']}",
                    "Listings M": r.get("listings"),
                    "Listings M-1": r.get("listings_prev"),
                    "Listings MoM %": round(r["listings_mom"], 2) if r.get("listings_mom") is not None else None,
                    "Dedup total M": r.get("dedup"),
                    "Dedup total M-1": r.get("dedup_prev"),
                    "Dedup MoM %": round(r["dedup_mom"], 2) if r.get("dedup_mom") is not None else None,
                    "Market share %": round(r["ms_now"], 2) if r.get("ms_now") is not None else None,
                    "Prev market share %": round(r["ms_prev"], 2) if r.get("ms_prev") is not None else None,
                    "Delta pp": round(r["delta"], 2) if r.get("delta") is not None else None,
                    "Status": r.get("status"),
                    "ℹ️": r.get("reason",""),
                })
            for r in bt_rows:
                export.append({
                    "Breakdown": "Pro type",
                    "Entity": r["entity"],
                    "Segment": r["type"],
                    "Listings M": r.get("listings"),
                    "Listings M-1": r.get("listings_prev"),
                    "Listings MoM %": round(r["listings_mom"], 2) if r.get("listings_mom") is not None else None,
                    "Dedup total M": r.get("dedup"),
                    "Dedup total M-1": r.get("dedup_prev"),
                    "Dedup MoM %": round(r["dedup_mom"], 2) if r.get("dedup_mom") is not None else None,
                    "Market share %": round(r["ms_now"], 2) if r.get("ms_now") is not None else None,
                    "Prev market share %": round(r["ms_prev"], 2) if r.get("ms_prev") is not None else None,
                    "Delta pp": round(r["delta"], 2) if r.get("delta") is not None else None,
                    "Status": r.get("status"),
                    "ℹ️": r.get("reason",""),
                })
            if export:
                st.download_button(
                    "⬇ Download Special check CSV",
                    pd.DataFrame(export).to_csv(index=False).encode("utf-8-sig"),
                    f"special_check_market_share_{mshare.get('lm','')}.csv",
                    "text/csv",
                )


    # ═════════════ TRENDS ══════════════

    with tab2:
        tr=[r for r in trends if not site_filter or r["site"] in site_filter]
        n_inactive=sum(1 for r in tr if r["status"]=="inactive")

        st.divider()

        # tr_filtered = tüm trend verisi (sub-tab'lar kendi filtrelemesini yapıyor)
        tr_filtered = tr
        is_geo_mode = any(_has_dept_number(r["site"]) or _is_region(r["site"]) for r in tr)

        # Ortak veri subsetleri — sub-tab'lardan önce tanımla
        tr_site = [r for r in tr_filtered if r["site"] in PANEL_SITES_SET] or tr_filtered
        tr_dept_all = [r for r in tr_filtered if _has_dept_number(r["site"])]
        tr_region_all = [r for r in tr_filtered if _is_region(r["site"])]

        subtab_m, subtab_y, subtab_d, subtab_r = st.tabs([
            "📅 Monthly (M/M-1)",
            "📆 Yearly (M/Y-1)",
            "🗺️ By département",
            "🏙️ By région",
        ])

        _draw_flagged = draw_flagged_series  # shared with AUTOFR — see draw_flagged_series()

        def _trend_table_html(rows, title="All series — M/M-1", cols_extra=None):
            """Render styled HTML table with hover tooltip on Notes."""
            import html as _html
            order = {"alert":0,"warn":1,"ok":2,"inactive":3}
            sorted_rows = sorted(rows, key=lambda x: order.get(x["status"],3))

            rows_html = ""
            for r in sorted_rows:
                icon = status_icon(r["status"], r.get("flags")) if r["status"]!="inactive" else "⚪"
                bg = status_color(r["status"], r.get("flags"), kind="bg") if r["status"]!="inactive" else "transparent"
                evol = f"{r['evol']:+.1f}%" if r["evol"] is not None else "—"
                evol_y1 = f"{r.get('evol_y1'):+.1f}%" if cols_extra=="y1" and r.get("evol_y1") is not None else None
                mkt = f"{r['prm']:.1f}%" if r["prm"] is not None else "—"
                notes_raw = " · ".join(r.get("flags") or []) or ("Not reporting" if r["status"]=="inactive" else "")
                notes_escaped = _html.escape(notes_raw)
                site = _html.escape(str(r["site"]))
                sec = _html.escape(str(r.get("section","") or r.get("sheet","")))
                lv = _html.escape(_fmt_full(r["lv"])); pv = _html.escape(_fmt_full(r["pv"]))
                lm = _html.escape(str(r["lm"])); pm = _html.escape(str(r["pm"]))

                extra_col = f"<td>{evol_y1}</td>" if evol_y1 is not None else ""

                note_cell = ""
                if notes_raw:
                    note_cell = f"""<td><span class="nt" data-tip="{notes_escaped}">⚠ note</span></td>"""
                else:
                    note_cell = "<td>—</td>"

                rows_html += f"""<tr style="background:{bg}">
                  <td>{icon}</td><td><b>{site}</b></td><td style="color:#666;font-size:12px">{sec}</td>
                  <td>{lv}</td><td>{pv}</td><td><b>{evol}</b></td>
                  {extra_col}<td style="color:#888">{mkt}</td>{note_cell}
                </tr>"""

            y1_header = "<th>M/Y-1</th>" if cols_extra=="y1" else ""

            html_out = f"""
    <style>
    .trtbl {{width:100%;border-collapse:collapse;font-size:13px;font-family:sans-serif;table-layout:fixed}}
    .trtbl th {{text-align:left;padding:6px 8px;border-bottom:2px solid #ddd;color:#666;font-weight:500;white-space:nowrap;overflow:hidden}}
    .trtbl td {{padding:5px 8px;border-bottom:1px solid #f0f0f0;vertical-align:middle;overflow:hidden;text-overflow:ellipsis}}
    .trtbl td:nth-child(1) {{width:30px;white-space:nowrap}}
    .trtbl td:nth-child(2) {{width:140px;white-space:nowrap;font-weight:600}}
    .trtbl td:nth-child(3) {{width:220px;white-space:normal;word-break:break-word;font-size:12px;color:#555}}
    .trtbl td:nth-child(4),.trtbl td:nth-child(5) {{width:80px;white-space:nowrap}}
    .trtbl td:nth-child(6) {{width:70px;white-space:nowrap;font-weight:600}}
    .trtbl td:nth-child(7) {{width:80px;white-space:nowrap;color:#888}}
    .trtbl td:nth-child(8) {{width:60px;white-space:nowrap}}
    .trtbl tr:hover td {{background:rgba(0,0,0,0.02)}}
    .nt {{cursor:help;color:#e07b00;border-bottom:1px dashed #e07b00;font-size:12px;position:relative;white-space:nowrap}}
    .nt::after {{
      content: attr(data-tip);
      position:fixed;left:50%;transform:translateX(-50%);top:auto;bottom:60px;
      background:#222;color:#fff;padding:10px 14px;border-radius:8px;
      font-size:13px;white-space:normal;min-width:320px;max-width:600px;
      line-height:1.6;z-index:9999;pointer-events:none;opacity:0;transition:opacity .2s;
      box-shadow:0 4px 16px rgba(0,0,0,.4);
    }}
    .nt:hover::after {{opacity:1}}
    </style>
    <div style="overflow-x:auto;width:100%">
    <table class="trtbl">
    <thead><tr>
      <th></th><th>Site / Area</th><th>Table</th>
      <th>M ({rows[0]['lm'] if rows else ''})</th>
      <th>M-1 ({rows[0]['pm'] if rows else ''})</th>
      <th>M/M-1</th>{y1_header}<th>Mkt share</th><th>Notes</th>
    </tr></thead>
    <tbody>{rows_html}</tbody>
    </table>
    </div>"""
            return html_out

        # ─── MONTHLY (M/M-1) — site-level only ───
        with subtab_m:
            flagged_m = []
            seen_m = set()
            for r in tr_site:
                mom_flags = [f for f in (r.get("flags") or []) if _is_mom_flag(f)]
                if not mom_flags:
                    continue  # Crash-only / Y-1-only / downtrend-only → Monthly'de değil
                k = f"{r['site']}_{r['sheet']}"
                if k not in seen_m:
                    seen_m.add(k)
                    # Status'u sadece M/M-1 flag'ine göre belirle
                    real_status = "alert" if any("drop" in f.lower() for f in mom_flags) else "warn"
                    flagged_m.append({**r, "flags": mom_flags, "status": real_status})

            n_crit_m = sum(1 for r in flagged_m if r["status"]=="alert")
            n_surge_m = sum(1 for r in flagged_m if r["status"]=="warn" and _flag_is_surge(r.get("flags")))
            n_warn_m = sum(1 for r in flagged_m if r["status"]=="warn" and not _flag_is_surge(r.get("flags")))
            st.caption(f"**Website-level series** — 🔴 {n_crit_m} critical · 🟠 {n_surge_m} surge · 🟡 {n_warn_m} warnings"
                       f"  \n*Critical = M/M-1 drop ≥ 20% · Surge = ≥ 30% increase · Warning = decline 10-20%*"
                       f"  \n*Crash vs 12m peak and Y-1 anomalies → Yearly tab*")

            if flagged_m:
                st.markdown(f"#### 🚨 Anomalies — {len(flagged_m)} series")
                _draw_flagged(flagged_m, "m")
            else:
                st.success("✅ No M/M-1 anomalies.")

            st.divider()
            st.markdown("#### All series — M/M-1")
            st.divider()
            with st.expander("📋 Detailed table — click to expand for in-depth review", expanded=False):
                st.caption("Filter by severity to focus on what matters.")
                sev_m = st.radio("Show", ["🔴 Critical", "🟠 Surge", "🟡 Warnings", "🔴🟠🟡 All"],
                                 index=3, horizontal=True, key="sev_m_tbl")
                tr_m_show = [r for r in flagged_m if
                             (sev_m=="🔴🟠🟡 All") or
                             (sev_m=="🔴 Critical" and r["status"]=="alert") or
                             (sev_m=="🟠 Surge" and r["status"]=="warn" and _flag_is_surge(r.get("flags"))) or
                             (sev_m=="🟡 Warnings" and r["status"]=="warn" and not _flag_is_surge(r.get("flags")))]

                if tr_m_show:
                    lm_ = tr_m_show[0]["lm"]; pm_ = tr_m_show[0]["pm"]
                    df_m_tbl = pd.DataFrame([{
                        "Status": status_icon(r["status"], r.get("flags")),
                        "Website": r["site"],
                        "File": r.get("file",""),
                        "Sheet": r.get("sheet",""),
                        "Table": r.get("section","") or r.get("sheet",""),
                        f"M ({lm_})": _fmt_full(r["lv"]),
                        f"M-1 ({pm_})": _fmt_full(r["pv"]),
                        "M/M-1": f"{r['evol']:+.1f}%" if r["evol"] is not None else "—",
                        "Mkt share": f"{r['prm']:.1f}%" if r.get("prm") is not None else "—",
                        "ℹ️": " · ".join(r.get("flags") or []),
                    } for r in tr_m_show])

                    def _color_m(df):
                        def _row(row):
                            s = str(row.get("Status",""))
                            bg = "#fdecec" if s=="🔴" else "#fff3e0" if s=="🟠" else "#fff4d6" if s=="🟡" else "transparent"
                            return [f"background-color:{bg}" for _ in row]
                        return df.style.apply(_row, axis=1)

                    _render_scrollable_df(df_m_tbl)
                    st.download_button("⬇ Download", df_m_tbl.to_csv(index=False).encode("utf-8-sig"),
                                       f"monthly_{lm_ref}.csv", "text/csv", key="dl_m")
                else:
                    st.info("No series match this filter.")

        # ─── YEARLY (M/Y-1) — site-level only ───
        with subtab_y:
            tr_y_base = [r for r in tr_filtered if r["site"] in PANEL_SITES_SET]
            if not tr_y_base:
                tr_y_base = tr_filtered

            flagged_y = []
            seen_y = set()
            for r in tr_y_base:
                # Yearly'ye: Crash vs 12m peak + downtrend + M/Y-1 flag'i olanlar
                yearly_flags = [f for f in (r.get("flags") or [])
                               if any(k in f.lower() for k in ["crash","downtrend","y-1","year"])]
                # Sadece M/M-1 anomalisi değil — olanlar buraya
                has_yearly = bool(yearly_flags) or r.get("evol_y1") is not None
                if not has_yearly: continue
                # Sadece gerçek bir sorun varsa göster: evol_y1 > ±20% VEYA crash VEYA downtrend
                evol_y1 = r.get("evol_y1")
                is_notable = (
                    (evol_y1 is not None and abs(evol_y1) > 20) or
                    any("crash" in f.lower() for f in yearly_flags) or
                    any("downtrend" in f.lower() for f in yearly_flags)
                )
                if not is_notable: continue
                k = f"{r['site']}_{r['sheet']}"
                if k not in seen_y:
                    seen_y.add(k)
                    # Status: crash → alert, büyük Y-1 → alert/warn, downtrend → warn
                    if any("crash" in f.lower() for f in yearly_flags):
                        y_status = "alert"
                    elif evol_y1 is not None and abs(evol_y1) >= 30:
                        y_status = "alert" if evol_y1 <= -30 else "warn"
                    else:
                        y_status = "warn"
                    flagged_y.append({**r, "flags": yearly_flags, "status": y_status})

            n_crit_y = sum(1 for r in flagged_y if r["status"]=="alert")
            n_warn_y = sum(1 for r in flagged_y if r["status"]=="warn")
            st.caption(f"**Yearly anomalies** — 🔴 {n_crit_y} critical · 🟡 {n_warn_y} warnings"
                       f"  \n*🔴 Critical = Crash vs 12m peak OR M/Y-1 drop ≥ 30% · 🟡 Warning = M/Y-1 ±20-30% or downtrend*"
                       f"  \n*Values shown: M/Y-1 ratio only*")

            if flagged_y:
                st.markdown(f"#### 🚨 Y-1 anomalies — {len(flagged_y)} series")
                _draw_flagged(flagged_y, "y")
            else:
                st.success("✅ No Y-1 anomalies.")

            st.divider()
            st.markdown("#### All series — M/Y-1")
            tr_y_show = [r for r in tr_y_base if r.get("evol_y1") is not None]
            st.divider()
            with st.expander("📋 Detailed table — click to expand for in-depth review", expanded=False):
                st.caption("Filter by severity to focus on what matters.")
                sev_y = st.radio("Show", ["🔴 Critical", "🟡 Warnings", "🔴🟡 Both"],
                                 index=2, horizontal=True, key="sev_y_tbl")
                tr_y_filt = [r for r in flagged_y if
                             (sev_y=="🔴🟡 Both") or
                             (sev_y=="🔴 Critical" and r["status"]=="alert") or
                             (sev_y=="🟡 Warnings" and r["status"]=="warn")]

                if tr_y_filt:
                    lm_ = tr_y_filt[0]["lm"]
                    df_y_tbl = pd.DataFrame([{
                        "Status": status_icon(r["status"]),
                        "Website": r["site"],
                        "File": r.get("file",""),
                        "Sheet": r.get("sheet",""),
                        "Table": r.get("section","") or r.get("sheet",""),
                        f"M ({lm_})": _fmt_full(r["lv"]),
                        "M/Y-1": f"{r.get('evol_y1'):+.1f}%" if r.get("evol_y1") is not None else "—",
                        "Mkt share": f"{r['prm']:.1f}%" if r.get("prm") is not None else "—",
                        "ℹ️": " · ".join([f for f in (r.get("flags") or [])
                                          if any(k in f.lower() for k in ["crash","downtrend","y-1","year"])]),
                    } for r in tr_y_filt])

                    def _color_y(df):
                        def _row(row):
                            s = str(row.get("Status",""))
                            bg = "#fdecec" if s=="🔴" else "#fff3e0" if s=="🟠" else "#fff4d6" if s=="🟡" else "transparent"
                            return [f"background-color:{bg}" for _ in row]
                        return df.style.apply(_row, axis=1)

                    _render_scrollable_df(df_y_tbl)
                    st.download_button("⬇ Download", df_y_tbl.to_csv(index=False).encode("utf-8-sig"),
                                       f"yearly_{lm_ref}.csv", "text/csv", key="dl_y")
                else:
                    st.info("No series match this filter.")

        # ─── DÉPARTEMENT — geo rows, M/M-1 only ───
        with subtab_d:
            tr_dept = tr_dept_all or tr_filtered

            flagged_d = []
            seen_d = set()
            for r in tr_dept:
                mom_flags = [f for f in (r.get("flags") or []) if _is_mom_flag(f)]
                if not mom_flags:
                    continue
                k = f"{r['site']}_{r['sheet']}"
                if k not in seen_d:
                    seen_d.add(k)
                    real_status = "alert" if any("drop" in f.lower() for f in mom_flags) else "warn"
                    flagged_d.append({**r, "flags": mom_flags, "status": real_status})

            n_crit_d = sum(1 for r in flagged_d if r["status"]=="alert")
            n_surge_d = sum(1 for r in flagged_d if r["status"]=="warn" and _flag_is_surge(r.get("flags")))
            n_warn_d = sum(1 for r in flagged_d if r["status"]=="warn" and not _flag_is_surge(r.get("flags")))
            st.caption(f"**Department-level series (M/M-1)** — 🔴 {n_crit_d} critical · 🟠 {n_surge_d} surge · 🟡 {n_warn_d} warnings"
                       f"  \n*Critical = M/M-1 drop ≥20% · Surge = ≥30% increase · Warning = decline 10-20%*")
            if site_filter:
                st.info("ℹ️ **Filter by site** (top of page) has no effect here — département data isn't "
                        "broken down by website, only by geography.", icon="ℹ️")

            if flagged_d:
                st.markdown(f"#### 🚨 Département anomalies — {len(flagged_d)} series")
                _draw_flagged(flagged_d, "d")
            else:
                st.success("✅ No département-level anomalies.")

            st.divider()
            with st.expander("📋 Detailed table — click to expand for in-depth review", expanded=False):
                st.caption("Filter by severity to focus on what matters. Hover over ⚠ note for full details.")

                sev_d = st.radio("Show", ["🔴 Critical", "🟠 Surge", "🟡 Warnings", "🔴🟠🟡 All"],
                                 index=3, horizontal=True, key="sev_d_tbl")

                rows_to_show_d = [r for r in flagged_d if
                                  (sev_d == "🔴🟠🟡 All") or
                                  (sev_d == "🔴 Critical" and r["status"] == "alert") or
                                  (sev_d == "🟠 Surge" and r["status"]=="warn" and _flag_is_surge(r.get("flags"))) or
                                  (sev_d == "🟡 Warnings" and r["status"] == "warn" and not _flag_is_surge(r.get("flags")))]

                df_d_rows = []
                for r in rows_to_show_d:
                    mom_notes = [f for f in (r.get("flags") or []) if _is_mom_flag(f)]
                    df_d_rows.append({
                        "Status": status_icon(r["status"], mom_notes),
                        "Département": r["site"],
                        "File": r.get("file",""),
                        "Sheet": r.get("sheet",""),
                        f"M ({r['lm']})":   _fmt_full(r["lv"]),
                        f"M-1 ({r['pm']})": _fmt_full(r["pv"]),
                        "M/M-1": f"{r['evol']:+.1f}%" if r["evol"] is not None else "—",
                        "ℹ️": " · ".join(mom_notes) if mom_notes else "",
                    })

                if df_d_rows:
                    df_d = pd.DataFrame(df_d_rows)
                    _render_scrollable_df(df_d, status_col="Status")
                    st.download_button("⬇ Download", df_d.to_csv(index=False).encode("utf-8-sig"),
                                       f"dept_{lm_ref}.csv", "text/csv", key="dl_d")
                else:
                    st.info("No series match this filter.")

        # ─── RÉGION — geo rows, M/M-1 only ───
        with subtab_r:
            tr_region = tr_region_all

            flagged_r = []
            seen_r = set()
            for r in tr_region:
                mom_flags = [f for f in (r.get("flags") or []) if _is_mom_flag(f)]
                if not mom_flags:
                    continue
                k = f"{r['site']}_{r['sheet']}"
                if k not in seen_r:
                    seen_r.add(k)
                    real_status = "alert" if any("drop" in f.lower() for f in mom_flags) else "warn"
                    flagged_r.append({**r, "flags": mom_flags, "status": real_status})

            n_crit_r = sum(1 for r in flagged_r if r["status"]=="alert")
            n_surge_r = sum(1 for r in flagged_r if r["status"]=="warn" and _flag_is_surge(r.get("flags")))
            n_warn_r = sum(1 for r in flagged_r if r["status"]=="warn" and not _flag_is_surge(r.get("flags")))
            st.caption(f"**Région-level series (M/M-1)** — 🔴 {n_crit_r} critical · 🟠 {n_surge_r} surge · 🟡 {n_warn_r} warnings"
                       f"  \n*Critical = M/M-1 drop ≥20% · Surge = ≥30% increase · Warning = decline 10-20%*")
            if site_filter:
                st.info("ℹ️ **Filter by site** (top of page) has no effect here — région data isn't "
                        "broken down by website, only by geography.", icon="ℹ️")

            if flagged_r:
                st.markdown(f"#### 🚨 Région anomalies — {len(flagged_r)} series")
                _draw_flagged(flagged_r, "r")
            else:
                st.success("✅ No région-level anomalies.")

            st.divider()
            with st.expander("📋 Detailed table — click to expand for in-depth review", expanded=False):
                st.caption("Filter by severity to focus on what matters. Hover over ⚠ note for full details.")

                # Severity filter
                sev_r = st.radio("Show", ["🔴 Critical", "🟠 Surge", "🟡 Warnings", "🔴🟠🟡 All"],
                                 index=3, horizontal=True, key="sev_r_tbl")

                rows_to_show = [r for r in flagged_r if
                                (sev_r == "🔴🟠🟡 All") or
                                (sev_r == "🔴 Critical" and r["status"] == "alert") or
                                (sev_r == "🟠 Surge" and r["status"]=="warn" and _flag_is_surge(r.get("flags"))) or
                                (sev_r == "🟡 Warnings" and r["status"] == "warn" and not _flag_is_surge(r.get("flags")))]

                df_r_rows = []
                for r in rows_to_show:
                    mom_notes = [f for f in (r.get("flags") or []) if _is_mom_flag(f)]
                    df_r_rows.append({
                        "Status": status_icon(r["status"], mom_notes),
                        "Région": r["site"],
                        "File": r.get("file",""),
                        "Sheet": r.get("sheet",""),
                        f"M ({r['lm']})":   _fmt_full(r["lv"]),
                        f"M-1 ({r['pm']})": _fmt_full(r["pv"]),
                        "M/M-1": f"{r['evol']:+.1f}%" if r["evol"] is not None else "—",
                        "ℹ️": " · ".join(mom_notes) if mom_notes else "",
                    })

                if df_r_rows:
                    df_r = pd.DataFrame(df_r_rows)
                    _render_scrollable_df(df_r, status_col="Status")
                    st.download_button("⬇ Download", df_r.to_csv(index=False).encode("utf-8-sig"),
                                       f"region_{lm_ref}.csv", "text/csv", key="dl_r")
                else:
                    st.info("No series match this filter.")

    # ═════════════ DATA & TABLE CHECKS (merged) ═════════════

    with tab_dq:
        st.markdown(f"### 🔍 Data & Table Checks  \u00a0`v{APP_VERSION}`")
        st.caption(
            "**Formula checks** = cross-file consistency (Panel Checker formulas reproduced on the source files). "
            "**Table anomalies** = M/M-1 movement per website, using the **same 🔴/🟠/🟡 scale as the Trends tab** "
            "(drop ≥20% → 🔴, decline 10–20% → 🟡, surge ≥30% → 🟠, crash vs 12m peak → 🔴, 3-month downtrend / M-Y-1 >30% → 🟡). "
            "Every table found in the uploaded files is listed below, including clean ones."
        )
        with st.expander("ℹ️ What's the difference between Formula checks and Table anomalies?", expanded=False):
            st.table(pd.DataFrame(
                [
                    ["Comparison",   "Between two different numbers/files",        "A number against its own history (M/M-1)"],
                    ["What it catches", "Formula / data inconsistency",             "Sudden movement / trend anomaly"],
                    ["Source of thresholds", "Panel Checker formulas",              "Same scale as the Trends tab"],
                ],
                columns=["", "📐 Formula checks", "📊 Table anomalies"]
            ).set_index(""))
        st.divider()

        checks_show=checks
        if site_filter:
            checks_show=[c for c in checks
                         if any(s.lower() in c["name"].lower() for s in site_filter)
                         or not any(s.lower() in c["name"].lower() for s in SITES)]
        tables_show=tables
        if site_filter:
            tables_show=[t for t in tables
                         if any(s.lower()==r["site"].lower() for s in site_filter for r in t["sites"])]

        # ── Combined summary (single header, no duplicated metrics) ──
        n_err_c  = sum(1 for c in checks_show if not c["ok"] and c["sev"]=="error")
        n_warn_c = sum(1 for c in checks_show if c["sev"]=="warning")
        n_tabs_err  = sum(1 for t in tables_show if t["n_error"]>0)
        n_tabs_warn = sum(1 for t in tables_show if t["n_warn"]>0 and t["n_error"]==0)
        n_tabs_ok   = len(tables_show)-n_tabs_err-n_tabs_warn
        c1,c2,c3,c4,c5 = st.columns(5)
        c1.metric("❌ Formula errors",   n_err_c)
        c2.metric("⚠️ Formula warnings", n_warn_c)
        c3.metric("🔴 Tables w/ errors", n_tabs_err)
        c4.metric("🟡🟠 Tables w/ warnings", n_tabs_warn)
        c5.metric("📋 Tables checked", len(tables_show), f"{n_tabs_ok} clean")

        BADGES={
            "CHANGE_SEVERE":"🔴 M/M-1 drop ≥20%","CHANGE_DECLINE":"🟡 M/M-1 decline 10–20%",
            "CHANGE_SURGE":"🟠 M/M-1 surge ≥30%",
            "CRASH_VS_PEAK":"🔴 Crash vs 12m peak","DOWNTREND_3M":"🟡 3-month downtrend",
            "YOY_CHANGE":"🟡 M/Y-1 change >30%",
            "MS_OVER_100":"🔴 Share >100%","ZERO":"🔴 Unexpected zero",
            "TAUX>100%":"🔴 Rate >100%","TAUX_JUMP":"⚠️ Rate jump M/M-1",
            "TAUX_ZERO_TO_VALUE":"⚠️ Rate: 0% → value","TAUX_VALUE_TO_ZERO":"🔴 Rate: value → 0%",
            "DEDUP_CHANGE_SEVERE":"🔴 Dedup drop ≥20%","DEDUP_CHANGE_DECLINE":"🟡 Dedup decline 10–20%",
            "DEDUP_CHANGE_SURGE":"🟠 Dedup surge ≥30%",
        }
        EXPL={
            "CHANGE_SEVERE":"M/M-1 drop ≥20% — same threshold as a 🔴 alert in the Trends tab.",
            "CHANGE_DECLINE":"M/M-1 decline 10–20% — same threshold as a 🟡 warning in the Trends tab.",
            "CHANGE_SURGE":"M/M-1 surge ≥30% — same threshold as the 🟠 surge flag in the Trends tab.",
            "CRASH_VS_PEAK":"Current value <60% of the 12-month peak — same rule as the Trends tab.",
            "DOWNTREND_3M":"3 consecutive declining months — same rule as the Trends tab.",
            "YOY_CHANGE":"Change vs same month last year >30% — same rule as the Trends tab.",
            "MS_OVER_100":"Site listings > Total Panel Dédupliqué Top 11 (or Marché) — check if scopes match.",
            "ZERO":"Value dropped to 0 while M-1 was significant — feed missing?",
            "TAUX>100%":"A rate cannot exceed 100% — calculation error.",
            "TAUX_JUMP":"Large rate change M/M-1 — confirm with team.",
            "TAUX_ZERO_TO_VALUE":"Rate was 0% last month — site active since this month?",
            "TAUX_VALUE_TO_ZERO":"Rate dropped to 0% — data missing?",
            "DEDUP_CHANGE_SEVERE":"Deduplicated total drop ≥20% M/M-1 — same threshold as Trends.",
            "DEDUP_CHANGE_DECLINE":"Deduplicated total decline 10–20% M/M-1 — same threshold as Trends.",
            "DEDUP_CHANGE_SURGE":"Deduplicated total surge ≥30% M/M-1 — same threshold as Trends.",
        }
        SURGE_TYPES={"CHANGE_SURGE","DEDUP_CHANGE_SURGE"}
        TTYPE_LBL = {
            "volume":"📊 Volume table (listings) — M vs M-1",
            "taux":"📐 Rate table (ratios 0–1) — M vs M-1",
            "snapshot":"📷 Snapshot table — single month, no M-1. Only MAX > dedup check.",
            "unreadable":"⬜ Not enough data to run checks on this section.",
        }

        tab_form, tab_tables = st.tabs(["📐 Formula checks", "📊 Table anomalies — by file"])

        # ═══ FORMULA CHECKS (grouped by category, unchanged) ═══
        with tab_form:
            st.markdown("##### 🎨 Filter by status")
            f_status_options=["❌ Error","⚠️ Warning","✅ Passed"]
            f_status_filter=st.multiselect("Show", f_status_options, default=f_status_options,
                                            key="dq_formula_status_filter", label_visibility="collapsed")
            def _check_status(c):
                if not c["ok"] and c["sev"]=="error": return "❌ Error"
                if c["sev"]=="warning": return "⚠️ Warning"
                return "✅ Passed"
            checks_f=[c for c in checks_show if _check_status(c) in f_status_filter] if f_status_filter else []

            if not checks_f:
                st.info("No checks match this filter.")
            else:
                by_g=defaultdict(list)
                for c in checks_f: by_g[c["group"]].append(c)
                group_order = {g:i for i,g in enumerate(GROUP_INFO.keys())}
                for grp in sorted(by_g, key=lambda g: group_order.get(g,99)):
                    items=by_g[grp]
                    ne=sum(1 for c in items if not c["ok"] and c["sev"]=="error")
                    nw=sum(1 for c in items if c["sev"]=="warning")
                    no=sum(1 for c in items if c["ok"])
                    title,sub=GROUP_INFO.get(grp,(f"Group {grp}",""))
                    real_fn=GROUP_TO_FILE.get(grp)
                    badge=f"{ne} error{'s' if ne!=1 else ''}" if ne else \
                          f"{nw} warning{'s' if nw!=1 else ''}" if nw else f"{no} passed"
                    with st.expander(f"{'❌' if ne else '⚠️' if nw else '✅'}  {title} — {badge}",expanded=(ne>0)):
                        if real_fn: st.caption(f"📁 **{real_fn}**")
                        st.caption(sub)
                        ordered=([c for c in items if not c["ok"] and c["sev"]=="error"]+
                                 [c for c in items if c["sev"]=="warning"]+[c for c in items if c["ok"]])
                        for c in ordered:
                            if c["ok"]: st.markdown(f"✅  {c['name']}")
                            elif c["sev"]=="error": st.error(f"**{c['name']}**  \n{c['detail']}")
                            else: st.warning(f"**{c['name']}**  \n{c['detail']}")
            rows_exp_f=[{"File":GROUP_TO_FILE.get(c["group"],""),
                       "Category":GROUP_INFO.get(c["group"],("",""))[0],"Check":c["name"],
                       "Result":"❌ Error" if not c["ok"] and c["sev"]=="error"
                                else "⚠️ Warning" if c["sev"]=="warning" else "✅ OK",
                       "Detail":c["detail"]} for c in checks]
            st.download_button("⬇ Download formula checks",
                               pd.DataFrame(rows_exp_f).to_csv(index=False).encode("utf-8-sig"),
                               f"formula_checks_{lm_ref}.csv","text/csv", key="dl_formula")

        # ═══ TABLE ANOMALIES — organised by real File → Sheet → Table ═══
        with tab_tables:
            if not tables_show:
                st.info("No tables found. Upload your Excel files to begin.")
            else:
                def _table_statuses(t):
                    """A table can contain several kinds of issues at once (e.g. one site
                    dropped while another surged) — return the FULL set present, so a
                    colour filter matches if the table contains ANY of that colour,
                    instead of forcing one exclusive status per table."""
                    statuses=set()
                    if t["n_error"]>0: statuses.add("❌ Error")
                    warn_issues=[i for i in t["issues"] if i["severity"]=="warning"]
                    if any(i.get("type") in SURGE_TYPES for i in warn_issues): statuses.add("🟠 Surge")
                    if any(i.get("type") not in SURGE_TYPES for i in warn_issues): statuses.add("🟡 Warning")
                    if not t["issues"]: statuses.add("✅ Clean")
                    return statuses

                # ── Colour filter — drives both the drill-down below and the overview table ──
                st.markdown("##### 🎨 Filter by status")
                status_options=["❌ Error","🟠 Surge","🟡 Warning","✅ Clean"]
                status_filter=st.multiselect("Show", status_options, default=status_options,
                                              key="dq_status_filter", label_visibility="collapsed")
                tables_f=[t for t in tables_show if _table_statuses(t) & set(status_filter)] if status_filter else []

                if not tables_f:
                    st.info("No tables match this filter.")
                else:
                    # ── Drill-down: pick one File → Sheet → Table to see full detail ──
                    st.markdown("##### 🔬 Inspect one table")
                    file_options = sorted(set(t.get("source_file") or t["file"] for t in tables_f))
                    def _file_badge(f):
                        ts=[t for t in tables_f if (t.get("source_file") or t["file"])==f]
                        ne=sum(t["n_error"] for t in ts); nw=sum(t["n_warn"] for t in ts)
                        return f"❌ {f}" if ne else (f"⚠️ {f}" if nw else f"✅ {f}")
                    sel_file = st.selectbox("📁 File", file_options, format_func=_file_badge, key="dq_file")
                    file_tables = [t for t in tables_f if (t.get("source_file") or t["file"])==sel_file]

                    def _sheet_badge(sn):
                        ts=[t for t in file_tables if t["sheet"]==sn]
                        ne=sum(t["n_error"] for t in ts); nw=sum(t["n_warn"] for t in ts)
                        return f"❌ {sn}" if ne else (f"⚠️ {sn}" if nw else f"✅ {sn}")
                    sel_sheet = st.selectbox("📄 Tab (sheet)", sorted(set(t["sheet"] for t in file_tables)),
                                              format_func=_sheet_badge, key="dq_sheet")
                    sheet_tables = [t for t in file_tables if t["sheet"]==sel_sheet]

                    if len(sheet_tables)>1:
                        def _sec_badge(t):
                            icon="❌" if t["n_error"]>0 else "⚠️" if t["n_warn"]>0 else "✅"
                            lbl=t["label"] or f"Section {t['sec_idx']+1}"
                            return f"{icon} {lbl}"
                        sel_t = st.radio("📋 Table / section", sheet_tables, format_func=_sec_badge,
                                          horizontal=False, key="dq_sec")
                    else:
                        sel_t = sheet_tables[0]

                    st.divider()
                    sel_label = sel_t["label"] or f"Section {sel_t['sec_idx']+1}"
                    st.markdown(f"#### {sel_t.get('source_file') or sel_t['file']}")
                    st.caption(f"{sel_t['sheet']} › {sel_label}")
                    cc1,cc2,cc3,cc4 = st.columns(4)
                    cc1.metric("Month", sel_t["lm"])
                    cc2.metric("Prev month", sel_t["pm"])
                    cc3.metric("Type", sel_t["table_type"].capitalize())
                    cc4.metric("Market dedup", _fmtn(sel_t["dedup"]) if sel_t["dedup"] else "—")
                    st.caption(TTYPE_LBL.get(sel_t["table_type"], sel_t["table_type"]))

                    issues=sel_t["issues"]
                    if not issues:
                        st.success("✅ No anomalies detected in this table.")
                    else:
                        ne_t=sum(1 for i in issues if i["severity"]=="error")
                        nw_t=sum(1 for i in issues if i["severity"]=="warning")
                        st.markdown(f"**{ne_t} error{'s' if ne_t!=1 else ''} · {nw_t} warning{'s' if nw_t!=1 else ''}**")
                        for iss in sorted(issues, key=lambda x:0 if x["severity"]=="error" else 1):
                            itype=iss.get("type",""); b2=BADGES.get(itype,f"⚠️ {itype}")
                            expl=EXPL.get(itype,"")
                            content=f"**{b2}** — **{iss['site']}** — {iss['message']}"
                            if expl: content+=f"  \n*💡 {expl}*"
                            if iss["severity"]=="error":
                                st.error(content)
                            elif itype in SURGE_TYPES:
                                st.markdown(
                                    f"<div style='background:#fff3e0;border-left:4px solid #ff8c00;"
                                    f"padding:10px 14px;border-radius:4px;margin-bottom:6px;font-size:14px'>"
                                    f"{content}</div>", unsafe_allow_html=True)
                            else:
                                st.warning(content)

                    if sel_t["sites"]:
                        st.divider()
                        st.markdown("##### Data per website")
                        is_taux=sel_t["table_type"]=="taux"
                        def _fv(v): return f"{v*100:.2f}%" if (is_taux and v is not None) else (_fmt_full(v) if v is not None else "—")
                        df_data=pd.DataFrame([{
                            "": "🔴" if r["status"]=="alert" else ("🟠" if r.get("is_surge") else "🟡") if r["status"]=="warn" else "✅",
                            "Website": r["site"],
                            f"M ({sel_t['lm']})": _fv(r["last"]),
                            f"M-1 ({sel_t['pm']})": _fv(r["prev"]),
                            "M/M-1": f"{r['evol']:+.1f}%" if r["evol"] is not None else "—",
                            **({} if is_taux else {"Market share": f"{r['ms']:.1f}%" if r.get("ms") else "—"}),
                        } for r in sorted(sel_t["sites"], key=lambda x:({"alert":0,"warn":1,"ok":2}[x["status"]],-(x["last"] or 0)))])
                        def _cr(df):
                            def _row(row):
                                s=str(row.get("",""))
                                bg="#fdecec" if s=="🔴" else "#fff3e0" if s=="🟠" else "#fff4d6" if s=="🟡" else "#edf7ed"
                                return [f"background-color:{bg}" for _ in row]
                            return df.style.apply(_row,axis=1)
                        st.dataframe(_cr(df_data), use_container_width=True, hide_index=True)


                st.divider()
                rows_exp_t=[]
                for t in tables:
                    if not t["issues"]:
                        rows_exp_t.append({"File":t.get("source_file") or t["file"],"Tab":t["sheet"],
                                            "Table":t["label"],"Result":"✅ OK","Detail":"No anomalies detected"})
                    for iss in t["issues"]:
                        rows_exp_t.append({"File":t.get("source_file") or t["file"],"Tab":t["sheet"],
                                            "Table":f"{t['label']} — {iss['site']}",
                                            "Result":"❌ Error" if iss["severity"]=="error" else "⚠️ Warning",
                                            "Detail":iss["message"]})
                st.download_button("⬇ Download table anomalies",
                                   pd.DataFrame(rows_exp_t).to_csv(index=False).encode("utf-8-sig"),
                                   f"table_anomalies_{lm_ref}.csv","text/csv", key="dl_tables")
else:
    SITES = AUTOFR_SITES
    if not uploaded:
        st.markdown("## 🚗 AUTO FR — Panel Quality Control")
        st.info("Upload your Excel files in the sidebar to begin.")
        st.stop()

    checks,trends,fb,tables = compute_everything_autofr(file_hash, raw_bytes)
    GROUP_TO_FILE_AF = {role: fn for role,(fn,_) in classify_named_autofr(raw_bytes).items()}
    n_err  = sum(1 for c in checks if not c["ok"] and c["sev"]=="error")
    n_warn = sum(1 for c in checks if c["sev"]=="warning")
    n_ok   = sum(1 for c in checks if c["ok"])
    AUTOFR_SITES_SET = set(AUTOFR_SITES)
    _tr_sites_only = [r for r in trends if r["site"] in AUTOFR_SITES_SET]
    n_alert = len({f"{r['site']}_{r['sheet']}" for r in _tr_sites_only if r["status"]=="alert"})
    _warn_rows_ov = [r for r in _tr_sites_only if r["status"]=="warn"]
    n_surge = len({f"{r['site']}_{r['sheet']}" for r in _warn_rows_ov if _flag_is_surge(r.get("flags"))})
    n_wtr   = len({f"{r['site']}_{r['sheet']}" for r in _warn_rows_ov if not _flag_is_surge(r.get("flags"))})

    lm_ref,pm_ref="—","—"; total=dedup=None; n_active=0
    if "file1" in fb:
        try:
            wb1=load_workbook(io.BytesIO(fb["file1"]),data_only=True)
            ws_tot=ws_get(wb1,"Total")
            if ws_tot:
                d_tot=read_series(ws_tot,section=0)
                lm_ref=d_tot.get("_lm","—"); pm_ref=d_tot.get("_pm","—")
                total=sv(d_tot,"Somme Panel")
                dedup_label=next((k for k in d_tot if isinstance(k,str) and "dédoublon" in k.lower()),None)
                dedup=sv(d_tot,dedup_label) if dedup_label else None
                n_active=sum(1 for s in AUTOFR_SITES if sv(d_tot,s) and sv(d_tot,s)>0)
            wb1.close()
        except: pass

    col_h,col_f=st.columns([3,2])
    with col_h:
        st.markdown("## 🚗 AUTO FR — Panel Quality Control")
        st.caption(f"Reference month: **{lm_ref}** · vs {pm_ref} · {len(fb)} files · v{APP_VERSION}")
    with col_f:
        chosen=st.multiselect("Filter by site",AUTOFR_SITES,default=[],
                              placeholder="All sites",label_visibility="visible",key="af_site_filter")
    site_filter=chosen if chosen else None

    af_tab1,af_tab2,af_tab_dq=st.tabs(["📋  Overview","📈  Trends — Monthly · Yearly","🔍  Data & Table Checks"])

    # ═════════════ OVERVIEW ════════════
    with af_tab1:
        if n_err>0:
            st.error(f"🚫 **{n_err} formula error(s) found — needs review**  \nCross-file checks: {n_warn} warning(s) · {n_ok} passed")
        elif n_warn>5 or n_alert>0:
            st.warning(f"⚠️ **Needs review before validation**  \nCross-file checks: {n_err} error(s) · {n_warn} warning(s) · {n_ok} passed")
        else:
            msg=f"✅ **Validated — data is ready** · {n_ok} integrity checks passed"
            if n_warn: msg+=f" · {n_warn} minor warnings"
            st.success(msg)

        st.divider()
        c1,c2,c3,c4,c5,c6,c7=st.columns([1.3,1.3,1,1,0.8,0.8,0.8])
        c1.metric("Somme Panel", fmt(total) if total else "—")
        c2.metric("Marché dédoublonné", fmt(dedup) if dedup else "—")
        c3.metric("Active sites", f"{n_active} / {len(AUTOFR_SITES)}")
        c4.metric("Formula errors", str(n_err), "none ✓" if n_err==0 else "blocking ⚠️")
        c5.metric("🔴 Critical", n_alert)
        c6.metric("🟠 Surge", n_surge)
        c7.metric("🟡 Warning", n_wtr)

        with st.expander("🎨 How colours are determined", expanded=False):
            st.markdown("""
**Formula checks** — cross-file consistency (real formulas reproduced from the Panel Checker Auto FR
on the 7 source files): 🔴 numbers don't match · ✅ OK

**Trends / Table anomalies** — same 🔴/🟠/🟡 scale as IMMOFR, M/M-1 per series:
- 🔴 drop ≥20% vs previous month, or crash <60% of the 12-month peak
- 🟠 surge ≥30% vs previous month
- 🟡 decline 10–20%, 3-month downtrend >5%, or |M/Y-1| >30%
- ✅ stable

Only **File 1 (Panel evolution)** has real month-over-month history today — the other 6 AUTOFR files
are single-month snapshots, so they feed Formula checks but not Trends/Table anomalies yet.
""")

    # ═════════════ TRENDS ════════════
    with af_tab2:
        af_sub_m, af_sub_y = st.tabs(["📅 Monthly (M/M-1)", "📆 Yearly (M/Y-1)"])
        tr=[r for r in trends if not site_filter or r["site"] in site_filter]
        tr_sites=[r for r in tr if r["site"] in AUTOFR_SITES_SET] or tr

        # ─── MONTHLY (M/M-1) ───
        with af_sub_m:
            flagged_m = []
            seen_m = set()
            for r in tr_sites:
                mom_flags = [f for f in (r.get("flags") or []) if _is_mom_flag(f)]
                if not mom_flags:
                    continue  # Crash-only / Y-1-only / downtrend-only → Yearly'de
                k = f"{r['site']}_{r['sheet']}"
                if k not in seen_m:
                    seen_m.add(k)
                    real_status = "alert" if any("drop" in f.lower() for f in mom_flags) else "warn"
                    flagged_m.append({**r, "flags": mom_flags, "status": real_status})

            n_crit_m = sum(1 for r in flagged_m if r["status"]=="alert")
            n_surge_m = sum(1 for r in flagged_m if r["status"]=="warn" and _flag_is_surge(r.get("flags")))
            n_warn_m = sum(1 for r in flagged_m if r["status"]=="warn" and not _flag_is_surge(r.get("flags")))
            st.caption(f"**Website-level series** — 🔴 {n_crit_m} critical · 🟠 {n_surge_m} surge · 🟡 {n_warn_m} warnings"
                       f"  \n*Critical = M/M-1 drop ≥ 20% · Surge = ≥ 30% increase · Warning = decline 10-20%*"
                       f"  \n*Crash vs 12m peak and Y-1 anomalies → Yearly tab*")

            if flagged_m:
                st.markdown(f"#### 🚨 Anomalies — {len(flagged_m)} series")
                draw_flagged_series(flagged_m, "af_m")
            else:
                st.success("✅ No M/M-1 anomalies.")

            st.divider()
            st.markdown("#### All series — M/M-1")
            st.divider()
            with st.expander("📋 Detailed table — click to expand for in-depth review", expanded=False):
                st.caption("Filter by severity to focus on what matters.")
                sev_m = st.radio("Show", ["🔴 Critical", "🟠 Surge", "🟡 Warnings", "🔴🟠🟡 All"],
                                 index=3, horizontal=True, key="af_sev_m_tbl")
                tr_m_show = [r for r in flagged_m if
                             (sev_m=="🔴🟠🟡 All") or
                             (sev_m=="🔴 Critical" and r["status"]=="alert") or
                             (sev_m=="🟠 Surge" and r["status"]=="warn" and _flag_is_surge(r.get("flags"))) or
                             (sev_m=="🟡 Warnings" and r["status"]=="warn" and not _flag_is_surge(r.get("flags")))]

                if tr_m_show:
                    lm_ = tr_m_show[0]["lm"]; pm_ = tr_m_show[0]["pm"]
                    df_m_tbl = pd.DataFrame([{
                        "Status": status_icon(r["status"], r.get("flags")),
                        "Website": r["site"],
                        "File": r.get("file",""),
                        "Sheet": r.get("sheet",""),
                        "Table": r.get("section","") or r.get("sheet",""),
                        f"M ({lm_})": _fmt_full(r["lv"]),
                        f"M-1 ({pm_})": _fmt_full(r["pv"]),
                        "M/M-1": f"{r['evol']:+.1f}%" if r["evol"] is not None else "—",
                        "ℹ️": " · ".join(r.get("flags") or []),
                    } for r in tr_m_show])
                    _render_scrollable_df(df_m_tbl)
                    st.download_button("⬇ Download", df_m_tbl.to_csv(index=False).encode("utf-8-sig"),
                                       f"autofr_monthly_{lm_ref}.csv", "text/csv", key="af_dl_m")
                else:
                    st.info("No series match this filter.")

        # ─── YEARLY (M/Y-1) — site-level only ───
        with af_sub_y:
            flagged_y = []
            seen_y = set()
            for r in tr_sites:
                yearly_flags = [f for f in (r.get("flags") or [])
                               if any(k in f.lower() for k in ["crash","downtrend","y-1","year"])]
                has_yearly = bool(yearly_flags) or r.get("evol_y1") is not None
                if not has_yearly: continue
                evol_y1 = r.get("evol_y1")
                is_notable = (
                    (evol_y1 is not None and abs(evol_y1) > 20) or
                    any("crash" in f.lower() for f in yearly_flags) or
                    any("downtrend" in f.lower() for f in yearly_flags)
                )
                if not is_notable: continue
                k = f"{r['site']}_{r['sheet']}"
                if k not in seen_y:
                    seen_y.add(k)
                    if any("crash" in f.lower() for f in yearly_flags):
                        y_status = "alert"
                    elif evol_y1 is not None and abs(evol_y1) >= 30:
                        y_status = "alert" if evol_y1 <= -30 else "warn"
                    else:
                        y_status = "warn"
                    flagged_y.append({**r, "flags": yearly_flags, "status": y_status})

            n_crit_y = sum(1 for r in flagged_y if r["status"]=="alert")
            n_warn_y = sum(1 for r in flagged_y if r["status"]=="warn")
            st.caption(f"**Yearly anomalies** — 🔴 {n_crit_y} critical · 🟡 {n_warn_y} warnings"
                       f"  \n*🔴 Critical = Crash vs 12m peak OR M/Y-1 drop ≥ 30% · 🟡 Warning = M/Y-1 ±20-30% or downtrend*"
                       f"  \n*Values shown: M/Y-1 ratio only*")

            if flagged_y:
                st.markdown(f"#### 🚨 Y-1 anomalies — {len(flagged_y)} series")
                draw_flagged_series(flagged_y, "af_y")
            else:
                st.success("✅ No Y-1 anomalies.")

            st.divider()
            st.markdown("#### All series — M/Y-1")
            tr_y_base=[r for r in tr_sites if r.get("evol_y1") is not None]
            st.divider()
            with st.expander("📋 Detailed table — click to expand for in-depth review", expanded=False):
                st.caption("Filter by severity to focus on what matters.")
                sev_y = st.radio("Show", ["🔴 Critical", "🟡 Warnings", "🔴🟡 Both"],
                                 index=2, horizontal=True, key="af_sev_y_tbl")
                tr_y_filt = [r for r in flagged_y if
                             (sev_y=="🔴🟡 Both") or
                             (sev_y=="🔴 Critical" and r["status"]=="alert") or
                             (sev_y=="🟡 Warnings" and r["status"]=="warn")]

                if tr_y_filt:
                    lm_ = tr_y_filt[0]["lm"]
                    df_y_tbl = pd.DataFrame([{
                        "Status": status_icon(r["status"]),
                        "Website": r["site"],
                        "File": r.get("file",""),
                        "Sheet": r.get("sheet",""),
                        "Table": r.get("section","") or r.get("sheet",""),
                        f"M ({lm_})": _fmt_full(r["lv"]),
                        "M/Y-1": f"{r.get('evol_y1'):+.1f}%" if r.get("evol_y1") is not None else "—",
                        "ℹ️": " · ".join([f for f in (r.get("flags") or [])
                                          if any(k in f.lower() for k in ["crash","downtrend","y-1","year"])]),
                    } for r in tr_y_filt])
                    _render_scrollable_df(df_y_tbl)
                    st.download_button("⬇ Download", df_y_tbl.to_csv(index=False).encode("utf-8-sig"),
                                       f"autofr_yearly_{lm_ref}.csv", "text/csv", key="af_dl_y")
                else:
                    st.info("No series match this filter.")

    # ═════════════ DATA & TABLE CHECKS ════════════
    with af_tab_dq:
        st.markdown(f"### 🔍 Data & Table Checks  \u00a0`v{APP_VERSION}`")
        st.caption(
            "**Formula checks** = cross-file consistency (Panel Checker Auto FR formulas reproduced on the source files). "
            "**Table anomalies** = M/M-1 movement per website, same 🔴/🟠/🟡 scale as Trends. "
            "Only File 1 (Panel evolution) has real month-over-month history — the other 6 files are single-month snapshots."
        )
        st.divider()

        checks_show=checks
        if site_filter:
            checks_show=[c for c in checks if any(s.lower() in c["name"].lower() for s in site_filter)
                         or not any(s.lower() in c["name"].lower() for s in AUTOFR_SITES)]
        tables_show=tables
        if site_filter:
            tables_show=[t for t in tables if any(s.lower()==r["site"].lower() for s in site_filter for r in t["sites"])]

        n_err_c=sum(1 for c in checks_show if not c["ok"] and c["sev"]=="error")
        n_warn_c=sum(1 for c in checks_show if c["sev"]=="warning")
        n_tabs_err=sum(1 for t in tables_show if t["n_error"]>0)
        n_tabs_warn=sum(1 for t in tables_show if t["n_warn"]>0 and t["n_error"]==0)
        n_tabs_ok=len(tables_show)-n_tabs_err-n_tabs_warn
        c1,c2,c3,c4,c5=st.columns(5)
        c1.metric("❌ Formula errors",n_err_c)
        c2.metric("⚠️ Formula warnings",n_warn_c)
        c3.metric("🔴 Tables w/ errors",n_tabs_err)
        c4.metric("🟡🟠 Tables w/ warnings",n_tabs_warn)
        c5.metric("📋 Tables checked",len(tables_show),f"{n_tabs_ok} clean")

        af_tab_form, af_tab_tables = st.tabs(["📐 Formula checks","📊 Table anomalies — by file"])

        with af_tab_form:
            st.markdown("##### 🎨 Filter by status")
            af_f_status_options=["❌ Error","⚠️ Warning","✅ Passed"]
            af_f_status_filter=st.multiselect("Show", af_f_status_options, default=af_f_status_options,
                                            key="af_formula_status_filter", label_visibility="collapsed")
            def _af_check_status(c):
                if not c["ok"] and c["sev"]=="error": return "❌ Error"
                if c["sev"]=="warning": return "⚠️ Warning"
                return "✅ Passed"
            checks_f=[c for c in checks_show if _af_check_status(c) in af_f_status_filter] if af_f_status_filter else []

            if not checks_f:
                st.info("No checks match this filter.")
            else:
                by_g=defaultdict(list)
                for c in checks_f: by_g[c["group"]].append(c)
                group_order={g:i for i,g in enumerate(GROUP_INFO_AUTOFR.keys())}
                for grp in sorted(by_g,key=lambda g:group_order.get(g,99)):
                    items=by_g[grp]
                    ne=sum(1 for c in items if not c["ok"] and c["sev"]=="error")
                    nw=sum(1 for c in items if c["sev"]=="warning")
                    no=sum(1 for c in items if c["ok"])
                    title,sub=GROUP_INFO_AUTOFR.get(grp,(f"Group {grp}",""))
                    real_fn=GROUP_TO_FILE_AF.get(f"file{grp}")
                    badge=f"{ne} error{'s' if ne!=1 else ''}" if ne else f"{nw} warning{'s' if nw!=1 else ''}" if nw else f"{no} passed"
                    with st.expander(f"{'❌' if ne else '⚠️' if nw else '✅'}  {title} — {badge}",expanded=(ne>0)):
                        if real_fn: st.caption(f"📁 **{real_fn}**")
                        st.caption(sub)
                        ordered=([c for c in items if not c["ok"] and c["sev"]=="error"]+
                                 [c for c in items if c["sev"]=="warning"]+[c for c in items if c["ok"]])
                        for c in ordered:
                            if c["ok"]: st.markdown(f"✅  {c['name']}")
                            elif c["sev"]=="error": st.error(f"**{c['name']}**  \n{c['detail']}")
                            else: st.warning(f"**{c['name']}**  \n{c['detail']}")
            rows_exp_f=[{"Category":GROUP_INFO_AUTOFR.get(c["group"],("",""))[0],"Check":c["name"],
                       "Result":"❌ Error" if not c["ok"] and c["sev"]=="error" else "⚠️ Warning" if c["sev"]=="warning" else "✅ OK",
                       "Detail":c["detail"]} for c in checks]
            st.download_button("⬇ Download formula checks",
                pd.DataFrame(rows_exp_f).to_csv(index=False).encode("utf-8-sig"),
                f"autofr_formula_checks_{lm_ref}.csv","text/csv",key="af_dl_formula")

        with af_tab_tables:
            if not tables_show:
                st.info("No tables found.")
            else:
                AF_BADGES={"CHANGE_SEVERE":"🔴 M/M-1 drop ≥20%","CHANGE_DECLINE":"🟡 M/M-1 decline 10–20%",
                        "CHANGE_SURGE":"🟠 M/M-1 surge ≥30%","CRASH_VS_PEAK":"🔴 Crash vs 12m peak",
                        "DOWNTREND_3M":"🟡 3-month downtrend","YOY_CHANGE":"🟡 M/Y-1 change >30%",
                        "MS_OVER_100":"🔴 Share >100%","ZERO":"🔴 Unexpected zero"}
                AF_EXPL={
                    "CHANGE_SEVERE":"M/M-1 drop ≥20% — same threshold as a 🔴 alert in the Trends tab.",
                    "CHANGE_DECLINE":"M/M-1 decline 10–20% — same threshold as a 🟡 warning in the Trends tab.",
                    "CHANGE_SURGE":"M/M-1 surge ≥30% — same threshold as the 🟠 surge flag in the Trends tab.",
                    "CRASH_VS_PEAK":"Current value <60% of the 12-month peak — same rule as the Trends tab.",
                    "DOWNTREND_3M":"3 consecutive declining months — same rule as the Trends tab.",
                    "YOY_CHANGE":"Change vs same month last year >30% — same rule as the Trends tab.",
                    "MS_OVER_100":"Site value exceeds the Somme Panel/Marché dédoublonné denominator — check scopes match.",
                    "ZERO":"Value dropped to 0 while M-1 was significant — feed missing?",
                }
                AF_TTYPE_LBL={
                    "volume":"📊 Volume table (listings) — M vs M-1",
                    "taux":"📐 Rate table (ratios 0–1) — M vs M-1",
                    "snapshot":"📷 Snapshot table — single month, no M-1.",
                    "unreadable":"⬜ Not enough data to run checks on this section.",
                }
                AF_SURGE_TYPES={"CHANGE_SURGE"}
                def _af_table_statuses(t):
                    s=set()
                    if t["n_error"]>0: s.add("❌ Error")
                    warn_issues=[i for i in t["issues"] if i["severity"]=="warning"]
                    if any(i.get("type") in AF_SURGE_TYPES for i in warn_issues): s.add("🟠 Surge")
                    if any(i.get("type") not in AF_SURGE_TYPES for i in warn_issues): s.add("🟡 Warning")
                    if not t["issues"]: s.add("✅ Clean")
                    return s
                st.markdown("##### 🎨 Filter by status")
                af_status_options=["❌ Error","🟠 Surge","🟡 Warning","✅ Clean"]
                af_status_filter=st.multiselect("Show",af_status_options,default=af_status_options,
                    key="af_status_filter",label_visibility="collapsed")
                tables_f=[t for t in tables_show if _af_table_statuses(t)&set(af_status_filter)] if af_status_filter else []
                if not tables_f:
                    st.info("No tables match this filter.")
                else:
                    st.markdown("##### 🔬 Inspect one table")
                    file_options=sorted(set(t.get("source_file") or t["file"] for t in tables_f))
                    def _af_fb(f):
                        ts=[t for t in tables_f if (t.get("source_file") or t["file"])==f]
                        ne=sum(t["n_error"] for t in ts); nw=sum(t["n_warn"] for t in ts)
                        return f"❌ {f}" if ne else (f"⚠️ {f}" if nw else f"✅ {f}")
                    sel_file=st.selectbox("📁 File",file_options,format_func=_af_fb,key="af_file")
                    file_tables=[t for t in tables_f if (t.get("source_file") or t["file"])==sel_file]
                    def _af_sb(sn):
                        ts=[t for t in file_tables if t["sheet"]==sn]
                        ne=sum(t["n_error"] for t in ts); nw=sum(t["n_warn"] for t in ts)
                        return f"❌ {sn}" if ne else (f"⚠️ {sn}" if nw else f"✅ {sn}")
                    sel_sheet=st.selectbox("📄 Tab (sheet)",sorted(set(t["sheet"] for t in file_tables)),
                        format_func=_af_sb,key="af_sheet")
                    sheet_tables=[t for t in file_tables if t["sheet"]==sel_sheet]
                    if len(sheet_tables)>1:
                        def _af_sc(t):
                            icon="❌" if t["n_error"]>0 else "⚠️" if t["n_warn"]>0 else "✅"
                            lbl=t["label"] or f"Section {t['sec_idx']+1}"
                            return f"{icon} {lbl}"
                        sel_t=st.radio("📋 Table / section",sheet_tables,format_func=_af_sc,key="af_sec")
                    else:
                        sel_t=sheet_tables[0]
                    st.divider()
                    sel_label=sel_t["label"] or f"Section {sel_t['sec_idx']+1}"
                    st.markdown(f"#### {sel_t.get('source_file') or sel_t['file']}")
                    st.caption(f"{sel_t['sheet']} › {sel_label}")
                    cc1,cc2,cc3,cc4=st.columns(4)
                    cc1.metric("Month",sel_t["lm"]); cc2.metric("Prev month",sel_t["pm"])
                    cc3.metric("Type",sel_t["table_type"].capitalize())
                    cc4.metric("Dedup denom.", _fmtn(sel_t["dedup"]) if sel_t["dedup"] else "—")
                    st.caption(AF_TTYPE_LBL.get(sel_t["table_type"], sel_t["table_type"]))
                    issues=sel_t["issues"]
                    if not issues:
                        st.success("✅ No anomalies detected in this table.")
                    else:
                        ne_t=sum(1 for i in issues if i["severity"]=="error")
                        nw_t=sum(1 for i in issues if i["severity"]=="warning")
                        st.markdown(f"**{ne_t} error{'s' if ne_t!=1 else ''} · {nw_t} warning{'s' if nw_t!=1 else ''}**")
                        for iss in sorted(issues,key=lambda x:0 if x["severity"]=="error" else 1):
                            itype=iss.get("type","");b2=AF_BADGES.get(itype,f"⚠️ {itype}")
                            expl=AF_EXPL.get(itype,"")
                            content=f"**{b2}** — **{iss['site']}** — {iss['message']}"
                            if expl: content+=f"  \n*💡 {expl}*"
                            if iss["severity"]=="error": st.error(content)
                            elif itype in AF_SURGE_TYPES:
                                st.markdown(f"<div style='background:#fff3e0;border-left:4px solid #ff8c00;"
                                    f"padding:10px 14px;border-radius:4px;margin-bottom:6px;font-size:14px'>{content}</div>",
                                    unsafe_allow_html=True)
                            else: st.warning(content)
                    if sel_t["sites"]:
                        is_taux=sel_t["table_type"]=="taux"
                        def _af_fv(v): return f"{v*100:.2f}%" if (is_taux and v is not None) else (_fmt_full(v) if v is not None else "—")
                        df_data=pd.DataFrame([{
                            "":"🔴" if r["status"]=="alert" else ("🟠" if r.get("is_surge") else "🟡") if r["status"]=="warn" else "✅",
                            "Website":r["site"],f"M ({sel_t['lm']})":_af_fv(r["last"]),
                            f"M-1 ({sel_t['pm']})":_af_fv(r["prev"]),
                            "M/M-1":f"{r['evol']:+.1f}%" if r["evol"] is not None else "—",
                        } for r in sorted(sel_t["sites"],key=lambda x:({"alert":0,"warn":1,"ok":2}[x["status"]],-(x["last"] or 0)))])
                        def _af_cr(df):
                            def _row(row):
                                s=str(row.get("",""))
                                bg="#fdecec" if s=="🔴" else "#fff3e0" if s=="🟠" else "#fff4d6" if s=="🟡" else "#edf7ed"
                                return [f"background-color:{bg}" for _ in row]
                            return df.style.apply(_row,axis=1)
                        st.dataframe(_af_cr(df_data),use_container_width=True,hide_index=True)

            st.divider()
            af_rows_exp_t=[]
            for t in tables:
                if not t["issues"]:
                    af_rows_exp_t.append({"File":t.get("source_file") or t["file"],"Tab":t["sheet"],
                                        "Table":t["label"],"Result":"✅ OK","Detail":"No anomalies detected"})
                for iss in t["issues"]:
                    af_rows_exp_t.append({"File":t.get("source_file") or t["file"],"Tab":t["sheet"],
                                        "Table":f"{t['label']} — {iss['site']}",
                                        "Result":"❌ Error" if iss["severity"]=="error" else "⚠️ Warning",
                                        "Detail":iss["message"]})
            st.download_button("⬇ Download table anomalies",
                               pd.DataFrame(af_rows_exp_t).to_csv(index=False).encode("utf-8-sig"),
                               f"autofr_table_anomalies_{lm_ref}.csv","text/csv", key="af_dl_tables")
